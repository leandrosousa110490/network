import sys
import os
import json
import csv
import duckdb
import polars as pl
import gc
import weakref
try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.worksheet.dimensions import ColumnDimension
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    PARQUET_AVAILABLE = True
except ImportError:
    PARQUET_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import eel
    from eel_dashboard import create_dashboard
    EEL_AVAILABLE = True
except ImportError:
    EEL_AVAILABLE = False
    print("Eel not available. Install with: pip install eel")
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget,
    QPushButton, QTextEdit, QTableWidget, QTableWidgetItem, QComboBox,
    QLabel, QFileDialog, QMessageBox, QSplitter, QGroupBox, QTreeWidget,
    QTreeWidgetItem, QHeaderView, QDialog, QFormLayout, QLineEdit,
    QCheckBox, QSpinBox, QDialogButtonBox, QListWidget, QListWidgetItem,
    QMenu, QAction, QInputDialog, QRadioButton, QButtonGroup, QTabWidget,
    QAbstractItemView, QProgressBar, QCompleter, QProgressDialog, QScrollArea, QStyle
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QStringListModel, QRegExp
from PyQt5.QtGui import QFont, QTextCursor, QSyntaxHighlighter, QTextCharFormat, QColor, QTextDocument, QPixmap, QPainter

def build_mysql_connection_string(connection_data):
    """Build MySQL connection string for DuckDB MySQL extension"""
    params = []
    
    # Required parameters
    if connection_data.get('host'):
        params.append(f"host={connection_data['host']}")
    else:
        params.append("host=localhost")
    
    # Port (always include it)
    port = connection_data.get('port', 3306)
    params.append(f"port={port}")
    
    if connection_data.get('database'):
        params.append(f"database={connection_data['database']}")
    
    if connection_data.get('username'):
        params.append(f"user={connection_data['username']}")
    
    if connection_data.get('password'):
        params.append(f"password={connection_data['password']}")
    
    # SSL parameters - DuckDB MySQL extension format
    # Check if any SSL parameter is provided
    has_ssl = any([
        connection_data.get('ssl_ca'),
        connection_data.get('ssl_cert'),
        connection_data.get('ssl_key'),
        connection_data.get('ssl_mode')
    ])
    
    if has_ssl:
        # If SSL CA is provided, use it
        if connection_data.get('ssl_ca'):
            ssl_ca_path = connection_data['ssl_ca'].strip()
            if ssl_ca_path:
                params.append(f"ssl_ca={ssl_ca_path}")
        
        # SSL certificate
        if connection_data.get('ssl_cert'):
            ssl_cert_path = connection_data['ssl_cert'].strip()
            if ssl_cert_path:
                params.append(f"ssl_cert={ssl_cert_path}")
        
        # SSL key
        if connection_data.get('ssl_key'):
            ssl_key_path = connection_data['ssl_key'].strip()
            if ssl_key_path:
                params.append(f"ssl_key={ssl_key_path}")
        
        # SSL mode
        ssl_mode = connection_data.get('ssl_mode', 'required')
        if ssl_mode and ssl_mode != 'preferred':
            params.append(f"ssl_mode={ssl_mode}")
    
    connection_string = ' '.join(params)
    
    # Debug: print connection string (without password)
    debug_params = [p for p in params if not p.startswith('password=')]
    print(f"[DEBUG] MySQL Connection string (without password): {' '.join(debug_params)}")
    
    return connection_string

class SQLSyntaxHighlighter(QSyntaxHighlighter):
    """SQL syntax highlighter for the text editor"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_highlighting_rules()
    
    def setup_highlighting_rules(self):
        """Setup syntax highlighting rules for SQL"""
        self.highlighting_rules = []
        
        # SQL Keywords format
        keyword_format = QTextCharFormat()
        keyword_format.setForeground(QColor(0, 100, 200))  # Blue
        keyword_format.setFontWeight(QFont.Bold)
        
        # SQL keywords pattern
        sql_keywords = [
            'SELECT', 'FROM', 'WHERE', 'JOIN', 'INNER', 'LEFT', 'RIGHT', 'FULL',
            'GROUP', 'ORDER', 'BY', 'HAVING', 'LIMIT', 'OFFSET', 'DISTINCT', 'AS',
            'AND', 'OR', 'NOT', 'IN', 'EXISTS', 'BETWEEN', 'LIKE', 'IS', 'NULL',
            'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER', 'TABLE',
            'INDEX', 'VIEW', 'DATABASE', 'SCHEMA', 'PRIMARY', 'KEY', 'FOREIGN',
            'REFERENCES', 'CONSTRAINT', 'UNIQUE', 'CHECK', 'DEFAULT', 'AUTO_INCREMENT',
            'CASE', 'WHEN', 'THEN', 'ELSE', 'END', 'IF', 'UNION', 'ALL', 'INTERSECT',
            'EXCEPT', 'WITH', 'RECURSIVE', 'CTE', 'WINDOW', 'PARTITION', 'OVER',
            'ROW_NUMBER', 'RANK', 'DENSE_RANK', 'LAG', 'LEAD', 'FIRST_VALUE',
            'LAST_VALUE', 'COUNT', 'SUM', 'AVG', 'MIN', 'MAX', 'STDDEV', 'VARIANCE', 'USE', 'ON'
        ]
        
        for keyword in sql_keywords:
            pattern = QRegExp(r'\b' + keyword + r'\b', Qt.CaseInsensitive)
            self.highlighting_rules.append((pattern, keyword_format))
        
        # String literals format (single quotes)
        string_format = QTextCharFormat()
        string_format.setForeground(QColor(0, 150, 0))  # Green
        pattern = QRegExp(r"'[^']*'")
        self.highlighting_rules.append((pattern, string_format))
        
        # String literals format (double quotes)
        string_format2 = QTextCharFormat()
        string_format2.setForeground(QColor(0, 150, 0))  # Green
        pattern = QRegExp(r'"[^"]*"')
        self.highlighting_rules.append((pattern, string_format2))
        
        # Numbers format
        number_format = QTextCharFormat()
        number_format.setForeground(QColor(200, 100, 0))  # Orange
        pattern = QRegExp(r'\b\d+(\.\d+)?\b')
        self.highlighting_rules.append((pattern, number_format))
        
        # Comments format (-- style)
        comment_format = QTextCharFormat()
        comment_format.setForeground(QColor(128, 128, 128))  # Gray
        comment_format.setFontItalic(True)
        pattern = QRegExp(r'--[^\n]*')
        self.highlighting_rules.append((pattern, comment_format))
        
        # Comments format (/* */ style)
        multiline_comment_format = QTextCharFormat()
        multiline_comment_format.setForeground(QColor(128, 128, 128))  # Gray
        multiline_comment_format.setFontItalic(True)
        pattern = QRegExp(r'/\*.*\*/')
        pattern.setMinimal(True)
        self.highlighting_rules.append((pattern, multiline_comment_format))
        
        # Functions format
        function_format = QTextCharFormat()
        function_format.setForeground(QColor(150, 0, 150))  # Purple
        function_format.setFontWeight(QFont.Bold)
        pattern = QRegExp(r'\b[A-Za-z_][A-Za-z0-9_]*(?=\()')
        self.highlighting_rules.append((pattern, function_format))
    
    def highlightBlock(self, text):
        """Apply highlighting rules to a block of text"""
        for pattern, format_obj in self.highlighting_rules:
            expression = QRegExp(pattern)
            index = expression.indexIn(text)
            while index >= 0:
                length = expression.matchedLength()
                self.setFormat(index, length, format_obj)
                index = expression.indexIn(text, index + length)
        
        self.setCurrentBlockState(0)
    
    def add_custom_keywords(self, keywords):
        """Add custom keywords to the highlighting rules"""
        if not isinstance(keywords, (list, tuple)):
            return
        
        # Create format for custom keywords (same as SQL keywords)
        keyword_format = QTextCharFormat()
        keyword_format.setForeground(QColor(0, 100, 200))  # Blue
        keyword_format.setFontWeight(QFont.Bold)
        
        # Add new keyword patterns
        for keyword in keywords:
            pattern = QRegExp(r'\b' + str(keyword) + r'\b', Qt.CaseInsensitive)
            self.highlighting_rules.append((pattern, keyword_format))
        
        # Rehighlight the document
        self.rehighlight()

# SQL Auto-completion keywords and functions - configurable list
SQL_AUTOCOMPLETE_KEYWORDS = [
    # SQL Keywords
    'SELECT', 'FROM', 'WHERE', 'JOIN', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN',
    'GROUP BY', 'ORDER BY', 'HAVING', 'LIMIT', 'OFFSET', 'DISTINCT', 'AS', 'AND', 'OR', 'NOT',
    'IN', 'BETWEEN', 'LIKE', 'IS NULL', 'IS NOT NULL', 'EXISTS', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END',
    'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER', 'TABLE', 'INDEX', 'VIEW', 'DATABASE',
    'UNION', 'UNION ALL', 'INTERSECT', 'EXCEPT', 'WITH', 'RECURSIVE', 'CTE', 'USE', 'ON',
    
    # DuckDB specific functions
    'COUNT', 'SUM', 'AVG', 'MIN', 'MAX', 'STDDEV', 'VARIANCE',
    'CONCAT', 'SUBSTRING', 'LENGTH', 'UPPER', 'LOWER', 'TRIM', 'LTRIM', 'RTRIM',
    'CAST', 'COALESCE', 'NULLIF', 'GREATEST', 'LEAST',
    'DATE', 'TIME', 'TIMESTAMP', 'INTERVAL', 'EXTRACT', 'DATE_PART',
    'NOW', 'CURRENT_DATE', 'CURRENT_TIME', 'CURRENT_TIMESTAMP',
    'ROW_NUMBER', 'RANK', 'DENSE_RANK', 'LAG', 'LEAD', 'FIRST_VALUE', 'LAST_VALUE',
    'OVER', 'PARTITION BY', 'WINDOW', 'FRAME', 'ROWS', 'RANGE',
    
    # Data types
    'INTEGER', 'BIGINT', 'SMALLINT', 'TINYINT', 'DECIMAL', 'NUMERIC', 'REAL', 'DOUBLE',
    'VARCHAR', 'CHAR', 'TEXT', 'BLOB', 'DATE', 'TIME', 'TIMESTAMP', 'BOOLEAN',
    'ARRAY', 'STRUCT', 'MAP', 'JSON',
    
    # Common table references
    'localdb.', 'information_schema.', 'pg_catalog.',
    
    # Operators
    '=', '!=', '<>', '<', '>', '<=', '>=', '+', '-', '*', '/', '%', '||',
]

class SQLTextEdit(QTextEdit):
    """Custom QTextEdit with SQL auto-completion functionality and syntax highlighting"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent  # Store reference to parent application
        self.setup_autocomplete()
        self.setup_syntax_highlighting()
        self.setup_text_highlighting()
        
        # Connect selection change signal to highlight matching text
        self.selectionChanged.connect(self.on_selection_changed)
    
    def focusInEvent(self, event):
        """Override focus in event to track active editor"""
        super().focusInEvent(event)
        if self.parent_app and hasattr(self.parent_app, 'last_active_sql_editor'):
            self.parent_app.last_active_sql_editor = self
    
    def setup_syntax_highlighting(self):
        """Setup SQL syntax highlighting"""
        self.highlighter = SQLSyntaxHighlighter(self.document())
    
    def setup_text_highlighting(self):
        """Setup text highlighting for matching words"""
        self.highlight_format = QTextCharFormat()
        self.highlight_format.setBackground(QColor(255, 255, 0, 100))  # Light yellow background
        self.highlight_selections = []  # Store extra selections for highlighting
    
    def on_selection_changed(self):
        """Handle selection changes to highlight matching text"""
        # Clear previous highlights
        self.clear_highlights()
        
        # Get selected text
        cursor = self.textCursor()
        selected_text = cursor.selectedText().strip()
        
        # Only highlight if text is selected and is a word (not empty or whitespace)
        if selected_text and len(selected_text) > 1 and selected_text.replace('_', '').isalnum():
            self.highlight_matching_text(selected_text)
    
    def highlight_matching_text(self, text):
        """Highlight all occurrences of the given text"""
        if not text:
            return
        
        document = self.document()
        cursor = QTextCursor(document)
        
        # Find all occurrences of the text
        while True:
            cursor = document.find(text, cursor, QTextDocument.FindWholeWords)
            if cursor.isNull():
                break
            
            # Create extra selection for highlighting
            extra_selection = QTextEdit.ExtraSelection()
            extra_selection.cursor = cursor
            extra_selection.format = self.highlight_format
            self.highlight_selections.append(extra_selection)
        
        # Apply all highlights
        self.setExtraSelections(self.highlight_selections)
    
    def clear_highlights(self):
        """Clear all text highlights"""
        self.highlight_selections.clear()
        self.setExtraSelections([])
    
    def setup_autocomplete(self):
        """Setup auto-completion with SQL keywords"""
        # Create completer with SQL keywords
        self.completer = QCompleter(SQL_AUTOCOMPLETE_KEYWORDS)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.completer.setCompletionMode(QCompleter.PopupCompletion)
        self.completer.setFilterMode(Qt.MatchContains)
        
        # Style the completer popup
        popup = self.completer.popup()
        popup.setStyleSheet("""
            QListView {
                background-color: #f0f0f0;
                border: 2px solid #0064c8;
                selection-background-color: #0064c8;
                selection-color: white;
                font-weight: bold;
                color: #0064c8;
            }
            QListView::item {
                padding: 4px;
                border-bottom: 1px solid #e0e0e0;
            }
            QListView::item:selected {
                background-color: #0064c8;
                color: white;
            }
        """)
        
        # Connect completer to text edit
        self.completer.setWidget(self)
        self.completer.activated.connect(self.insert_completion)
    
    def insert_completion(self, completion):
        """Insert the selected completion into the text"""
        cursor = self.textCursor()
        
        # Get the current word being typed
        cursor.select(QTextCursor.WordUnderCursor)
        current_word = cursor.selectedText()
        
        # Replace current word with completion
        cursor.insertText(completion)
        self.setTextCursor(cursor)
        
        # Hide the completer popup after insertion
        self.completer.popup().hide()
    
    def keyPressEvent(self, event):
        """Handle key press events for auto-completion"""
        # Handle completer popup
        if self.completer.popup().isVisible():
            if event.key() in (Qt.Key_Enter, Qt.Key_Return, Qt.Key_Escape, Qt.Key_Tab, Qt.Key_Backtab):
                event.ignore()
                return
        
        # Call parent implementation first
        super().keyPressEvent(event)
        
        # Trigger completion on certain keys that should hide the popup
        if event.key() in (Qt.Key_Space, Qt.Key_Period, Qt.Key_Comma, Qt.Key_Semicolon):
            self.completer.popup().hide()
            return
        
        # Get current cursor and word
        cursor = self.textCursor()
        cursor.select(QTextCursor.WordUnderCursor)
        current_word = cursor.selectedText().strip()
        
        # Show completion if word is long enough and has potential matches
        if len(current_word) >= 2:
            # Set completion prefix
            self.completer.setCompletionPrefix(current_word)
            
            # Only show popup if there are matches
            if self.completer.completionCount() > 0:
                popup = self.completer.popup()
                popup.setCurrentIndex(self.completer.completionModel().index(0, 0))
                
                # Position popup at cursor
                cursor_rect = self.cursorRect()
                cursor_rect.setWidth(popup.sizeHintForColumn(0) + popup.verticalScrollBar().sizeHint().width())
                self.completer.complete(cursor_rect)
            else:
                # Hide popup if no matches found
                self.completer.popup().hide()
        else:
            # Hide popup if word is too short
            self.completer.popup().hide()
    
    def add_custom_completions(self, completions):
        """Add custom completions to the existing list and update syntax highlighting"""
        if isinstance(completions, (list, tuple)):
            current_completions = SQL_AUTOCOMPLETE_KEYWORDS.copy()
            current_completions.extend(completions)
            
            # Update completer model
            model = QStringListModel(current_completions)
            self.completer.setModel(model)
            
            # Update syntax highlighter with new keywords
            if hasattr(self, 'highlighter'):
                self.highlighter.add_custom_keywords(completions)
    
    def add_custom_keywords_to_highlighter(self, keywords):
        """Add custom keywords to syntax highlighting"""
        if hasattr(self, 'highlighter'):
            self.highlighter.add_custom_keywords(keywords)
    
    def set_completions(self, completions):
        """Replace all completions with a new list"""
        if isinstance(completions, (list, tuple)):
            model = QStringListModel(completions)
            self.completer.setModel(model)
    
    def insertFromMimeData(self, source):
        """Override to handle pasted text formatting - strip rich text and use plain text only"""
        if source.hasText():
            # Get plain text from clipboard, stripping all formatting
            plain_text = source.text()
            
            # Create a new QMimeData with only plain text
            from PyQt5.QtCore import QMimeData
            plain_mime_data = QMimeData()
            plain_mime_data.setText(plain_text)
            
            # Call parent with plain text only
            super().insertFromMimeData(plain_mime_data)
        else:
            # If no text, use default behavior
            super().insertFromMimeData(source)
    
    def contextMenuEvent(self, event):
        """Create custom context menu with 'Run Selected' option"""
        context_menu = self.createStandardContextMenu()
        
        # Add separator and 'Run Selected' option if text is selected
        if self.textCursor().hasSelection():
            context_menu.addSeparator()
            run_selected_action = context_menu.addAction("Run Selected")
            run_selected_action.triggered.connect(self.run_selected_text)
        
        context_menu.exec_(event.globalPos())
    
    def run_selected_text(self):
        """Signal to run the selected text"""
        try:
            # Get selected text first to check if there's anything selected
            selected_text = self.textCursor().selectedText().strip()
            if not selected_text:
                print("No text selected")
                return
            
            print(f"Selected text: {selected_text[:50]}...")  # Debug output
            
            # Try to find the parent with execute_selected_query method
            parent = self.parent()
            while parent:
                if hasattr(parent, 'execute_selected_query'):
                    print(f"Found execute_selected_query in {type(parent).__name__}")
                    parent.execute_selected_query()
                    return
                parent = parent.parent()
            
            print("Could not find execute_selected_query method in any parent")
        except Exception as e:
            print(f"Error running selected text: {e}")

class PDFViewer(QWidget):
    """PDF viewer widget with zoom and navigation controls"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.pdf_document = None
        self.current_page = 0
        self.zoom_factor = 1.0
        self.init_ui()
    
    def init_ui(self):
        """Initialize the PDF viewer UI"""
        layout = QVBoxLayout()
        
        # Control panel
        control_panel = QHBoxLayout()
        
        # Navigation buttons
        self.prev_btn = QPushButton("Previous")
        self.next_btn = QPushButton("Next")
        self.page_label = QLabel("Page: 0/0")
        
        # Zoom controls
        self.zoom_in_btn = QPushButton("Zoom In")
        self.zoom_out_btn = QPushButton("Zoom Out")
        self.zoom_fit_btn = QPushButton("Fit to Width")
        self.zoom_label = QLabel("100%")
        
        control_panel.addWidget(self.prev_btn)
        control_panel.addWidget(self.next_btn)
        control_panel.addWidget(self.page_label)
        control_panel.addStretch()
        control_panel.addWidget(self.zoom_out_btn)
        control_panel.addWidget(self.zoom_label)
        control_panel.addWidget(self.zoom_in_btn)
        control_panel.addWidget(self.zoom_fit_btn)
        
        layout.addLayout(control_panel)
        
        # PDF display area
        self.scroll_area = QScrollArea()
        self.pdf_label = QLabel()
        self.pdf_label.setAlignment(Qt.AlignCenter)
        self.pdf_label.setStyleSheet("background-color: white; border: 1px solid gray;")
        self.scroll_area.setWidget(self.pdf_label)
        self.scroll_area.setWidgetResizable(True)
        
        layout.addWidget(self.scroll_area)
        
        self.setLayout(layout)
        
        # Connect signals
        self.prev_btn.clicked.connect(self.previous_page)
        self.next_btn.clicked.connect(self.next_page)
        self.zoom_in_btn.clicked.connect(self.zoom_in)
        self.zoom_out_btn.clicked.connect(self.zoom_out)
        self.zoom_fit_btn.clicked.connect(self.fit_to_width)
        
        # Initially disable controls
        self.update_controls()
    
    def load_pdf(self, file_path):
        """Load a PDF file"""
        if not PDF_AVAILABLE:
            QMessageBox.warning(self, "PDF Support", 
                              "PDF support is not available. Please install PyMuPDF:\npip install PyMuPDF")
            return False
        
        try:
            self.pdf_document = fitz.open(file_path)
            self.current_page = 0
            self.zoom_factor = 1.0
            self.display_page()
            self.update_controls()
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load PDF: {str(e)}")
            return False
    
    def display_page(self):
        """Display the current page"""
        if not self.pdf_document or self.current_page >= len(self.pdf_document):
            return
        
        try:
            page = self.pdf_document[self.current_page]
            
            # Create transformation matrix for zoom
            mat = fitz.Matrix(self.zoom_factor, self.zoom_factor)
            
            # Render page to pixmap
            pix = page.get_pixmap(matrix=mat)
            
            # Convert to QPixmap
            img_data = pix.tobytes("ppm")
            qpixmap = QPixmap()
            qpixmap.loadFromData(img_data)
            
            # Display in label
            self.pdf_label.setPixmap(qpixmap)
            self.pdf_label.resize(qpixmap.size())
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to display page: {str(e)}")
    
    def previous_page(self):
        """Go to previous page"""
        if self.current_page > 0:
            self.current_page -= 1
            self.display_page()
            self.update_controls()
    
    def next_page(self):
        """Go to next page"""
        if self.pdf_document and self.current_page < len(self.pdf_document) - 1:
            self.current_page += 1
            self.display_page()
            self.update_controls()
    
    def zoom_in(self):
        """Zoom in"""
        self.zoom_factor = min(self.zoom_factor * 1.25, 5.0)
        self.display_page()
        self.update_zoom_label()
    
    def zoom_out(self):
        """Zoom out"""
        self.zoom_factor = max(self.zoom_factor / 1.25, 0.1)
        self.display_page()
        self.update_zoom_label()
    
    def fit_to_width(self):
        """Fit page to scroll area width"""
        if not self.pdf_document:
            return
        
        try:
            page = self.pdf_document[self.current_page]
            page_rect = page.rect
            
            # Calculate zoom factor to fit width
            available_width = self.scroll_area.viewport().width() - 20  # margin
            self.zoom_factor = available_width / page_rect.width
            self.zoom_factor = max(0.1, min(self.zoom_factor, 5.0))
            
            self.display_page()
            self.update_zoom_label()
        except Exception as e:
            print(f"Error fitting to width: {e}")
    
    def update_controls(self):
        """Update control button states and labels"""
        if self.pdf_document:
            total_pages = len(self.pdf_document)
            self.page_label.setText(f"Page: {self.current_page + 1}/{total_pages}")
            
            self.prev_btn.setEnabled(self.current_page > 0)
            self.next_btn.setEnabled(self.current_page < total_pages - 1)
            
            self.zoom_in_btn.setEnabled(True)
            self.zoom_out_btn.setEnabled(True)
            self.zoom_fit_btn.setEnabled(True)
        else:
            self.page_label.setText("Page: 0/0")
            self.prev_btn.setEnabled(False)
            self.next_btn.setEnabled(False)
            self.zoom_in_btn.setEnabled(False)
            self.zoom_out_btn.setEnabled(False)
            self.zoom_fit_btn.setEnabled(False)
    
    def update_zoom_label(self):
        """Update zoom percentage label"""
        self.zoom_label.setText(f"{int(self.zoom_factor * 100)}%")
    
    def close_pdf(self):
        """Close the current PDF"""
        if self.pdf_document:
            self.pdf_document.close()
            self.pdf_document = None
            self.pdf_label.clear()
            self.pdf_label.setText("No PDF loaded")
            self.update_controls()

class SavedQueryManagerDialog(QDialog):
    def __init__(self, parent=None, saved_queries=None):
        super().__init__(parent)
        self.parent_app = parent
        self.saved_queries = saved_queries or []
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle('Manage Saved Queries')
        self.setModal(True)
        self.resize(800, 600)
        
        layout = QVBoxLayout()
        
        # Query list
        self.query_list = QListWidget()
        self.populate_query_list()
        layout.addWidget(QLabel('Saved Queries:'))
        layout.addWidget(self.query_list)
        
        # Query details
        details_group = QGroupBox('Query Details')
        details_layout = QVBoxLayout()
        
        self.name_edit = QLineEdit()
        self.description_edit = QTextEdit()
        self.description_edit.setMaximumHeight(80)
        self.query_edit = QTextEdit()
        self.query_edit.setFont(QFont('Courier', 10))
        self.date_label = QLabel()
        
        details_layout.addWidget(QLabel('Name:'))
        details_layout.addWidget(self.name_edit)
        details_layout.addWidget(QLabel('Description:'))
        details_layout.addWidget(self.description_edit)
        details_layout.addWidget(QLabel('Query:'))
        details_layout.addWidget(self.query_edit)
        details_layout.addWidget(self.date_label)
        
        details_group.setLayout(details_layout)
        layout.addWidget(details_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.load_btn = QPushButton('Load Query')
        self.save_btn = QPushButton('Save Changes')
        self.delete_btn = QPushButton('Delete Query')
        self.close_btn = QPushButton('Close')
        
        button_layout.addWidget(self.load_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.close_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Connect signals
        self.query_list.currentRowChanged.connect(self.on_query_selected)
        self.load_btn.clicked.connect(self.load_selected_query)
        self.save_btn.clicked.connect(self.save_query_changes)
        self.delete_btn.clicked.connect(self.delete_selected_query)
        self.close_btn.clicked.connect(self.accept)
        
        # Initially disable buttons
        self.load_btn.setEnabled(False)
        self.save_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)
    
    def populate_query_list(self):
        self.query_list.clear()
        for query in self.saved_queries:
            item_text = f"{query['name']}"
            if query.get('description'):
                item_text += f" - {query['description'][:50]}..."
            self.query_list.addItem(item_text)
    
    def on_query_selected(self, row):
        if 0 <= row < len(self.saved_queries):
            query = self.saved_queries[row]
            self.name_edit.setText(query['name'])
            self.description_edit.setPlainText(query.get('description', ''))
            self.query_edit.setPlainText(query['query'])
            self.date_label.setText(f"Saved: {query.get('date_saved', 'Unknown')}")
            
            self.load_btn.setEnabled(True)
            self.save_btn.setEnabled(True)
            self.delete_btn.setEnabled(True)
        else:
            self.clear_details()
    
    def clear_details(self):
        self.name_edit.clear()
        self.description_edit.clear()
        self.query_edit.clear()
        self.date_label.clear()
        
        self.load_btn.setEnabled(False)
        self.save_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)
    
    def load_selected_query(self):
        row = self.query_list.currentRow()
        if 0 <= row < len(self.saved_queries):
            query = self.saved_queries[row]
            
            # Load into current tab of parent application
            current_tab_index = self.parent_app.query_tab_widget.currentIndex()
            if current_tab_index in self.parent_app.query_tabs:
                tab_data = self.parent_app.query_tabs[current_tab_index]
                tab_data['sql_editor'].setPlainText(query['query'])
                QMessageBox.information(self, 'Query Loaded', f'Query "{query["name"]}" has been loaded into the current tab.')
                self.accept()
            else:
                QMessageBox.warning(self, 'No Active Tab', 'No active query tab found in the main application.')
    
    def save_query_changes(self):
        row = self.query_list.currentRow()
        if 0 <= row < len(self.saved_queries):
            # Update the query data
            self.saved_queries[row]['name'] = self.name_edit.text().strip()
            self.saved_queries[row]['description'] = self.description_edit.toPlainText().strip()
            self.saved_queries[row]['query'] = self.query_edit.toPlainText().strip()
            
            # Save to file
            try:
                with open('saved_queries.json', 'w') as f:
                    json.dump(self.saved_queries, f, indent=2)
                QMessageBox.information(self, 'Changes Saved', 'Query changes have been saved successfully.')
                self.populate_query_list()
            except Exception as e:
                QMessageBox.critical(self, 'Save Error', f'Failed to save changes: {str(e)}')
    
    def delete_selected_query(self):
        row = self.query_list.currentRow()
        if 0 <= row < len(self.saved_queries):
            query_name = self.saved_queries[row]['name']
            reply = QMessageBox.question(self, 'Delete Query', 
                                       f'Are you sure you want to delete the query "{query_name}"?',
                                       QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                # Remove from list
                del self.saved_queries[row]
                
                # Save to file
                try:
                    with open('saved_queries.json', 'w') as f:
                        json.dump(self.saved_queries, f, indent=2)
                    QMessageBox.information(self, 'Query Deleted', f'Query "{query_name}" has been deleted.')
                    self.populate_query_list()
                    self.clear_details()
                except Exception as e:
                    QMessageBox.critical(self, 'Delete Error', f'Failed to delete query: {str(e)}')


class DatabaseConnectionDialog(QDialog):
    """Dialog for configuring database connections"""
    
    def __init__(self, parent=None, connection_data=None):
        super().__init__(parent)
        self.setWindowTitle('Database Connection Configuration')
        self.setModal(True)
        self.resize(400, 500)
        
        # Initialize form
        self.init_ui()
        
        # Initialize theme system
        self.current_theme = 'light'
        
        # Load existing connection data if provided
        if connection_data:
            self.load_connection_data(connection_data)
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Form layout for connection details
        form_layout = QFormLayout()
        
        # Connection name
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText('e.g., Production MySQL')
        form_layout.addRow('Connection Name:', self.name_edit)
        
        # Database type
        self.db_type_combo = QComboBox()
        self.db_type_combo.addItems(['mysql', 'mariadb'])
        form_layout.addRow('Database Type:', self.db_type_combo)
        
        # Host
        self.host_edit = QLineEdit()
        self.host_edit.setText('localhost')
        form_layout.addRow('Host:', self.host_edit)
        
        # Port
        self.port_spin = QSpinBox()
        self.port_spin.setRange(1, 65535)
        self.port_spin.setValue(3306)
        form_layout.addRow('Port:', self.port_spin)
        
        # Database name
        self.database_edit = QLineEdit()
        self.database_edit.setPlaceholderText('Database name')
        form_layout.addRow('Database:', self.database_edit)
        
        # Username
        self.username_edit = QLineEdit()
        form_layout.addRow('Username:', self.username_edit)
        
        # Password
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        form_layout.addRow('Password:', self.password_edit)
        
        layout.addLayout(form_layout)
        
        # SSL Configuration Group
        ssl_group = QGroupBox('SSL Configuration (Optional)')
        ssl_layout = QFormLayout(ssl_group)
        
        # SSL Mode
        self.ssl_mode_combo = QComboBox()
        self.ssl_mode_combo.addItems(['preferred', 'disabled', 'required', 'verify_ca', 'verify_identity'])
        ssl_layout.addRow('SSL Mode:', self.ssl_mode_combo)
        
        # SSL CA Certificate
        ssl_ca_layout = QHBoxLayout()
        self.ssl_ca_edit = QLineEdit()
        self.ssl_ca_edit.setPlaceholderText('Path to CA certificate file')
        self.ssl_ca_browse_btn = QPushButton('Browse')
        self.ssl_ca_browse_btn.clicked.connect(self.browse_ssl_ca)
        ssl_ca_layout.addWidget(self.ssl_ca_edit)
        ssl_ca_layout.addWidget(self.ssl_ca_browse_btn)
        ssl_layout.addRow('SSL CA Certificate:', ssl_ca_layout)
        
        # SSL Client Certificate
        ssl_cert_layout = QHBoxLayout()
        self.ssl_cert_edit = QLineEdit()
        self.ssl_cert_edit.setPlaceholderText('Path to client certificate file')
        self.ssl_cert_browse_btn = QPushButton('Browse')
        self.ssl_cert_browse_btn.clicked.connect(self.browse_ssl_cert)
        ssl_cert_layout.addWidget(self.ssl_cert_edit)
        ssl_cert_layout.addWidget(self.ssl_cert_browse_btn)
        ssl_layout.addRow('SSL Client Certificate:', ssl_cert_layout)
        
        # SSL Client Key
        ssl_key_layout = QHBoxLayout()
        self.ssl_key_edit = QLineEdit()
        self.ssl_key_edit.setPlaceholderText('Path to client private key file')
        self.ssl_key_browse_btn = QPushButton('Browse')
        self.ssl_key_browse_btn.clicked.connect(self.browse_ssl_key)
        ssl_key_layout.addWidget(self.ssl_key_edit)
        ssl_key_layout.addWidget(self.ssl_key_browse_btn)
        ssl_layout.addRow('SSL Client Key:', ssl_key_layout)
        
        layout.addWidget(ssl_group)
        
        # Test connection button
        self.test_btn = QPushButton('Test Connection')
        self.test_btn.clicked.connect(self.test_connection)
        layout.addWidget(self.test_btn)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def browse_ssl_ca(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Select SSL CA Certificate', '', 'Certificate Files (*.pem *.crt *.cer);;All Files (*)')
        if file_path:
            self.ssl_ca_edit.setText(file_path)
    
    def browse_ssl_cert(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Select SSL Client Certificate', '', 'Certificate Files (*.pem *.crt *.cer);;All Files (*)')
        if file_path:
            self.ssl_cert_edit.setText(file_path)
    
    def browse_ssl_key(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Select SSL Client Key', '', 'Key Files (*.pem *.key);;All Files (*)')
        if file_path:
            self.ssl_key_edit.setText(file_path)
    

        
        if self.username_edit.text():
            params.append(f"user={self.username_edit.text()}")
        
        if self.password_edit.text():
            params.append(f"password={self.password_edit.text()}")
        
        # SSL parameters
        if self.ssl_mode_combo.currentText() != 'preferred':
            params.append(f"ssl_mode={self.ssl_mode_combo.currentText()}")
        
        if self.ssl_ca_edit.text():
            params.append(f"ssl_ca={self.ssl_ca_edit.text()}")
        
        if self.ssl_cert_edit.text():
            params.append(f"ssl_cert={self.ssl_cert_edit.text()}")
        
        if self.ssl_key_edit.text():
            params.append(f"ssl_key={self.ssl_key_edit.text()}")
        
        return ' '.join(params)
    
    def get_connection_data(self):
        """Get all connection data as a dictionary"""
        return {
            'name': self.name_edit.text(),
            'type': self.db_type_combo.currentText(),
            'host': self.host_edit.text(),
            'port': self.port_spin.value(),
            'database': self.database_edit.text(),
            'username': self.username_edit.text(),
            'password': self.password_edit.text(),
            'ssl_mode': self.ssl_mode_combo.currentText(),
            'ssl_ca': self.ssl_ca_edit.text(),
            'ssl_cert': self.ssl_cert_edit.text(),
            'ssl_key': self.ssl_key_edit.text()
        }
    
    def load_connection_data(self, data):
        """Load connection data into the form"""
        self.name_edit.setText(data.get('name', ''))
        self.db_type_combo.setCurrentText(data.get('type', 'mysql'))
        self.host_edit.setText(data.get('host', 'localhost'))
        self.port_spin.setValue(data.get('port', 3306))
        self.database_edit.setText(data.get('database', ''))
        self.username_edit.setText(data.get('username', ''))
        self.password_edit.setText(data.get('password', ''))
        self.ssl_mode_combo.setCurrentText(data.get('ssl_mode', 'preferred'))
        self.ssl_ca_edit.setText(data.get('ssl_ca', ''))
        self.ssl_cert_edit.setText(data.get('ssl_cert', ''))
        self.ssl_key_edit.setText(data.get('ssl_key', ''))
    
    def test_connection(self):
        """Test the database connection"""
        connection_data = self.get_connection_data()
        
        # Validate required fields first
        if not connection_data['name'].strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a connection name.')
            return
        
        if not connection_data['host'].strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a host.')
            return
        
        if not connection_data['database'].strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a database name.')
            return
        
        if not connection_data['username'].strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a username.')
            return
        
        # Test the connection
        try:
            import duckdb
            test_conn = duckdb.connect(':memory:')
            
            # Install and load MySQL extension
            try:
                test_conn.execute('INSTALL mysql')
                test_conn.execute('LOAD mysql')
            except Exception as ext_error:
                # Extension might already be installed
                try:
                    test_conn.execute('LOAD mysql')
                except:
                    raise Exception(f"Failed to load MySQL extension. Please install it first.\nError: {str(ext_error)}")
            
            # Build connection string
            connection_string = self.build_connection_string(connection_data)
            
            # Try to attach the database
            try:
                attach_query = f"ATTACH '{connection_string}' AS test_db (TYPE mysql, READ_ONLY)"
                test_conn.execute(attach_query)
            except Exception as attach_error:
                error_msg = str(attach_error).lower()
                
                if 'io error' in error_msg or 'failed to connect' in error_msg:
                    raise Exception(
                        f"Failed to connect to MySQL database.\n\n"
                        f"Connection details:\n"
                        f"  Host: {connection_data.get('host', 'N/A')}\n"
                        f"  Port: {connection_data.get('port', 3306)}\n"
                        f"  Database: {connection_data.get('database', 'N/A')}\n"
                        f"  User: {connection_data.get('username', 'N/A')}\n\n"
                        f"Please verify:\n"
                        f"  1. MySQL server is running and accessible\n"
                        f"  2. Host and port are correct\n"
                        f"  3. Username and password are correct\n"
                        f"  4. Database name exists\n"
                        f"  5. User has permission to access the database\n"
                        f"  6. Firewall allows connection to MySQL port\n\n"
                        f"Original error: {str(attach_error)}"
                    )
                else:
                    raise attach_error
            
            # Try to list tables to verify connection
            test_conn.execute("SELECT table_name FROM information_schema.tables WHERE table_catalog = 'test_db' LIMIT 1")
            
            # Clean up
            test_conn.execute('DETACH test_db')
            test_conn.close()
            
            QMessageBox.information(self, 'Connection Test', 'Connection successful!\n\nYou can now save and use this connection.')
            
        except Exception as e:
            error_message = str(e)
            QMessageBox.critical(self, 'Connection Test Failed', f'Failed to connect to database:\n\n{error_message}')
    
    def build_connection_string(self, connection_data):
        """Build connection string from connection data"""
        return build_mysql_connection_string(connection_data)
    
    def accept(self):
        """Validate form before accepting"""
        if not self.name_edit.text().strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a connection name.')
            return
        
        if not self.host_edit.text().strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a host.')
            return
        
        if not self.database_edit.text().strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a database name.')
            return
        
        if not self.username_edit.text().strip():
            QMessageBox.warning(self, 'Validation Error', 'Please enter a username.')
            return
        
        super().accept()
    
    def get_selected_mode(self):
        """Return the selected display mode"""
        return 'webengine' if self.webengine_radio.isChecked() else 'browser'


class ConnectionManagerDialog(QDialog):
    """Dialog for managing saved database connections"""
    
    def __init__(self, parent=None, connections=None):
        super().__init__(parent)
        self.setWindowTitle('Manage Database Connections')
        self.setModal(True)
        self.resize(600, 400)
        
        self.connections = connections or []
        self.init_ui()
        self.refresh_connection_list()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Connection list
        list_layout = QHBoxLayout()
        
        self.connection_list = QListWidget()
        self.connection_list.itemSelectionChanged.connect(self.on_selection_changed)
        list_layout.addWidget(self.connection_list)
        
        # Buttons
        button_layout = QVBoxLayout()
        
        self.new_btn = QPushButton('New Connection')
        self.new_btn.clicked.connect(self.new_connection)
        button_layout.addWidget(self.new_btn)
        
        self.edit_btn = QPushButton('Edit Connection')
        self.edit_btn.clicked.connect(self.edit_connection)
        self.edit_btn.setEnabled(False)
        button_layout.addWidget(self.edit_btn)
        
        self.connect_btn = QPushButton('Connect')
        self.connect_btn.clicked.connect(self.connect_to_selected)
        self.connect_btn.setEnabled(False)
        button_layout.addWidget(self.connect_btn)
        
        self.delete_btn = QPushButton('Delete')
        self.delete_btn.clicked.connect(self.delete_connection)
        self.delete_btn.setEnabled(False)
        button_layout.addWidget(self.delete_btn)
        
        button_layout.addStretch()
        
        list_layout.addLayout(button_layout)
        layout.addLayout(list_layout)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Close)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def refresh_connection_list(self):
        """Refresh the connection list display"""
        self.connection_list.clear()
        
        for connection in self.connections:
            item = QListWidgetItem()
            item.setText(f"{connection['name']} ({connection['type']}@{connection['host']}:{connection['port']})")
            item.setData(Qt.UserRole, connection)
            self.connection_list.addItem(item)
    
    def on_selection_changed(self):
        """Handle selection change in connection list"""
        has_selection = bool(self.connection_list.currentItem())
        self.edit_btn.setEnabled(has_selection)
        self.connect_btn.setEnabled(has_selection)
        self.delete_btn.setEnabled(has_selection)
    
    def new_connection(self):
        """Create a new connection"""
        dialog = DatabaseConnectionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            connection_data = dialog.get_connection_data()
            
            # Check if connection name already exists
            existing_names = [conn['name'] for conn in self.connections]
            if connection_data['name'] in existing_names:
                QMessageBox.warning(self, 'Duplicate Name', 'A connection with this name already exists.')
                return
            
            self.connections.append(connection_data)
            self.refresh_connection_list()
    
    def edit_connection(self):
        """Edit the selected connection"""
        current_item = self.connection_list.currentItem()
        if not current_item:
            return
        
        connection_data = current_item.data(Qt.UserRole)
        dialog = DatabaseConnectionDialog(self, connection_data)
        
        if dialog.exec_() == QDialog.Accepted:
            updated_data = dialog.get_connection_data()
            
            # Find and update the connection
            for i, conn in enumerate(self.connections):
                if conn['name'] == connection_data['name']:
                    self.connections[i] = updated_data
                    break
            
            self.refresh_connection_list()
    
    def connect_to_selected(self):
        """Connect to the selected database"""
        current_item = self.connection_list.currentItem()
        if not current_item:
            return
        
        connection_data = current_item.data(Qt.UserRole)
        
        # Close this dialog and connect
        self.accept()
        
        # Get parent window and connect
        if hasattr(self.parent(), 'connect_to_database'):
            self.parent().connect_to_database(connection_data)
    
    def delete_connection(self):
        """Delete the selected connection"""
        current_item = self.connection_list.currentItem()
        if not current_item:
            return
        
        connection_data = current_item.data(Qt.UserRole)
        
        reply = QMessageBox.question(self, 'Delete Connection', 
                                   f'Are you sure you want to delete the connection "{connection_data["name"]}"?',
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.connections = [conn for conn in self.connections if conn['name'] != connection_data['name']]
            self.refresh_connection_list()
    
    def get_connections(self):
        """Get the current list of connections"""
        return self.connections

class DelimiterSelectionDialog(QDialog):
    """Dialog for selecting CSV delimiter with auto-detection"""
    
    def __init__(self, parent=None, file_path=None):
        super().__init__(parent)
        self.setWindowTitle('CSV Delimiter Selection')
        self.setModal(True)
        self.resize(400, 300)
        
        self.file_path = file_path
        self.detected_delimiter = None
        self.selected_delimiter = None
        
        self.init_ui()
        
        if file_path:
            self.detect_delimiter()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # File info
        if self.file_path:
            file_label = QLabel(f'File: {os.path.basename(self.file_path)}')
            layout.addWidget(file_label)
        
        # Detection result
        self.detection_label = QLabel('Detecting delimiter...')
        layout.addWidget(self.detection_label)
        
        layout.addWidget(QLabel('\nSelect delimiter:'))
        
        # Delimiter options
        self.delimiter_group = QButtonGroup()
        
        # Common delimiters
        delimiters = [
            (',', 'Comma (,)'),
            (';', 'Semicolon (;)'),
            ('\t', 'Tab'),
            ('|', 'Pipe (|)'),
            (' ', 'Space'),
        ]
        
        self.delimiter_radios = {}
        for delimiter, label in delimiters:
            radio = QRadioButton(label)
            radio.delimiter = delimiter
            self.delimiter_group.addButton(radio)
            self.delimiter_radios[delimiter] = radio
            layout.addWidget(radio)
        
        # Custom delimiter
        custom_layout = QHBoxLayout()
        self.custom_radio = QRadioButton('Custom:')
        self.custom_input = QLineEdit()
        self.custom_input.setMaxLength(1)
        self.custom_input.setFixedWidth(50)
        self.custom_input.textChanged.connect(self.on_custom_delimiter_changed)
        
        custom_layout.addWidget(self.custom_radio)
        custom_layout.addWidget(self.custom_input)
        custom_layout.addStretch()
        
        self.delimiter_group.addButton(self.custom_radio)
        layout.addLayout(custom_layout)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def detect_delimiter(self):
        """Detect the most likely delimiter in the CSV file"""
        try:
            with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                # Read first few lines for detection
                sample = ''
                for i, line in enumerate(f):
                    if i >= 10:  # Use first 10 lines for detection
                        break
                    sample += line
            
            # Use csv.Sniffer to detect delimiter
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample, delimiters=',;\t| ')
                self.detected_delimiter = dialect.delimiter
            except csv.Error:
                # Fallback: count occurrences of common delimiters
                delimiters = [',', ';', '\t', '|', ' ']
                counts = {}
                
                for delimiter in delimiters:
                    counts[delimiter] = sample.count(delimiter)
                
                # Choose delimiter with highest count
                if counts:
                    self.detected_delimiter = max(counts, key=counts.get)
                else:
                    self.detected_delimiter = ','
            
            # Update UI
            self.detection_label.setText(f'Detected delimiter: {self.get_delimiter_display_name(self.detected_delimiter)}')
            
            # Select the detected delimiter
            if self.detected_delimiter in self.delimiter_radios:
                self.delimiter_radios[self.detected_delimiter].setChecked(True)
            else:
                self.custom_radio.setChecked(True)
                self.custom_input.setText(self.detected_delimiter)
            
            
        except Exception as e:
            self.detection_label.setText(f'Detection failed: {str(e)}')
            # Default to comma
            self.delimiter_radios[','].setChecked(True)
    
    def get_delimiter_display_name(self, delimiter):
        """Get display name for delimiter"""
        names = {
            ',': 'Comma (,)',
            ';': 'Semicolon (;)',
            '\t': 'Tab',
            '|': 'Pipe (|)',
            ' ': 'Space'
        }
        return names.get(delimiter, f'Custom ({delimiter})')
    
    def on_custom_delimiter_changed(self):
        """Handle custom delimiter input change"""
        if self.custom_input.text():
            self.custom_radio.setChecked(True)
    
    def get_selected_delimiter(self):
        """Get the currently selected delimiter"""
        for radio in self.delimiter_group.buttons():
            if radio.isChecked():
                if radio == self.custom_radio:
                    return self.custom_input.text() or ','
                else:
                    return radio.delimiter
        return ','
    

    
    def accept(self):
        """Accept dialog and store selected delimiter"""
        self.selected_delimiter = self.get_selected_delimiter()
        super().accept()
    
    def get_delimiter(self):
        """Get the selected delimiter"""
        return self.selected_delimiter


class ExcelSheetSelectionDialog(QDialog):
    """Dialog for selecting Excel sheet from available sheets"""
    
    def __init__(self, parent=None, file_path=None):
        super().__init__(parent)
        self.setWindowTitle('Excel Sheet Selection')
        self.setModal(True)
        self.resize(400, 200)
        
        self.file_path = file_path
        self.selected_sheet = None
        self.available_sheets = []
        
        self.init_ui()
        
        if file_path:
            self.load_sheet_names()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # File info
        if self.file_path:
            file_label = QLabel(f'File: {os.path.basename(self.file_path)}')
            layout.addWidget(file_label)
        
        layout.addWidget(QLabel('\nSelect sheet:'))
        
        # Sheet selection dropdown
        self.sheet_combo = QComboBox()
        layout.addWidget(self.sheet_combo)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def load_sheet_names(self):
        """Load available sheet names from Excel file without loading data"""
        try:
            # Use openpyxl to get sheet names only (much faster)
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path, read_only=True, data_only=False)
            self.available_sheets = wb.sheetnames
            wb.close()
            
        except Exception as e:
            # Fallback: try with pandas ExcelFile (also fast for sheet names)
            try:
                import pandas as pd
                with pd.ExcelFile(self.file_path) as xls:
                    self.available_sheets = xls.sheet_names
            except:
                # Last fallback: use polars to get sheet names
                try:
                    import polars as pl
                    xl_file = pl.read_excel(self.file_path, sheet_id=None)
                    if isinstance(xl_file, dict):
                        self.available_sheets = list(xl_file.keys())
                    else:
                        self.available_sheets = ['Sheet1']
                except:
                    self.available_sheets = ['Sheet1']  # Final fallback
        
        # Populate combo box
        self.sheet_combo.addItems(self.available_sheets)
        
        # Set first sheet as default
        if self.available_sheets:
            self.sheet_combo.setCurrentIndex(0)
    
    def get_selected_sheet(self):
        """Get the currently selected sheet name"""
        return self.sheet_combo.currentText()
    
    def accept(self):
        """Accept dialog and store selected sheet"""
        self.selected_sheet = self.get_selected_sheet()
        super().accept()
    
    def get_sheet_name(self):
        """Get the selected sheet name"""
        return self.selected_sheet


class ExportDelimiterDialog(QDialog):
    """Dialog for selecting CSV export delimiter"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('CSV Export Options')
        self.setModal(True)
        self.selected_delimiter = ','
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Title
        title_label = QLabel('Select CSV delimiter:')
        title_label.setStyleSheet('font-weight: bold; font-size: 12px;')
        layout.addWidget(title_label)
        
        # Delimiter options
        self.delimiter_group = QButtonGroup()
        
        # Common delimiters
        delimiters = [
            (',', 'Comma (,)'),
            (';', 'Semicolon (;)'),
            ('\t', 'Tab'),
            ('|', 'Pipe (|)'),
        ]
        
        for delimiter, display_name in delimiters:
            radio = QRadioButton(display_name)
            if delimiter == ',':
                radio.setChecked(True)  # Default to comma
            radio.toggled.connect(lambda checked, d=delimiter: self.set_delimiter(d) if checked else None)
            self.delimiter_group.addButton(radio)
            layout.addWidget(radio)
        
        # Custom delimiter option
        custom_layout = QHBoxLayout()
        self.custom_radio = QRadioButton('Custom:')
        self.custom_input = QLineEdit()
        self.custom_input.setMaxLength(1)
        self.custom_input.setFixedWidth(50)
        self.custom_input.textChanged.connect(self.on_custom_delimiter_changed)
        
        custom_layout.addWidget(self.custom_radio)
        custom_layout.addWidget(self.custom_input)
        custom_layout.addStretch()
        
        self.delimiter_group.addButton(self.custom_radio)
        layout.addLayout(custom_layout)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def set_delimiter(self, delimiter):
        self.selected_delimiter = delimiter
    
    def on_custom_delimiter_changed(self):
        if self.custom_input.text():
            self.custom_radio.setChecked(True)
            self.selected_delimiter = self.custom_input.text()
    
    def get_delimiter(self):
        return self.selected_delimiter


class SQLQueryThread(QThread):
    """Thread for executing SQL queries to prevent UI freezing"""
    result_ready = pyqtSignal(list, list)  # columns, data
    error_occurred = pyqtSignal(str)
    
    def __init__(self, connection, query):
        super().__init__()
        self.connection = connection
        self.query = query
    
    def run(self):
        try:
            result = self.connection.execute(self.query).fetchall()
            columns = [desc[0] for desc in self.connection.description]
            self.result_ready.emit(columns, result)
        except Exception as e:
            self.error_occurred.emit(str(e))

class StreamingQueryThread(QThread):
    """Thread for executing streaming SQL queries with pagination"""
    batch_ready = pyqtSignal(list, list, int, bool)  # columns, data, total_count, has_more
    error_occurred = pyqtSignal(str)
    progress_update = pyqtSignal(int)  # rows processed
    
    def __init__(self, connection, query, batch_size=10000, offset=0):
        super().__init__()
        self.connection = connection
        self.query = query
        self.batch_size = batch_size
        self.offset = offset
        self._is_cancelled = False
    
    def cancel(self):
        """Cancel the query execution"""
        self._is_cancelled = True
    
    def run(self):
        try:
            # Emit initial progress
            self.progress_update.emit(0)
            
            # Check if this is a SELECT query that supports pagination
            query_upper = self.query.strip().upper()
            is_select_query = query_upper.startswith('SELECT')
            
            if is_select_query:
                # Handle SELECT queries with pagination
                total_count = 0
                
                try:
                    # Create a count query by wrapping the original query
                    # Strip trailing semicolon to avoid syntax errors in subquery
                    clean_query = self.query.rstrip().rstrip(';')
                    count_query = f"SELECT COUNT(*) FROM ({clean_query}) AS count_subquery"
                    count_result = self.connection.execute(count_query).fetchone()
                    total_count = count_result[0] if count_result else 0
                    self.progress_update.emit(25)  # 25% progress after count
                except:
                    # If count fails, we'll stream without knowing total
                    total_count = -1
                    self.progress_update.emit(25)
                
                if self._is_cancelled:
                    return
                
                # Handle pagination based on whether query already has LIMIT
                import re
                # More precise LIMIT detection using regex to avoid false positives
                limit_pattern = r'\bLIMIT\s+\d+\b'
                has_limit_clause = bool(re.search(limit_pattern, query_upper, re.IGNORECASE))
                
                if has_limit_clause:
                    # Query already has LIMIT, always wrap it to handle pagination properly
                    # This ensures consistent behavior regardless of the original LIMIT value
                    # Strip trailing semicolon to avoid syntax errors in subquery
                    clean_query = self.query.rstrip().rstrip(';')
                    paginated_query = f"SELECT * FROM ({clean_query}) AS subquery LIMIT {self.batch_size} OFFSET {self.offset}"
                else:
                    # Add LIMIT and OFFSET to the original query for pagination
                    # Strip trailing semicolon to avoid syntax errors
                    clean_query = self.query.rstrip().rstrip(';')
                    paginated_query = f"{clean_query} LIMIT {self.batch_size} OFFSET {self.offset}"
                self.progress_update.emit(50)  # 50% progress after query preparation
                
                if self._is_cancelled:
                    return
                
                # Execute the paginated query
                cursor = self.connection.execute(paginated_query)
                columns = [desc[0] for desc in cursor.description]
                self.progress_update.emit(75)  # 75% progress after query execution
            else:
                # Handle non-SELECT queries (ALTER, CREATE, INSERT, UPDATE, DELETE, etc.)
                # These don't support LIMIT/OFFSET and should be executed directly
                self.progress_update.emit(25)
                
                if self._is_cancelled:
                    return
                
                # Execute the query directly without pagination
                clean_query = self.query.rstrip().rstrip(';')
                cursor = self.connection.execute(clean_query)
                
                # For non-SELECT queries, we may not have columns or data to return
                if cursor.description:
                    columns = [desc[0] for desc in cursor.description]
                else:
                    # For DDL/DML operations, create a simple result indicator
                    columns = ['Result']
                
                total_count = 1  # Indicate successful execution
                self.progress_update.emit(75)  # 75% progress after query execution
            
            if self._is_cancelled:
                return
            
            # Fetch the batch with memory optimization
            batch_data = []
            row_count = 0
            
            if is_select_query:
                # Process rows in smaller chunks to reduce memory usage for SELECT queries
                while row_count < self.batch_size:
                    chunk = cursor.fetchmany(1000)  # Fetch in smaller chunks
                    if not chunk:
                        break
                        
                    if self._is_cancelled:
                        return
                        
                    # Process chunk and add to batch_data
                    for row in chunk:
                        # Convert large objects to strings early to save memory
                        processed_row = []
                        for cell in row:
                            if isinstance(cell, (bytes, bytearray)) and len(cell) > 10000:
                                # Convert large binary data to summary
                                processed_row.append(f"<Binary data: {len(cell)} bytes>")
                            elif isinstance(cell, str) and len(cell) > 50000:
                                # Truncate very large strings
                                processed_row.append(cell[:50000] + "... (truncated)")
                            else:
                                processed_row.append(cell)
                        batch_data.append(tuple(processed_row))
                        row_count += 1
                        
                        if row_count >= self.batch_size:
                            break
                    
                    # Update progress based on rows processed
                    progress = min(95, 75 + (row_count / self.batch_size) * 20)
                    self.progress_update.emit(int(progress))
            else:
                # For non-SELECT queries, try to fetch any results or create success message
                try:
                    results = cursor.fetchall()
                    if results:
                        # Some non-SELECT queries return data (e.g., INSERT...RETURNING)
                        for row in results:
                            processed_row = []
                            for cell in row:
                                if isinstance(cell, (bytes, bytearray)) and len(cell) > 10000:
                                    processed_row.append(f"<Binary data: {len(cell)} bytes>")
                                elif isinstance(cell, str) and len(cell) > 50000:
                                    processed_row.append(cell[:50000] + "... (truncated)")
                                else:
                                    processed_row.append(cell)
                            batch_data.append(tuple(processed_row))
                            row_count += 1
                    else:
                        # No results returned, create success message
                        batch_data.append(("Query executed successfully",))
                        row_count = 1
                except:
                    # Some queries don't return fetchable results
                    batch_data.append(("Query executed successfully",))
                    row_count = 1
            
            if self._is_cancelled:
                return
            
            # Determine if there are more results (only relevant for SELECT queries)
            if is_select_query:
                has_more = len(batch_data) == self.batch_size and (total_count == -1 or self.offset + self.batch_size < total_count)
            else:
                has_more = False  # Non-SELECT queries don't have pagination
            
            self.progress_update.emit(100)  # 100% progress when complete
            self.batch_ready.emit(columns, batch_data, total_count, has_more)
            
            # Clean up batch_data reference to free memory
            del batch_data
            gc.collect()
            
        except Exception as e:
            if not self._is_cancelled:
                self.error_occurred.emit(str(e))

class FullExportQueryThread(QThread):
    """Thread for executing complete SQL queries for export purposes"""
    export_ready = pyqtSignal(list, list)  # columns, data
    error_occurred = pyqtSignal(str)
    progress_update = pyqtSignal(int)  # progress percentage
    
    def __init__(self, connection, query):
        super().__init__()
        self.connection = connection
        self.query = query
        self._is_cancelled = False
    
    def cancel(self):
        """Cancel the export operation"""
        self._is_cancelled = True
    
    def run(self):
        try:
            # Emit initial progress
            self.progress_update.emit(10)
            
            if self._is_cancelled:
                return
            
            # Clean the query by removing trailing semicolons
            clean_query = self.query.rstrip().rstrip(';')
            
            # Execute the complete query without pagination
            self.progress_update.emit(30)
            cursor = self.connection.execute(clean_query)
            columns = [desc[0] for desc in cursor.description]
            
            if self._is_cancelled:
                return
            
            self.progress_update.emit(50)
            
            # Fetch all data in batches to manage memory
            all_data = []
            batch_size = 10000
            
            while True:
                if self._is_cancelled:
                    return
                    
                batch = cursor.fetchmany(batch_size)
                if not batch:
                    break
                    
                all_data.extend(batch)
                self.progress_update.emit(min(90, 50 + (len(all_data) // batch_size) * 5))
            
            if self._is_cancelled:
                return
                
            self.progress_update.emit(100)
            self.export_ready.emit(columns, all_data)
            
        except Exception as e:
            if not self._is_cancelled:
                self.error_occurred.emit(str(e))
        finally:
            # Clean up
            gc.collect()

class DuckDBSQLApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.connection = duckdb.connect(':memory:')
        # Create the localdb schema
        self.connection.execute("CREATE SCHEMA IF NOT EXISTS localdb")
        self.loaded_tables = {}  # filename -> table_name mapping
        
        # Connection storage file
        self.connections_file = 'db_connections.json'
        self.saved_connections = self.load_saved_connections()
        self.active_connections = {}  # Store multiple active connections: {db_name: connection_data}
        
        # Query tab management
        self.query_tabs = {}  # tab_index -> {sql_editor, results_table, query_thread}
        self.tab_counter = 0
        
        # Split screen management
        self.split_screen_active = False
        self.split_screen_widget = None
        self.split_query_tabs = {}  # Track split screen tabs
        self.last_active_sql_editor = None  # Track the last active SQL editor
        
        self.init_ui()
        
        # Initialize theme system
        self.current_theme = 'light'
        self.load_theme_preference()
        

        
    def init_ui(self):
        self.setWindowTitle('DuckDB SQL Query Application')
        self.setGeometry(100, 100, 1200, 800)
        
        # Create menu bar
        self.create_menu_bar()
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout with no margins to use full space
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        
        # Create horizontal splitter for main content and schema tree
        main_splitter = QSplitter(Qt.Horizontal)
        
        # Left side container for tabbed queries
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        
        # Create tab widget for multiple queries
        self.query_tab_widget = QTabWidget()
        self.query_tab_widget.setTabsClosable(False)  # We'll use custom close buttons
        
        # Add initial query tab
        self.add_new_query_tab()
        
        left_layout.addWidget(self.query_tab_widget)
        
        # Database schema tree view (left side)
        schema_group = QGroupBox('Database Schema')
        schema_layout = QVBoxLayout(schema_group)
        
        self.schema_tree = QTreeWidget()
        self.schema_tree.setHeaderLabels(['Schema/Table', 'Type', 'Rows'])
        self.schema_tree.header().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.schema_tree.itemDoubleClicked.connect(self.on_table_double_click)
        self.schema_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.schema_tree.customContextMenuRequested.connect(self.show_schema_context_menu)
        self.schema_tree.itemExpanded.connect(self.on_schema_item_expanded)
        self.refresh_schema_tree()
        
        schema_layout.addWidget(self.schema_tree)
        
        # Add schema group to main splitter (left side)
        main_splitter.addWidget(schema_group)
        
        # Add right widget to main splitter
        main_splitter.addWidget(left_widget)
        
        # Set splitter proportions (25% left, 75% right) and ensure full UI usage
        main_splitter.setSizes([250, 750])
        main_splitter.setStretchFactor(0, 1)
        main_splitter.setStretchFactor(1, 3)
        
        # Ensure the splitter takes all available space
        from PyQt5.QtWidgets import QSizePolicy
        main_splitter.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        main_layout.addWidget(main_splitter)
        
        # Status info
        self.status_label = QLabel('Ready. Load a file to begin.')
        main_layout.addWidget(self.status_label)
        
    def get_supported_formats(self):
        """Return file formats that DuckDB can load directly"""
        formats = {
            'CSV Files': '*.csv',
            'Parquet Files': '*.parquet',
            'JSON Files': '*.json',
            'JSONL Files': '*.jsonl',
            'TSV Files': '*.tsv',
            'Excel Files': '*.xlsx *.xls',
            'All Supported': '*.csv *.parquet *.json *.jsonl *.tsv *.xlsx *.xls'
        }
        
        # Add PDF support if available
        if PDF_AVAILABLE:
            formats['PDF Files'] = '*.pdf'
            formats['All Supported'] += ' *.pdf'
            
        return formats
    
    def load_file(self):
        """Load a file into DuckDB"""
        formats = self.get_supported_formats()
        filter_str = ';;'.join([f'{name} ({ext})' for name, ext in formats.items()])
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, 'Select Data File', '', filter_str
        )
        
        if not file_path:
            return
            
        try:
            # Extract filename without extension for default table name
            filename = os.path.basename(file_path)
            default_table_name = os.path.splitext(filename)[0]
            
            # Clean default table name (remove special characters)
            default_table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in default_table_name)
            
            # Prompt user for table name
            table_name, ok = QInputDialog.getText(
                self, 'Table Name', 
                f'Enter a name for the table (from file: {filename}):',
                text=default_table_name
            )
            
            if not ok or not table_name.strip():
                return  # User cancelled or entered empty name
            
            # Clean the user-provided table name
            table_name = table_name.strip()
            table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
            
            # Determine file type and load accordingly
            file_ext = os.path.splitext(file_path)[1].lower()
            delimiter_info = ""  # For success message
            
            if file_ext == '.csv':
                # Show delimiter selection dialog
                dialog = DelimiterSelectionDialog(self, file_path)
                if dialog.exec_() == QDialog.Accepted:
                    delimiter = dialog.get_delimiter()
                    
                    # Store delimiter info for success message
                    delimiter_names = {
                        ',': 'comma',
                        ';': 'semicolon', 
                        '\t': 'tab',
                        '|': 'pipe',
                        ' ': 'space'
                    }
                    delimiter_info = f" (delimiter: {delimiter_names.get(delimiter, delimiter)})"
                    
                    # Escape the delimiter for SQL
                    escaped_delimiter = delimiter.replace("'", "''")
                    
                    # Use DuckDB's read_csv with specified delimiter
                    query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_csv('{file_path}', delim='{escaped_delimiter}')"
                else:
                    # User cancelled the dialog
                    return
            elif file_ext == '.parquet':
                query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_parquet('{file_path}')"
            elif file_ext in ['.json', '.jsonl']:
                query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_json_auto('{file_path}')"
            elif file_ext == '.tsv':
                query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_csv_auto('{file_path}', delim='\t')"
                delimiter_info = " (delimiter: tab)"
            elif file_ext in ['.xlsx', '.xls']:
                # Handle Excel files using Polars with sheet selection dialog
                dialog = ExcelSheetSelectionDialog(self, file_path)
                if dialog.exec_() == QDialog.Accepted:
                    sheet_name = dialog.get_sheet_name()
                else:
                    return  # User cancelled
                
                # Use Polars to read Excel file
                try:
                    if sheet_name:
                        # Load specific sheet
                        df = pl.read_excel(file_path, sheet_name=sheet_name)
                        delimiter_info = f" (sheet: {sheet_name})"
                    else:
                        # Load first sheet (default)
                        df = pl.read_excel(file_path)
                        delimiter_info = " (sheet: first sheet)"
                    
                    # Strip leading and trailing spaces from column names
                    df = df.rename({col: col.strip() for col in df.columns})
                    
                    # Convert Polars DataFrame to DuckDB table
                    self.connection.execute(f"CREATE OR REPLACE TABLE localdb.{table_name} AS SELECT * FROM df")
                    query = None  # Skip the normal query execution since we already loaded the data
                    
                except Exception as excel_error:
                    raise ValueError(f"Error loading Excel file: {str(excel_error)}")
            elif file_ext == '.pdf':
                # Handle PDF files - open in PDF viewer instead of loading as table
                if PDF_AVAILABLE:
                    self.open_pdf_viewer(file_path)
                    return  # Exit early since we're not loading into a table
                else:
                    raise ValueError("PDF support not available. Please install PyMuPDF (pip install PyMuPDF)")
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Execute the load query with error handling for type conversion
            conversion_error_occurred = False
            needs_column_rename = file_ext in ['.csv', '.tsv', '.parquet', '.json', '.jsonl']
            
            if query is not None:  # Skip execution for Excel files (already loaded)
                try:
                    self.connection.execute(query)
                    
                    # Strip leading and trailing spaces from column names for CSV/TSV/Parquet/JSON
                    if needs_column_rename:
                        # Get column names from temp table
                        columns_result = self.connection.execute(
                            f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}_temp' AND table_schema = 'localdb' ORDER BY ordinal_position"
                        ).fetchall()
                        
                        # Build SELECT with renamed columns
                        renamed_columns = []
                        for col_tuple in columns_result:
                            col_name = col_tuple[0]
                            stripped_name = col_name.strip()
                            if col_name != stripped_name:
                                # Column needs renaming - use AS clause
                                renamed_columns.append(f'"{col_name}" AS "{stripped_name}"')
                            else:
                                # Column is fine as-is
                                renamed_columns.append(f'"{col_name}"')
                        
                        # Create final table with renamed columns
                        rename_query = f"CREATE OR REPLACE TABLE localdb.{table_name} AS SELECT {', '.join(renamed_columns)} FROM localdb.{table_name}_temp"
                        self.connection.execute(rename_query)
                        
                        # Drop temp table
                        self.connection.execute(f"DROP TABLE IF EXISTS localdb.{table_name}_temp")
                except Exception as load_error:
                    # Check if it's a type conversion error
                    error_msg = str(load_error).lower()
                    if any(keyword in error_msg for keyword in ['conversion', 'cast', 'invalid', 'parse', 'type']):
                        conversion_error_occurred = True
                        # Retry with all columns as VARCHAR/text
                        try:
                            if file_ext == '.csv':
                                # For CSV files, use all_varchar option
                                escaped_delimiter = delimiter.replace("'", "''")
                                query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_csv('{file_path}', delim='{escaped_delimiter}', all_varchar=true)"
                            elif file_ext == '.parquet':
                                # For Parquet, we can't force all varchar, so we'll try with union_by_name
                                query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_parquet('{file_path}', union_by_name=true)"
                            elif file_ext in ['.json', '.jsonl']:
                                # For JSON, try with union_by_name and ignore_errors
                                query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_json_auto('{file_path}', union_by_name=true, ignore_errors=true)"
                            elif file_ext == '.tsv':
                                 # For TSV files, use all_varchar option
                                 query = f"CREATE OR REPLACE TABLE localdb.{table_name}_temp AS SELECT * FROM read_csv_auto('{file_path}', delim='\t', all_varchar=true)"
                            elif file_ext in ['.xlsx', '.xls']:
                                # For Excel files, try to reload with string conversion
                                try:
                                    if sheet_name:
                                        df = pl.read_excel(file_path, sheet_name=sheet_name, read_csv_options={"dtypes": str})
                                    else:
                                        df = pl.read_excel(file_path, read_csv_options={"dtypes": str})
                                    # Strip leading and trailing spaces from column names
                                    df = df.rename({col: col.strip() for col in df.columns})
                                    self.connection.execute(f"CREATE OR REPLACE TABLE localdb.{table_name} AS SELECT * FROM df")
                                except:
                                    # If that fails too, just convert all columns to string after loading
                                    # Strip leading and trailing spaces from column names
                                    df = df.rename({col: col.strip() for col in df.columns})
                                    df_str = df.with_columns([pl.col(col).cast(pl.Utf8) for col in df.columns])
                                    self.connection.execute(f"CREATE OR REPLACE TABLE localdb.{table_name} AS SELECT * FROM df_str")
                            
                            if file_ext not in ['.xlsx', '.xls']:  # Only execute query for non-Excel files
                                self.connection.execute(query)
                                
                                # Strip leading and trailing spaces from column names for CSV/TSV/Parquet/JSON
                                if file_ext in ['.csv', '.tsv', '.parquet', '.json', '.jsonl']:
                                    # Get column names from temp table
                                    columns_result = self.connection.execute(
                                        f"SELECT column_name FROM information_schema.columns WHERE table_name = '{table_name}_temp' AND table_schema = 'localdb' ORDER BY ordinal_position"
                                    ).fetchall()
                                    
                                    # Build SELECT with renamed columns
                                    renamed_columns = []
                                    for col_tuple in columns_result:
                                        col_name = col_tuple[0]
                                        stripped_name = col_name.strip()
                                        if col_name != stripped_name:
                                            # Column needs renaming - use AS clause
                                            renamed_columns.append(f'"{col_name}" AS "{stripped_name}"')
                                        else:
                                            # Column is fine as-is
                                            renamed_columns.append(f'"{col_name}"')
                                    
                                    # Create final table with renamed columns
                                    rename_query = f"CREATE OR REPLACE TABLE localdb.{table_name} AS SELECT {', '.join(renamed_columns)} FROM localdb.{table_name}_temp"
                                    self.connection.execute(rename_query)
                                    
                                    # Drop temp table
                                    self.connection.execute(f"DROP TABLE IF EXISTS localdb.{table_name}_temp")
                        except Exception as retry_error:
                            # If retry also fails, show the original error
                            raise load_error
                    else:
                        # If it's not a conversion error, re-raise the original exception
                        raise load_error
            
            # Store the mapping
            self.loaded_tables[filename] = table_name
            
            # Get row count
            count_result = self.connection.execute(f"SELECT COUNT(*) FROM localdb.{table_name}").fetchone()
            row_count = count_result[0]
            
            # Refresh the schema tree to show the new table
            self.refresh_schema_tree()
            
            # Show success message with conversion info if applicable
            success_msg = f'Successfully loaded {filename} as localdb.{table_name}{delimiter_info}'
            if conversion_error_occurred:
                success_msg += ' (columns converted to text due to type conflicts)'
            self.status_label.setText(success_msg)
            
            info_msg = f'File: {filename}\nTable: localdb.{table_name}\nRows: {row_count}{delimiter_info}'
            if conversion_error_occurred:
                info_msg += '\n\nNote: Some columns were converted to text due to type conversion errors.'
            
            QMessageBox.information(
                self, 'File Loaded', info_msg
            )
            
        except Exception as e:
            QMessageBox.critical(self, 'Error Loading File', f'Failed to load file:\n{str(e)}')
            self.status_label.setText('Error loading file')
    
    def open_pdf_viewer(self, file_path):
        """Open PDF file in the current query tab's result area"""
        if not PDF_AVAILABLE:
            QMessageBox.warning(self, 'PDF Support', 'PDF viewing requires PyMuPDF. Please install it with: pip install PyMuPDF')
            return
        
        try:
            # Get current tab index
            current_tab_index = self.query_tab_widget.currentIndex()
            if current_tab_index == -1:
                # No tabs open, create a new one
                self.add_new_query_tab()
                current_tab_index = self.query_tab_widget.currentIndex()
            
            # Get tab data
            tab_data = self.query_tabs[current_tab_index]
            
            # Clear current results
            results_table = tab_data['results_table']
            
            # Hide the results table
            results_table.setVisible(False)
            
            # Get the results group widget (parent of results_table)
            results_group = results_table.parent()
            while results_group and not isinstance(results_group, QGroupBox):
                results_group = results_group.parent()
            
            if results_group:
                # Update the group box title
                results_group.setTitle(f'PDF Viewer - {os.path.basename(file_path)}')
                
                # Get the results layout
                results_layout = results_group.layout()
                
                # Create PDF viewer widget if it doesn't exist
                if not hasattr(tab_data, 'pdf_viewer') or tab_data.get('pdf_viewer') is None:
                    pdf_viewer = PDFViewer()
                    tab_data['pdf_viewer'] = pdf_viewer
                    # Insert PDF viewer before the results table
                    results_layout.insertWidget(results_layout.count() - 1, pdf_viewer)
                else:
                    pdf_viewer = tab_data['pdf_viewer']
                
                # Show PDF viewer and load file
                pdf_viewer.setVisible(True)
                if pdf_viewer.load_pdf(file_path):
                    # Hide pagination controls when showing PDF
                    if 'page_info_label' in tab_data:
                        tab_data['page_info_label'].setVisible(False)
                    if 'first_page_btn' in tab_data:
                        tab_data['first_page_btn'].setVisible(False)
                    if 'prev_page_btn' in tab_data:
                        tab_data['prev_page_btn'].setVisible(False)
                    if 'next_page_btn' in tab_data:
                        tab_data['next_page_btn'].setVisible(False)
                    if 'last_page_btn' in tab_data:
                        tab_data['last_page_btn'].setVisible(False)
                    if 'page_size_combo' in tab_data:
                        tab_data['page_size_combo'].setVisible(False)
                    
                    self.status_label.setText(f'Opened PDF: {os.path.basename(file_path)}')
                else:
                    QMessageBox.warning(self, 'PDF Error', f'Failed to load PDF file: {file_path}')
                    # Show results table again if PDF loading failed
                    results_table.setVisible(True)
                    pdf_viewer.setVisible(False)
                    results_group.setTitle('Query Results')
                    self.status_label.setText('Error opening PDF file')
                
        except Exception as e:
            QMessageBox.critical(self, 'PDF Error', f'Error opening PDF: {str(e)}')
            self.status_label.setText('Error opening PDF viewer')
    
    def restore_query_results_view(self, tab_index=None):
        """Restore the query results table view and hide PDF viewer"""
        if tab_index is None:
            tab_index = self.query_tab_widget.currentIndex()
        
        if tab_index == -1 or tab_index not in self.query_tabs:
            return
        
        tab_data = self.query_tabs[tab_index]
        
        # Show results table
        results_table = tab_data['results_table']
        results_table.setVisible(True)
        
        # Hide PDF viewer if it exists
        if 'pdf_viewer' in tab_data and tab_data['pdf_viewer']:
            tab_data['pdf_viewer'].setVisible(False)
        
        # Show pagination controls
        if 'page_info_label' in tab_data:
            tab_data['page_info_label'].setVisible(True)
        if 'first_page_btn' in tab_data:
            tab_data['first_page_btn'].setVisible(True)
        if 'prev_page_btn' in tab_data:
            tab_data['prev_page_btn'].setVisible(True)
        if 'next_page_btn' in tab_data:
            tab_data['next_page_btn'].setVisible(True)
        if 'last_page_btn' in tab_data:
            tab_data['last_page_btn'].setVisible(True)
        if 'page_size_combo' in tab_data:
            tab_data['page_size_combo'].setVisible(True)
        
        # Restore group box title
        results_group = results_table.parent()
        while results_group and not isinstance(results_group, QGroupBox):
            results_group = results_group.parent()
        
        if results_group:
            results_group.setTitle('Query Results')
        
        self.status_label.setText('Restored query results view')
    
    def load_excel_folder(self):
        """Load all Excel files from a selected folder and merge them into a single table"""
        folder_path = QFileDialog.getExistingDirectory(
            self, 'Select Folder with Excel Files', ''
        )
        
        if not folder_path:
            return
            
        try:
            # Find all Excel files in the folder
            excel_files = []
            for file_name in os.listdir(folder_path):
                if file_name.lower().endswith(('.xlsx', '.xls')):
                    excel_files.append(os.path.join(folder_path, file_name))
            
            if not excel_files:
                QMessageBox.information(self, 'No Excel Files', 'No Excel files found in the selected folder.')
                return
            
            # Prompt user for table name
            folder_name = os.path.basename(folder_path)
            default_table_name = f'{folder_name}_combined'
            default_table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in default_table_name)
            
            table_name, ok = QInputDialog.getText(
                self, 'Table Name', 
                f'Enter a name for the combined table (from {len(excel_files)} Excel files):',
                text=default_table_name
            )
            
            if not ok or not table_name.strip():
                return
            
            # Clean the table name
            table_name = table_name.strip()
            table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)

            # Get sheet names from the first file for reference
            try:
                first_file_sheets = self.get_excel_sheet_names(excel_files[0])
                if not first_file_sheets:
                    first_file_sheets = ['Sheet1']
            except Exception:
                first_file_sheets = ['Sheet1']

            # Ask user which sheets to load
            sheet_dialog = QDialog(self)
            sheet_dialog.setWindowTitle('Select Sheets to Load')
            sheet_dialog.setModal(True)
            sheet_dialog.resize(400, 300)
            
            layout = QVBoxLayout(sheet_dialog)
            
            # Info label
            info_label = QLabel(f'Select which sheets to load from all Excel files.\nReference from first file: {os.path.basename(excel_files[0])}')
            info_label.setWordWrap(True)
            layout.addWidget(info_label)
            
            # Sheet selection
            sheet_list = QListWidget()
            sheet_list.setSelectionMode(QListWidget.MultiSelection)
            
            # Add "First sheet from all files" option at the top
            first_sheet_option = QListWidgetItem("[First sheet from all files]")
            first_sheet_option.setData(Qt.UserRole, "__FIRST_SHEET__")  # Special marker
            sheet_list.addItem(first_sheet_option)
            first_sheet_option.setSelected(True)  # Select by default
            
            # Add actual sheet names from first file
            for sheet_name in first_file_sheets:
                item = QListWidgetItem(sheet_name)
                sheet_list.addItem(item)
            
            layout.addWidget(QLabel('Available sheets:'))
            layout.addWidget(sheet_list)
            
            # Buttons
            button_layout = QHBoxLayout()
            select_all_btn = QPushButton('Select All')
            select_none_btn = QPushButton('Select None')
            ok_btn = QPushButton('OK')
            cancel_btn = QPushButton('Cancel')
            
            def select_all():
                for i in range(sheet_list.count()):
                    sheet_list.item(i).setSelected(True)
            
            def select_none():
                for i in range(sheet_list.count()):
                    sheet_list.item(i).setSelected(False)
            
            select_all_btn.clicked.connect(select_all)
            select_none_btn.clicked.connect(select_none)
            ok_btn.clicked.connect(sheet_dialog.accept)
            cancel_btn.clicked.connect(sheet_dialog.reject)
            
            button_layout.addWidget(select_all_btn)
            button_layout.addWidget(select_none_btn)
            button_layout.addStretch()
            button_layout.addWidget(ok_btn)
            button_layout.addWidget(cancel_btn)
            
            layout.addLayout(button_layout)
            
            if sheet_dialog.exec_() != QDialog.Accepted:
                return
            
            # Get selected sheets
            selected_sheets = []
            use_first_sheet_from_all = False
            
            for i in range(sheet_list.count()):
                item = sheet_list.item(i)
                if item.isSelected():
                    # Check if this is the special "First sheet from all files" option
                    if item.data(Qt.UserRole) == "__FIRST_SHEET__":
                        use_first_sheet_from_all = True
                    else:
                        selected_sheets.append(item.text())
            
            if not selected_sheets and not use_first_sheet_from_all:
                QMessageBox.warning(self, 'No Sheets Selected', 'Please select at least one sheet to load.')
                return

            # Process files with progress dialog
            self.process_excel_folder(excel_files, table_name, folder_path, selected_sheets, use_first_sheet_from_all)
             
        except Exception as e:
             QMessageBox.critical(self, 'Error Loading Folder', f'Failed to load Excel folder:\n{str(e)}')
             self.status_label.setText('Error loading Excel folder')
    
    def process_excel_folder(self, excel_files, table_name, folder_path, selected_sheets, use_first_sheet_from_all=False):
        """Process multiple Excel files and combine them into a single table"""
        # Create progress dialog
        progress = QProgressDialog('Processing Excel files...', 'Cancel', 0, len(excel_files), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.show()
        
        combined_data = []
        all_columns = set()
        file_info = []
        conversion_errors = False
        
        try:
            # First pass: collect all unique column names
            for i, file_path in enumerate(excel_files):
                if progress.wasCanceled():
                    return
                
                progress.setValue(i)
                progress.setLabelText(f'Analyzing schema: {os.path.basename(file_path)}')
                QApplication.processEvents()
                
                try:
                    # Determine which sheets to process for this file
                    sheets_to_process = []
                    
                    if use_first_sheet_from_all:
                        # Get the first sheet from this specific file
                        try:
                            file_sheets = self.get_excel_sheet_names(file_path)
                            if file_sheets:
                                sheets_to_process = [file_sheets[0]]
                        except Exception:
                            sheets_to_process = ['Sheet1']  # Fallback
                    else:
                        # Use the selected sheets from the first file
                        sheets_to_process = selected_sheets
                    
                    # Process the determined sheets
                    for sheet_name in sheets_to_process:
                        try:
                            # Read the sheet to get column names (polars doesn't support n_rows for Excel)
                            df = pl.read_excel(file_path, sheet_name=sheet_name)
                            # Strip leading and trailing spaces from column names
                            all_columns.update([col.strip() for col in df.columns])
                        except Exception as sheet_error:
                            print(f'Warning: Could not read sheet {sheet_name} from {file_path}: {sheet_error}')
                            continue
                            
                except Exception as file_error:
                    print(f'Warning: Could not analyze {file_path}: {file_error}')
                    continue
            
            # Convert to sorted list for consistent column order
            all_columns = sorted(list(all_columns))
            
            # Second pass: load and harmonize data
            for i, file_path in enumerate(excel_files):
                if progress.wasCanceled():
                    return
                
                progress.setValue(i)
                progress.setLabelText(f'Loading: {os.path.basename(file_path)}')
                QApplication.processEvents()
                
                try:
                    file_data = self.load_and_harmonize_excel(file_path, all_columns, selected_sheets, use_first_sheet_from_all)
                    if file_data:
                        combined_data.extend(file_data)
                        file_info.append({
                            'file': os.path.basename(file_path),
                            'rows': len(file_data)
                        })
                except Exception as file_error:
                    conversion_errors = True
                    print(f'Error processing {file_path}: {file_error}')
                    continue
            
            progress.setValue(len(excel_files))
            progress.setLabelText('Creating combined table...')
            QApplication.processEvents()
            
            if not combined_data:
                QMessageBox.warning(self, 'No Data', 'No data could be loaded from the Excel files.')
                return
            
            # Create combined DataFrame
            combined_df = pl.DataFrame(combined_data, schema={col: pl.Utf8 for col in all_columns})
            
            # Add source file column
            source_files = []
            for file_info_item in file_info:
                source_files.extend([file_info_item['file']] * file_info_item['rows'])
            
            if len(source_files) == len(combined_data):
                combined_df = combined_df.with_columns(pl.Series('source_file', source_files))
            
            # Load into DuckDB
            self.connection.execute(f'CREATE OR REPLACE TABLE localdb.{table_name} AS SELECT * FROM combined_df')
            
            # Store the mapping
            folder_name = os.path.basename(folder_path)
            self.loaded_tables[f'{folder_name}_combined'] = table_name
            
            # Get row count
            count_result = self.connection.execute(f'SELECT COUNT(*) FROM localdb.{table_name}').fetchone()
            row_count = count_result[0]
            
            # Refresh schema tree
            self.refresh_schema_tree()
            
            # Show success message
            success_msg = f'Successfully combined {len(file_info)} Excel files into localdb.{table_name}'
            if conversion_errors:
                success_msg += ' (some files had errors and were skipped)'
            self.status_label.setText(success_msg)
            
            # Detailed info message
            file_list = '\n'.join([f'- {info["file"]}: {info["rows"]} rows' for info in file_info])
            info_msg = f'Combined Table: localdb.{table_name}\nTotal Rows: {row_count}\nColumns: {len(all_columns)}\n\nFiles processed:\n{file_list}'
            if conversion_errors:
                info_msg += '\n\nNote: All columns were converted to text for compatibility. Some files may have been skipped due to errors.'
            
            QMessageBox.information(self, 'Folder Loaded', info_msg)
            
        finally:
             progress.close()
    
    def get_excel_sheet_names(self, file_path):
        """Get sheet names from Excel file using the same logic as ExcelSheetSelectionDialog"""
        try:
            # Use openpyxl to get sheet names only (much faster)
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=False)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception:
            # Fallback: try with pandas ExcelFile
            try:
                import pandas as pd
                with pd.ExcelFile(file_path) as xls:
                    return xls.sheet_names
            except:
                # Last fallback: use polars
                try:
                    import polars as pl
                    xl_file = pl.read_excel(file_path, sheet_id=None)
                    if isinstance(xl_file, dict):
                        return list(xl_file.keys())
                    else:
                        return ['Sheet1']
                except:
                    return ['Sheet1']  # Final fallback
    
    def load_and_harmonize_excel(self, file_path, all_columns, selected_sheets=None, use_first_sheet_from_all=False):
        """Load Excel file and harmonize its data to match the combined schema"""
        file_data = []
        
        try:
            # Determine which sheets to process
            if use_first_sheet_from_all:
                # Use first sheet from this file
                all_sheet_names = self.get_excel_sheet_names(file_path)
                sheet_names = [all_sheet_names[0]] if all_sheet_names else []
            elif selected_sheets:
                sheet_names = selected_sheets
            else:
                sheet_names = self.get_excel_sheet_names(file_path)
            
            for sheet_name in sheet_names:
                try:
                    # Load the sheet data, converting everything to string
                    df = pl.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Strip leading and trailing spaces from column names
                    df = df.rename({col: col.strip() for col in df.columns})
                    
                    # Convert all columns to string to avoid type conflicts
                    df = df.with_columns([pl.col(col).cast(pl.Utf8, strict=False) for col in df.columns])
                    
                    # Convert to list of dictionaries for easier manipulation
                    sheet_data = df.to_dicts()
                    
                    # Harmonize each row to match the combined schema
                    for row in sheet_data:
                        harmonized_row = {}
                        
                        # Add existing columns
                        for col in all_columns:
                            if col in row:
                                # Convert to string and handle None values
                                value = row[col]
                                if value is None or (isinstance(value, str) and value.lower() in ['nan', 'null', '']):
                                    harmonized_row[col] = None
                                else:
                                    harmonized_row[col] = str(value)
                            else:
                                # Column doesn't exist in this file, set to None
                                harmonized_row[col] = None
                        
                        file_data.append(harmonized_row)
                        
                except Exception as sheet_error:
                    print(f'Warning: Could not load sheet {sheet_name} from {file_path}: {sheet_error}')
                    continue
                    
        except Exception as file_error:
            print(f'Error loading {file_path}: {file_error}')
            raise
        
        return file_data
    
    def get_active_sql_editor(self):
        """
        Determine which SQL editor currently has focus or should receive input.
        Returns a tuple: (sql_editor, is_split_screen)
        """
        # First check if we have a tracked last active editor and it's still valid
        if (self.last_active_sql_editor and 
            hasattr(self.last_active_sql_editor, 'hasFocus')):
            
            # Check if it's a split screen editor
            if hasattr(self, 'split_query_tabs'):
                for tab_key, tab_data in self.split_query_tabs.items():
                    if tab_data['sql_editor'] == self.last_active_sql_editor:
                        return self.last_active_sql_editor, True
            
            # Check if it's the main editor
            current_tab_index = self.query_tab_widget.currentIndex()
            if (current_tab_index in self.query_tabs and 
                self.query_tabs[current_tab_index]['sql_editor'] == self.last_active_sql_editor):
                return self.last_active_sql_editor, False
        
        # Fallback: check if any split screen SQL editor has focus
        if hasattr(self, 'split_query_tabs'):
            for tab_key, tab_data in self.split_query_tabs.items():
                sql_editor = tab_data['sql_editor']
                if sql_editor.hasFocus():
                    self.last_active_sql_editor = sql_editor
                    return sql_editor, True
        
        # Check if main SQL editor has focus or fall back to it as default
        current_tab_index = self.query_tab_widget.currentIndex()
        if current_tab_index in self.query_tabs:
            main_sql_editor = self.query_tabs[current_tab_index]['sql_editor']
            if main_sql_editor.hasFocus():
                self.last_active_sql_editor = main_sql_editor
            return main_sql_editor, False
        
        # Fallback: return None if no valid editor found
        return None, False
    
    def on_table_double_click(self, item, column):
        """Handle double-click on table items in the schema tree"""
        if item.text(1) == 'table':  # Only handle table items, not schema
            full_table_name = item.data(0, Qt.UserRole)
            if full_table_name:
                # Get the active SQL editor (could be main or split screen)
                sql_editor, is_split_screen = self.get_active_sql_editor()
                
                if sql_editor:
                    # Insert the table name into the active SQL editor
                    current_text = sql_editor.toPlainText()
                    if current_text.strip():
                        # If there's existing text, add the table name at cursor position
                        cursor = sql_editor.textCursor()
                        cursor.insertText(full_table_name)
                    else:
                        # If empty, add a basic SELECT query
                        sql_editor.setPlainText(f'SELECT * FROM {full_table_name} LIMIT 10;')
                    
                    # Set focus to the SQL editor
                    sql_editor.setFocus()
    
    def execute_query(self):
        """Execute the SQL query for the current tab"""
        current_tab_index = self.query_tab_widget.currentIndex()
        self.execute_query_for_tab(current_tab_index)
    
    def execute_selected_query(self):
        """Execute only the selected text as a SQL query"""
        current_tab_index = self.query_tab_widget.currentIndex()
        if current_tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[current_tab_index]
        sql_editor = tab_data['sql_editor']
        
        # Get selected text
        cursor = sql_editor.textCursor()
        if not cursor.hasSelection():
            QMessageBox.warning(self, 'No Selection', 'Please select the SQL text you want to execute.')
            return
        
        # Fix for comments: Get selected text properly preserving line breaks
        # cursor.selectedText() converts line breaks to \u2029 which breaks SQL comments
        start_pos = cursor.selectionStart()
        end_pos = cursor.selectionEnd()
        
        # Get the full text and extract the selected portion
        full_text = sql_editor.toPlainText()
        selected_query = full_text[start_pos:end_pos].strip()
        
        if not selected_query:
            QMessageBox.warning(self, 'Empty Selection', 'Please select a valid SQL query.')
            return
        
        # Check if we have any data available (local tables or remote connections)
        if not self.loaded_tables and not self.active_connections:
            QMessageBox.warning(self, 'No Data', 'Please load a data file or connect to a database first.')
            return
        
        # Store current query and reset pagination
        tab_data['current_query'] = selected_query
        tab_data['current_page'] = 0
        tab_data['total_rows'] = 0
        
        # Execute streaming query
        self.execute_streaming_query(current_tab_index)
    
    def display_results(self, columns, data):
        """Legacy method - now handled by display_results_for_tab"""
        current_tab_index = self.query_tab_widget.currentIndex()
        self.display_results_for_tab(current_tab_index, columns, data)
    
    def handle_query_error(self, error_message):
        """Legacy method - now handled by handle_query_error_for_tab"""
        current_tab_index = self.query_tab_widget.currentIndex()
        self.handle_query_error_for_tab(current_tab_index, error_message)
    
    def create_menu_bar(self):
        """Create the application menu bar"""
        menubar = self.menuBar()
        
        # File menu
        db_menu = menubar.addMenu('File')
        
        # Load File action
        load_file_action = db_menu.addAction('Load File...')
        load_file_action.triggered.connect(self.load_file)
        
        # Load Folder of Excel Files action
        load_folder_action = db_menu.addAction('Load Folder of Excel Files...')
        load_folder_action.triggered.connect(self.load_excel_folder)
        
        db_menu.addSeparator()
        
        # Connect to Database action
        connect_action = db_menu.addAction('Connect to Database...')
        connect_action.triggered.connect(self.show_connection_dialog)
        
        # Manage Connections action
        manage_action = db_menu.addAction('Manage Connections...')
        manage_action.triggered.connect(self.show_connection_manager)
        
        db_menu.addSeparator()
        
        # Disconnect action
        self.disconnect_action = db_menu.addAction('Disconnect')
        self.disconnect_action.triggered.connect(self.disconnect_database)
        self.disconnect_action.setEnabled(False)
        
        db_menu.addSeparator()
        
        # Query management actions
        save_query_action = db_menu.addAction('Save Query...')
        save_query_action.triggered.connect(self.save_query)
        
        load_query_action = db_menu.addAction('Load Query...')
        load_query_action.triggered.connect(self.load_query)
        
        manage_queries_action = db_menu.addAction('Manage Saved Queries...')
        manage_queries_action.triggered.connect(self.show_query_manager)
        
        db_menu.addSeparator()
        
        # Export submenu
        export_menu = db_menu.addMenu('Export Results')
        
        # Export to Excel action
        self.export_excel_action = export_menu.addAction('Export to Excel...')
        self.export_excel_action.triggered.connect(lambda: self.export_results('excel', self.query_tab_widget.currentIndex()))
        self.export_excel_action.setEnabled(False)
        
        # Export to CSV action
        self.export_csv_action = export_menu.addAction('Export to CSV...')
        self.export_csv_action.triggered.connect(lambda: self.export_results('csv', self.query_tab_widget.currentIndex()))
        self.export_csv_action.setEnabled(False)
        
        # Export to JSON action
        self.export_json_action = export_menu.addAction('Export to JSON...')
        self.export_json_action.triggered.connect(lambda: self.export_results('json', self.query_tab_widget.currentIndex()))
        self.export_json_action.setEnabled(False)
        
        # Export to Parquet action
        self.export_parquet_action = export_menu.addAction('Export to Parquet...')
        self.export_parquet_action.triggered.connect(lambda: self.export_results('parquet', self.query_tab_widget.currentIndex()))
        self.export_parquet_action.setEnabled(False)
        
        # View menu
        view_menu = menubar.addMenu('View')
        
        # Theme submenu
        theme_menu = view_menu.addMenu('Theme')
        
        # Settings menu
        settings_menu = menubar.addMenu('Settings')
        

        
        # Theme actions
        self.light_theme_action = theme_menu.addAction('Light Theme')
        self.light_theme_action.setCheckable(True)
        self.light_theme_action.triggered.connect(lambda: self.set_theme('light'))
        
        self.dark_theme_action = theme_menu.addAction('Dark Theme')
        self.dark_theme_action.setCheckable(True)
        self.dark_theme_action.triggered.connect(lambda: self.set_theme('dark'))
        
        self.blue_theme_action = theme_menu.addAction('Blue Theme')
        self.blue_theme_action.setCheckable(True)
        self.blue_theme_action.triggered.connect(lambda: self.set_theme('blue'))
        
        self.green_theme_action = theme_menu.addAction('Green Theme')
        self.green_theme_action.setCheckable(True)
        self.green_theme_action.triggered.connect(lambda: self.set_theme('green'))
        
        self.high_contrast_theme_action = theme_menu.addAction('High Contrast Theme')
        self.high_contrast_theme_action.setCheckable(True)
        self.high_contrast_theme_action.triggered.connect(lambda: self.set_theme('high_contrast'))
        
        # Group theme actions
        from PyQt5.QtWidgets import QActionGroup
        self.theme_group = QActionGroup(self)
        self.theme_group.addAction(self.light_theme_action)
        self.theme_group.addAction(self.dark_theme_action)
        self.theme_group.addAction(self.blue_theme_action)
        self.theme_group.addAction(self.green_theme_action)
        self.theme_group.addAction(self.high_contrast_theme_action)
        
        # Set default theme
        self.light_theme_action.setChecked(True)
    
    def show_connection_dialog(self):
        """Show the database connection dialog"""
        dialog = DatabaseConnectionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            connection_data = dialog.get_connection_data()
            self.connect_to_database(connection_data)
    
    def show_connection_manager(self):
        """Show the connection manager dialog"""
        dialog = ConnectionManagerDialog(self, self.saved_connections)
        if dialog.exec_() == QDialog.Accepted:
            self.saved_connections = dialog.get_connections()
            self.save_connections()
    
    def connect_to_database(self, connection_data):
        """Connect to a database using the provided connection data"""
        try:
            # Create unique database name from connection name
            db_name = self.sanitize_db_name(connection_data['name'])
            
            # Check if this connection is already active
            if db_name in self.active_connections:
                QMessageBox.information(self, 'Already Connected', f'Already connected to {connection_data["name"]}')
                return
            
            # Install and load MySQL extension
            try:
                self.connection.execute('INSTALL mysql')
                self.connection.execute('LOAD mysql')
            except Exception as ext_error:
                # Extension might already be installed
                try:
                    self.connection.execute('LOAD mysql')
                except:
                    raise Exception(f"Failed to load MySQL extension. Please ensure DuckDB MySQL extension is available.\nError: {str(ext_error)}")
            
            # Build connection string
            connection_string = self.build_connection_string(connection_data)
            
            # Attach the database with unique name
            # Use different syntax based on DuckDB version
            try:
                # Try newer DuckDB syntax first
                attach_query = f"ATTACH '{connection_string}' AS {db_name} (TYPE mysql, READ_ONLY)"
                self.connection.execute(attach_query)
            except Exception as attach_error:
                # Try alternative syntax
                error_msg = str(attach_error).lower()
                
                # Provide helpful error messages
                if 'io error' in error_msg or 'failed to connect' in error_msg:
                    raise Exception(
                        f"Failed to connect to MySQL database.\n\n"
                        f"Connection details:\n"
                        f"  Host: {connection_data.get('host', 'N/A')}\n"
                        f"  Port: {connection_data.get('port', 3306)}\n"
                        f"  Database: {connection_data.get('database', 'N/A')}\n"
                        f"  User: {connection_data.get('username', 'N/A')}\n\n"
                        f"Please verify:\n"
                        f"  1. MySQL server is running and accessible\n"
                        f"  2. Host and port are correct\n"
                        f"  3. Username and password are correct\n"
                        f"  4. Database name exists\n"
                        f"  5. User has permission to access the database\n"
                        f"  6. Firewall allows connection to MySQL port\n\n"
                        f"Original error: {str(attach_error)}"
                    )
                else:
                    raise attach_error
            
            # Store connection data
            self.active_connections[db_name] = connection_data
            
            # Update UI
            self.disconnect_action.setEnabled(True)
            
            # Refresh schema tree to show remote tables
            self.refresh_schema_tree()
            
            QMessageBox.information(self, 'Connection Successful', f'Successfully connected to {connection_data["name"]}')
            
            # Save this connection if it's not already saved
            existing_names = [conn['name'] for conn in self.saved_connections]
            if connection_data['name'] not in existing_names:
                self.saved_connections.append(connection_data)
                self.save_connections()
            
        except Exception as e:
            error_message = str(e)
            QMessageBox.critical(self, 'Connection Failed', f'Failed to connect to database:\n\n{error_message}')
    
    def sanitize_db_name(self, name):
        """Convert connection name to valid database identifier"""
        # Replace spaces and special characters with underscores
        sanitized = ''.join(c if c.isalnum() else '_' for c in name)
        # Ensure it starts with a letter or underscore
        if sanitized and sanitized[0].isdigit():
            sanitized = '_' + sanitized
        # Limit length and ensure it's not empty
        sanitized = sanitized[:50] or 'db_connection'
        return sanitized.lower()
    
    def disconnect_database(self, db_name=None):
        """Disconnect from remote database(s)"""
        try:
            if db_name:
                # Disconnect specific database
                if db_name in self.active_connections:
                    self.connection.execute(f'DETACH {db_name}')
                    del self.active_connections[db_name]
            else:
                # Disconnect all databases
                for db_name in list(self.active_connections.keys()):
                    try:
                        self.connection.execute(f'DETACH {db_name}')
                    except:
                        pass  # Continue even if detach fails
                self.active_connections.clear()
            
            # Update UI
            if not self.active_connections:
                self.disconnect_action.setEnabled(False)
            
            self.refresh_schema_tree()
            
        except Exception as e:
            QMessageBox.warning(self, 'Disconnect Warning', f'Error during disconnect:\n{str(e)}')
    
    def build_connection_string(self, connection_data):
        """Build connection string from connection data"""
        return build_mysql_connection_string(connection_data)
    
    def refresh_schema_tree(self):
        """Refresh the schema tree to show current schemas only (lazy loading - optimized)"""
        self.schema_tree.clear()
        
        # Create localdb schema node (always show it)
        localdb_node = QTreeWidgetItem(self.schema_tree)
        localdb_node.setText(0, 'localdb')
        localdb_node.setText(1, 'schema')
        localdb_node.setData(0, Qt.UserRole, 'localdb')  # Store schema identifier
        
        # Always add placeholder to make it expandable (lazy check for tables)
        placeholder = QTreeWidgetItem(localdb_node)
        placeholder.setText(0, 'Loading...')
        placeholder.setText(1, 'placeholder')
        
        # Add remote database schemas for all active connections
        for db_name, connection_data in self.active_connections.items():
            # Get the database name from connection data
            database_name = connection_data.get('database', '')
            connection_name = connection_data.get('name', 'Remote Database')
            
            # Create node without checking table count (lazy loading)
            remote_node = QTreeWidgetItem(self.schema_tree)
            remote_node.setText(0, f'{connection_name} ({database_name})')
            remote_node.setText(1, 'schema')
            remote_node.setData(0, Qt.UserRole, db_name)  # Store db_name for disconnect
            
            # Add placeholder child to make node expandable
            placeholder = QTreeWidgetItem(remote_node)
            placeholder.setText(0, 'Loading...')
            placeholder.setText(1, 'placeholder')
    
    def update_autocomplete_with_tables(self):
        """Update SQL autocomplete with current table names from all schemas (cached)"""
        # Use cached table names if available to avoid repeated queries
        if not hasattr(self, '_cached_table_names') or not hasattr(self, '_cache_timestamp'):
            self._cached_table_names = []
            self._cache_timestamp = 0
        
        import time
        current_time = time.time()
        
        # Only refresh cache if it's older than 5 seconds
        if current_time - self._cache_timestamp > 5:
            table_names = []
            
            try:
                # Get localdb tables
                local_tables = self.connection.execute(
                    "SELECT table_name FROM information_schema.tables WHERE table_schema = 'localdb'"
                ).fetchall()
                
                for table in local_tables:
                    table_name = table[0]
                    table_names.extend([
                        table_name,
                        f'localdb.{table_name}'
                    ])
                
                # Get remote database tables (limit to avoid slowdown)
                for db_name, connection_data in self.active_connections.items():
                    try:
                        database_name = connection_data.get('database', '')
                        # Limit to first 1000 tables for performance
                        remote_tables = self.connection.execute(
                            f"SELECT table_name FROM information_schema.tables WHERE table_catalog = '{db_name}' AND table_schema = '{database_name}' LIMIT 1000"
                        ).fetchall()
                        
                        for table in remote_tables:
                            table_name = table[0]
                            table_names.extend([
                                table_name,
                                f'{db_name}.{table_name}'
                            ])
                    except:
                        pass
                
                self._cached_table_names = table_names
                self._cache_timestamp = current_time
            except Exception as e:
                pass  # Error updating autocomplete
        
        # Update autocomplete for all SQL editors with cached names
        if self._cached_table_names:
            for i in range(self.query_tab_widget.count()):
                tab_widget = self.query_tab_widget.widget(i)
                if hasattr(tab_widget, 'findChild'):
                    sql_editor = tab_widget.findChild(SQLTextEdit)
                    if sql_editor:
                        sql_editor.add_custom_completions(self._cached_table_names)
    
    def update_autocomplete_for_editor(self, sql_editor):
        """Update autocomplete for a specific SQL editor with current table names (cached)"""
        # Use cached table names if available
        if hasattr(self, '_cached_table_names') and self._cached_table_names:
            sql_editor.add_custom_completions(self._cached_table_names)
        else:
            # Build cache if not available
            self.update_autocomplete_with_tables()
    
    def on_schema_item_expanded(self, item):
        """Handle lazy loading when schema tree items are expanded"""
        # Check if this item has placeholder children that need to be replaced
        if item.childCount() > 0 and item.child(0).text(1) == 'placeholder':
            # Remove placeholder
            item.removeChild(item.child(0))
            
            # Load actual content based on item type
            if item.text(1) == 'schema':
                self.load_tables_for_schema(item)
            elif item.text(1) == 'table':
                self.load_columns_for_table(item)
    
    def load_tables_for_schema(self, schema_item):
        """Load tables for a schema node (optimized - no row counts)"""
        schema_identifier = schema_item.data(0, Qt.UserRole)
        
        if schema_identifier == 'localdb':
            # Load local tables (without counting rows for performance)
            try:
                tables = self.connection.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'localdb' ORDER BY table_name").fetchall()
                for table in tables:
                    table_name = table[0]
                    
                    table_item = QTreeWidgetItem(schema_item)
                    table_item.setText(0, table_name)
                    table_item.setText(1, 'table')
                    table_item.setText(2, '')  # Don't show row count for performance
                    table_item.setData(0, Qt.UserRole, f'localdb.{table_name}')
                    
                    # Add placeholder for columns to make table expandable
                    placeholder = QTreeWidgetItem(table_item)
                    placeholder.setText(0, 'Loading...')
                    placeholder.setText(1, 'placeholder')
            except:
                pass  # Error loading local tables
        else:
            # Load remote tables (without counting rows for performance)
            if schema_identifier in self.active_connections:
                connection_data = self.active_connections[schema_identifier]
                database_name = connection_data.get('database', '')
                
                try:
                    remote_tables = self.connection.execute(
                        f"SELECT table_name FROM information_schema.tables WHERE table_catalog = '{schema_identifier}' AND table_schema = '{database_name}' ORDER BY table_name"
                    ).fetchall()
                    
                    for table in remote_tables:
                        table_name = table[0]
                        
                        table_item = QTreeWidgetItem(schema_item)
                        table_item.setText(0, table_name)
                        table_item.setText(1, 'table')
                        table_item.setText(2, '')  # Don't show row count for performance
                        table_item.setData(0, Qt.UserRole, f'{schema_identifier}.{table_name}')
                        
                        # Add placeholder for columns to make table expandable
                        placeholder = QTreeWidgetItem(table_item)
                        placeholder.setText(0, 'Loading...')
                        placeholder.setText(1, 'placeholder')
                except:
                    pass  # Error loading remote tables
        
        # Update autocomplete with newly loaded table names (async to avoid blocking)
        QApplication.processEvents()  # Keep UI responsive
    
    def load_columns_for_table(self, table_item):
        """Load columns for a table node"""
        full_table_name = table_item.data(0, Qt.UserRole)
        if not full_table_name:
            return
        
        # Parse schema and table name
        if '.' in full_table_name:
            schema_name, table_name = full_table_name.split('.', 1)
        else:
            return
        
        try:
            if schema_name == 'localdb':
                # Load columns for local table
                columns = self.connection.execute(
                    f"SELECT column_name, data_type FROM information_schema.columns WHERE table_name = '{table_name}' AND table_schema = 'localdb' ORDER BY ordinal_position"
                ).fetchall()
            else:
                # Load columns for remote table
                if schema_name in self.active_connections:
                    connection_data = self.active_connections[schema_name]
                    database_name = connection_data.get('database', '')
                    columns = self.connection.execute(
                        f"SELECT column_name, data_type FROM information_schema.columns WHERE table_name = '{table_name}' AND table_catalog = '{schema_name}' AND table_schema = '{database_name}' ORDER BY ordinal_position"
                    ).fetchall()
                else:
                    return
            
            # Add column items
            for column in columns:
                column_name, data_type = column
                column_item = QTreeWidgetItem(table_item)
                column_item.setText(0, column_name)
                column_item.setText(1, 'column')
                column_item.setText(2, data_type)
        except:
            pass  # Error loading columns
    
    def show_schema_context_menu(self, position):
        """Show context menu for schema tree items"""
        item = self.schema_tree.itemAt(position)
        if not item:
            return
        
        menu = QMenu(self)
        
        # Check if this is a table item in the localdb schema
        if item.text(1) == 'table':
            full_table_name = item.data(0, Qt.UserRole)
            if full_table_name and full_table_name.startswith('localdb.'):
                # This is a localdb table - add remove option
                table_name = full_table_name.replace('localdb.', '')
                remove_action = QAction(f'Remove Table "{table_name}"', self)
                remove_action.triggered.connect(lambda: self.remove_table(table_name))
                menu.addAction(remove_action)
                
                menu.exec_(self.schema_tree.mapToGlobal(position))
                return
        
        # Check if this is a remote database schema node
        if item.text(1) == 'schema' and item.data(0, Qt.UserRole):
            db_name = item.data(0, Qt.UserRole)
            if db_name in self.active_connections:
                disconnect_action = QAction(f'Disconnect from {item.text(0)}', self)
                disconnect_action.triggered.connect(lambda: self.disconnect_specific_database(db_name))
                menu.addAction(disconnect_action)
                
                menu.exec_(self.schema_tree.mapToGlobal(position))
    
    def remove_table(self, table_name):
        """Remove a table from the localdb schema"""
        try:
            # Show confirmation dialog
            reply = QMessageBox.question(
                self, 
                'Confirm Table Removal',
                f'Are you sure you want to remove the table "{table_name}" from localdb?\n\nThis action cannot be undone.',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Drop the table
                self.connection.execute(f'DROP TABLE IF EXISTS localdb.{table_name}')
                
                # Remove from loaded_tables mapping if it exists
                table_to_remove = None
                for filename, loaded_table_name in self.loaded_tables.items():
                    if loaded_table_name == table_name:
                        table_to_remove = filename
                        break
                
                if table_to_remove:
                    del self.loaded_tables[table_to_remove]
                
                # Refresh the schema tree
                self.refresh_schema_tree()
                
                # Update status
                self.status_label.setText(f'Table "{table_name}" removed from localdb')
                
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to remove table "{table_name}":\n{str(e)}')
    
    def disconnect_specific_database(self, db_name):
        """Disconnect from a specific database"""
        if db_name in self.active_connections:
            connection_name = self.active_connections[db_name]['name']
            self.disconnect_database(db_name)
            self.status_label.setText(f'Disconnected from {connection_name}')
    
    def show_results_context_menu(self, position, tab_index):
        """Show context menu for results table"""
        if tab_index not in self.query_tabs:
            return
            
        results_table = self.query_tabs[tab_index]['results_table']
        item = results_table.itemAt(position)
        
        if not item:
            return
            
        menu = QMenu(self)
        
        # Copy cell value
        copy_cell_action = QAction('Copy Cell Value', self)
        copy_cell_action.triggered.connect(lambda: self.copy_cell_value(tab_index, item.row(), item.column()))
        menu.addAction(copy_cell_action)
        
        # Copy column with header
        copy_column_action = QAction('Copy Column with Header', self)
        copy_column_action.triggered.connect(lambda: self.copy_column_with_header(tab_index, item.column()))
        menu.addAction(copy_column_action)
        
        # Copy row with header
        copy_row_action = QAction('Copy Row with Header', self)
        copy_row_action.triggered.connect(lambda: self.copy_row_with_header(tab_index, item.row()))
        menu.addAction(copy_row_action)
        
        menu.addSeparator()
        
        # Copy entire table
        copy_table_action = QAction('Copy Entire Table', self)
        copy_table_action.triggered.connect(lambda: self.copy_entire_table(tab_index))
        menu.addAction(copy_table_action)
        
        # Add separator and Dashboard option
        menu.addSeparator()
        
        # Build Nodes Dashboard option
        nodes_dashboard_action = QAction('🕸️ Build Nodes Dashboard', self)
        nodes_dashboard_action.triggered.connect(lambda: self.open_nodes_dashboard(tab_index))
        menu.addAction(nodes_dashboard_action)

        # Build Main Dashboard option
        main_dashboard_action = QAction('📊 Build Dashboard', self)
        main_dashboard_action.triggered.connect(lambda: self.open_main_dashboard(tab_index))
        menu.addAction(main_dashboard_action)
        
        if EEL_AVAILABLE:
            dashboard_action = QAction('📊 Open Dashboard', self)
            dashboard_action.triggered.connect(lambda: self.open_eel_dashboard(tab_index))
            menu.addAction(dashboard_action)
        
        menu.exec_(results_table.mapToGlobal(position))
    
    def show_header_context_menu(self, position, tab_index):
        """Show context menu for table headers"""
        if tab_index not in self.query_tabs:
            return
            
        results_table = self.query_tabs[tab_index]['results_table']
        header = results_table.horizontalHeader()
        
        # Get the column index from the position
        column = header.logicalIndexAt(position)
        if column < 0:
            return
            
        menu = QMenu(self)
        
        # Copy header value
        copy_header_action = QAction('Copy Header', self)
        copy_header_action.triggered.connect(lambda: self.copy_header_value(tab_index, column))
        menu.addAction(copy_header_action)
        
        menu.exec_(header.mapToGlobal(position))
    
    def copy_header_value(self, tab_index, column):
        """Copy the header value to clipboard"""
        if tab_index not in self.query_tabs:
            return
            
        results_table = self.query_tabs[tab_index]['results_table']
        header_item = results_table.horizontalHeaderItem(column)
        
        if header_item:
            clipboard = QApplication.clipboard()
            clipboard.setText(header_item.text())
    
    def copy_header_value_for_split(self, tab_widget, tab_index, column):
        """Copy the header value to clipboard for split screen tabs"""
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        results_table = tab_data['results_table']
        header_item = results_table.horizontalHeaderItem(column)
        
        if header_item:
            clipboard = QApplication.clipboard()
            clipboard.setText(header_item.text())
    
    def copy_cell_value_for_split(self, tab_widget, tab_index, row, column):
        """Copy the value of a specific cell to clipboard for split screen tabs"""
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        results_table = tab_data['results_table']
        item = results_table.item(row, column)
        
        if item:
            clipboard = QApplication.clipboard()
            clipboard.setText(item.text())
    
    def copy_column_with_header_for_split(self, tab_widget, tab_index, column):
        """Copy entire column with header to clipboard for split screen tabs"""
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        results_table = tab_data['results_table']
        
        # Get column header
        header_item = results_table.horizontalHeaderItem(column)
        header_text = header_item.text() if header_item else f'Column {column + 1}'
        
        # Get full data from complete query execution
        try:
            columns, full_data = self.execute_complete_query_for_split(tab_key)
            if not columns or column >= len(columns):
                return
                
            # Collect column data with header
            column_data = [header_text]
            
            for row_data in full_data:
                if column < len(row_data):
                    cell_value = str(row_data[column]) if row_data[column] is not None else ''
                    column_data.append(cell_value)
                else:
                    column_data.append('')
            
            # Copy to clipboard
            clipboard = QApplication.clipboard()
            clipboard.setText('\n'.join(column_data))
            
        except Exception as e:
            QMessageBox.critical(self, 'Copy Error', f'Failed to copy column data:\n{str(e)}')
    
    def copy_row_with_header_for_split(self, tab_widget, tab_index, row):
        """Copy entire row with headers to clipboard for split screen tabs"""
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        results_table = tab_data['results_table']
        
        # Get headers
        headers = []
        for col in range(results_table.columnCount()):
            header_item = results_table.horizontalHeaderItem(col)
            header_text = header_item.text() if header_item else f'Column {col + 1}'
            headers.append(header_text)
        
        # Get row data
        row_data = []
        for col in range(results_table.columnCount()):
            item = results_table.item(row, col)
            cell_text = item.text() if item else ''
            row_data.append(cell_text)
        
        # Format as tab-separated values with headers
        result = '\t'.join(headers) + '\n' + '\t'.join(row_data)
        
        # Copy to clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText(result)
    
    def copy_entire_table_for_split(self, tab_widget, tab_index):
        """Copy entire table with headers to clipboard for split screen tabs"""
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if tab_key not in self.split_query_tabs:
            return
            
        # Get full data from complete query execution
        try:
            columns, full_data = self.execute_complete_query_for_split(tab_key)
            if not columns:
                return
                
            # Start with headers
            table_data = ['\t'.join(columns)]
            
            # Add all row data
            for row_data in full_data:
                formatted_row = []
                for cell_value in row_data:
                    cell_text = str(cell_value) if cell_value is not None else ''
                    formatted_row.append(cell_text)
                table_data.append('\t'.join(formatted_row))
            
            # Copy to clipboard
            clipboard = QApplication.clipboard()
            clipboard.setText('\n'.join(table_data))
            
        except Exception as e:
            QMessageBox.critical(self, 'Copy Error', f'Failed to copy table data:\n{str(e)}')
    
    def execute_complete_query_for_split(self, tab_key):
        """Execute the complete query without pagination to get all results for split screen tabs"""
        if tab_key not in self.split_query_tabs:
            raise Exception("Invalid tab key")
            
        tab_data = self.split_query_tabs[tab_key]
        query = tab_data['current_query']
        
        if not query:
            raise Exception("No query to execute")
            
        if not self.connection:
            raise Exception("No database connection available")
            
        # Execute the complete query without pagination
        cursor = self.connection.execute(query)
        columns = [desc[0] for desc in cursor.description]
        
        # Fetch all data
        full_data = cursor.fetchall()
        
        return columns, full_data
    
    def copy_cell_value(self, tab_index, row, column):
        """Copy the value of a specific cell to clipboard"""
        if tab_index not in self.query_tabs:
            return
            
        results_table = self.query_tabs[tab_index]['results_table']
        item = results_table.item(row, column)
        
        if item:
            clipboard = QApplication.clipboard()
            clipboard.setText(item.text())
    
    def copy_column_with_header(self, tab_index, column):
        """Copy entire column with header to clipboard"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        results_table = tab_data['results_table']
        
        # Get column header
        header_item = results_table.horizontalHeaderItem(column)
        header_text = header_item.text() if header_item else f'Column {column + 1}'
        
        # Get full data from complete query execution
        try:
            columns, full_data = self.execute_complete_query(tab_index)
            if not columns or column >= len(columns):
                return
                
            # Collect column data with header
            column_data = [header_text]
            
            for row_data in full_data:
                if column < len(row_data):
                    cell_value = str(row_data[column]) if row_data[column] is not None else ''
                    column_data.append(cell_value)
                else:
                    column_data.append('')
            
            # Copy to clipboard
            clipboard = QApplication.clipboard()
            clipboard.setText('\n'.join(column_data))
            
        except Exception as e:
            QMessageBox.critical(self, 'Copy Error', f'Failed to copy column data:\n{str(e)}')
    
    def copy_row_with_header(self, tab_index, row):
        """Copy entire row with headers to clipboard"""
        if tab_index not in self.query_tabs:
            return
            
        results_table = self.query_tabs[tab_index]['results_table']
        
        # Get headers
        headers = []
        for col in range(results_table.columnCount()):
            header_item = results_table.horizontalHeaderItem(col)
            header_text = header_item.text() if header_item else f'Column {col + 1}'
            headers.append(header_text)
        
        # Get row data
        row_data = []
        for col in range(results_table.columnCount()):
            item = results_table.item(row, col)
            cell_text = item.text() if item else ''
            row_data.append(cell_text)
        
        # Format as tab-separated values with headers
        result = '\t'.join(headers) + '\n' + '\t'.join(row_data)
        
        # Copy to clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText(result)
    
    def copy_entire_table(self, tab_index):
        """Copy entire table with headers to clipboard"""
        if tab_index not in self.query_tabs:
            return
            
        # Get full data from complete query execution
        try:
            columns, full_data = self.execute_complete_query(tab_index)
            if not columns:
                return
                
            # Start with headers
            table_data = ['\t'.join(columns)]
            
            # Add all row data
            for row_data in full_data:
                formatted_row = []
                for cell_value in row_data:
                    cell_text = str(cell_value) if cell_value is not None else ''
                    formatted_row.append(cell_text)
                table_data.append('\t'.join(formatted_row))
            
            # Copy to clipboard
            clipboard = QApplication.clipboard()
            clipboard.setText('\n'.join(table_data))
            
        except Exception as e:
            QMessageBox.critical(self, 'Copy Error', f'Failed to copy table data:\n{str(e)}')

    def open_nodes_dashboard(self, tab_index):
        """Open the Nodes Dashboard (node.py) with the current query results"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        query = tab_data.get('current_query')
        
        if not query:
            QMessageBox.warning(self, 'No Query', 'No active query to visualize.')
            return
            
        try:
            # Create temp file path
            import tempfile
            import subprocess
            
            # Create a temporary file name (we close the file descriptor because DuckDB will open it)
            fd, temp_path = tempfile.mkstemp(suffix='.parquet')
            os.close(fd)
            
            # Clean query (remove trailing semicolon)
            clean_query = query.strip().rstrip(';')
            
            # Normalize path for DuckDB (replace backslashes with forward slashes to avoid escape issues)
            duckdb_path = temp_path.replace('\\', '/')
            
            # Export directly using DuckDB
            export_query = f"COPY ({clean_query}) TO '{duckdb_path}' (FORMAT PARQUET)"
            self.connection.execute(export_query)
            
            # Launch node.py
            app_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'node.py')
            
            # Run in a separate process
            subprocess.Popen([sys.executable, app_path, temp_path])
            
            self.status_label.setText(f'Opened Nodes Dashboard')
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to open Nodes Dashboard:\n{str(e)}')

    def open_main_dashboard(self, tab_index):
        """Open the Main Dashboard (dash.py) with the current query results"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        query = tab_data.get('current_query')
        
        if not query:
            QMessageBox.warning(self, 'No Query', 'No active query to visualize.')
            return
            
        try:
            # Create temp file path
            import tempfile
            import subprocess
            
            # Create a temporary file name (we close the file descriptor because DuckDB will open it)
            fd, temp_path = tempfile.mkstemp(suffix='.parquet')
            os.close(fd)
            
            # Clean query (remove trailing semicolon)
            clean_query = query.strip().rstrip(';')
            
            # Normalize path for DuckDB (replace backslashes with forward slashes to avoid escape issues)
            duckdb_path = temp_path.replace('\\', '/')
            
            # Export directly using DuckDB
            export_query = f"COPY ({clean_query}) TO '{duckdb_path}' (FORMAT PARQUET)"
            self.connection.execute(export_query)
            
            # Launch dash.py
            main_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dash.py')
            
            # Run in a separate process
            subprocess.Popen([sys.executable, main_path, temp_path])
            
            self.status_label.setText(f'Opened Dashboard')
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to open Dashboard:\n{str(e)}')
    
    def execute_complete_query(self, tab_index):
        """Execute the complete query without pagination to get all results"""
        if tab_index not in self.query_tabs:
            raise Exception("Invalid tab index")
            
        tab_data = self.query_tabs[tab_index]
        query = tab_data['current_query']
        
        if not query:
            raise Exception("No query to execute")
            
        if not self.connection:
            raise Exception("No database connection available")
            
        # Execute the complete query without pagination
        cursor = self.connection.execute(query)
        columns = [desc[0] for desc in cursor.description]
        
        # Fetch all data
        full_data = cursor.fetchall()
        
        return columns, full_data
    
    def load_saved_connections(self):
        """Load saved connections from file"""
        try:
            if os.path.exists(self.connections_file):
                with open(self.connections_file, 'r') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Error loading connections: {e}")
        return []
    
    def save_connections(self):
        """Save database connections to file"""
        try:
            with open('db_connections.json', 'w') as f:
                json.dump(self.saved_connections, f, indent=2)
        except Exception as e:
            print(f"Error saving connections: {e}")
    
    def save_query(self):
        """Save the current query with a name and description"""
        current_tab_index = self.query_tab_widget.currentIndex()
        if current_tab_index not in self.query_tabs:
            QMessageBox.warning(self, 'No Query', 'No active query tab found.')
            return
        
        tab_data = self.query_tabs[current_tab_index]
        query_text = tab_data['sql_editor'].toPlainText().strip()
        
        if not query_text:
            QMessageBox.warning(self, 'Empty Query', 'Please enter a SQL query to save.')
            return
        
        # Get query name from user
        name, ok = QInputDialog.getText(self, 'Save Query', 'Enter a name for this query:')
        if not ok or not name.strip():
            return
        
        # Get query description from user
        description, ok = QInputDialog.getText(self, 'Save Query', 'Enter a description (optional):')
        if not ok:
            description = ''
        
        # Load existing saved queries
        saved_queries = self.load_saved_queries()
        
        # Check if query name already exists
        if any(q['name'] == name.strip() for q in saved_queries):
            reply = QMessageBox.question(self, 'Query Exists', 
                                       f'A query named "{name.strip()}" already exists. Do you want to overwrite it?',
                                       QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes:
                return
            
            # Remove existing query with same name
            saved_queries = [q for q in saved_queries if q['name'] != name.strip()]
        
        # Add new query
        from datetime import datetime
        new_query = {
            'name': name.strip(),
            'description': description.strip(),
            'query': query_text,
            'date_saved': datetime.now().isoformat()
        }
        
        saved_queries.append(new_query)
        
        # Save to file
        try:
            with open('saved_queries.json', 'w') as f:
                json.dump(saved_queries, f, indent=2)
            QMessageBox.information(self, 'Query Saved', f'Query "{name.strip()}" has been saved successfully.')
        except Exception as e:
            QMessageBox.critical(self, 'Save Error', f'Failed to save query: {str(e)}')
    
    def load_query(self):
        """Load a saved query into the current tab"""
        saved_queries = self.load_saved_queries()
        
        if not saved_queries:
            QMessageBox.information(self, 'No Saved Queries', 'No saved queries found.')
            return
        
        # Create a simple selection dialog
        query_names = [f"{q['name']} - {q['description'][:50]}..." if len(q['description']) > 50 else f"{q['name']} - {q['description']}" for q in saved_queries]
        
        selected_query, ok = QInputDialog.getItem(self, 'Load Query', 'Select a query to load:', query_names, 0, False)
        if not ok:
            return
        
        # Find the selected query
        selected_index = query_names.index(selected_query)
        query_data = saved_queries[selected_index]
        
        # Load into current tab
        current_tab_index = self.query_tab_widget.currentIndex()
        if current_tab_index in self.query_tabs:
            tab_data = self.query_tabs[current_tab_index]
            tab_data['sql_editor'].setPlainText(query_data['query'])
            QMessageBox.information(self, 'Query Loaded', f'Query "{query_data["name"]}" has been loaded.')
        else:
            QMessageBox.warning(self, 'No Active Tab', 'No active query tab found.')
    
    def show_query_manager(self):
        """Show the query management dialog"""
        saved_queries = self.load_saved_queries()
        dialog = SavedQueryManagerDialog(self, saved_queries)
        if dialog.exec_() == QDialog.Accepted:
            # Refresh saved queries if any changes were made
            pass
    
    def load_saved_queries(self):
        """Load saved queries from JSON file"""
        try:
            if os.path.exists('saved_queries.json'):
                with open('saved_queries.json', 'r') as f:
                    return json.load(f)
            return []
        except Exception as e:
            print(f"Error loading saved queries: {e}")
            return []
    
    def add_custom_autocomplete_entries(self, entries):
        """Add custom auto-complete entries to all SQL editors"""
        if not isinstance(entries, (list, tuple)):
            return
        
        # Add to global list
        SQL_AUTOCOMPLETE_KEYWORDS.extend(entries)
        
        # Update all existing SQL editors in tabs
        for i in range(self.query_tab_widget.count()):
            tab_widget = self.query_tab_widget.widget(i)
            if hasattr(tab_widget, 'findChild'):
                sql_editor = tab_widget.findChild(SQLTextEdit)
                if sql_editor:
                    sql_editor.add_custom_completions(entries)
    
    def get_current_sql_editor(self):
        """Get the SQL editor from the current tab"""
        current_tab = self.query_tab_widget.currentWidget()
        if current_tab and hasattr(current_tab, 'findChild'):
            return current_tab.findChild(SQLTextEdit)
        return None
    
    def closeEvent(self, event):
        """Clean up when closing the application"""
        if hasattr(self, 'query_thread') and self.query_thread.isRunning():
            self.query_thread.terminate()
            self.query_thread.wait()
        
        self.connection.close()
        event.accept()

    def set_theme(self, theme_name):
        """Set the application theme"""
        themes = {
            'light': {
                'background': '#ffffff',
                'text': '#000000',
                'button': '#f0f0f0',
                'button_hover': '#e0e0e0',
                'input': '#ffffff',
                'border': '#cccccc',
                'selection': '#0078d4',
                'menu': '#ffffff',
                'menu_hover': '#e0e0e0'
            },
            'dark': {
                'background': '#2b2b2b',
                'text': '#ffffff',
                'button': '#404040',
                'button_hover': '#505050',
                'input': '#3c3c3c',
                'border': '#555555',
                'selection': '#0078d4',
                'menu': '#2b2b2b',
                'menu_hover': '#404040'
            },
            'blue': {
                'background': '#1e3a5f',
                'text': '#ffffff',
                'button': '#2c5282',
                'button_hover': '#3c6382',
                'input': '#2a4a6b',
                'border': '#4a6fa5',
                'selection': '#63b3ed',
                'menu': '#1e3a5f',
                'menu_hover': '#2c5282'
            },
            'green': {
                'background': '#1a4d3a',
                'text': '#ffffff',
                'button': '#2d7d32',
                'button_hover': '#388e3c',
                'input': '#2e5d3e',
                'border': '#4caf50',
                'selection': '#81c784',
                'menu': '#1a4d3a',
                'menu_hover': '#2d7d32'
            },
            'high_contrast': {
                'background': '#000000',
                'text': '#ffffff',
                'button': '#ffffff',
                'button_hover': '#ffff00',
                'button_text': '#000000',
                'button_hover_text': '#000000',
                'input': '#000000',
                'input_text': '#ffffff',
                'border': '#ffffff',
                'selection': '#ffff00',
                'selection_text': '#000000',
                'menu': '#000000',
                'menu_hover': '#ffff00',
                'menu_hover_text': '#000000'
            }
        }
        
        if theme_name not in themes:
            return
            
        theme = themes[theme_name]
        
        # Apply theme stylesheet
        # Handle high contrast theme specific colors
        button_text = theme.get('button_text', theme['text'])
        button_hover_text = theme.get('button_hover_text', theme['text'])
        input_text = theme.get('input_text', theme['text'])
        selection_text = theme.get('selection_text', theme['text'])
        menu_hover_text = theme.get('menu_hover_text', theme['text'])
        
        stylesheet = f"""
        QMainWindow {{
            background-color: {theme['background']};
            color: {theme['text']};
        }}
        
        QWidget {{
            background-color: {theme['background']};
            color: {theme['text']};
        }}
        
        QPushButton {{
            background-color: {theme['button']};
            color: {button_text};
            border: 2px solid {theme['border']};
            padding: 5px 10px;
            border-radius: 3px;
            font-weight: bold;
        }}
        
        QPushButton:hover {{
            background-color: {theme['button_hover']};
            color: {button_hover_text};
            border: 2px solid {theme['border']};
        }}
        
        QPushButton:pressed {{
            background-color: {theme['border']};
            color: {button_hover_text};
        }}
        
        QLineEdit, QTextEdit, QPlainTextEdit {{
            background-color: {theme['input']};
            color: {input_text};
            border: 2px solid {theme['border']};
            padding: 3px;
            border-radius: 2px;
        }}
        
        QTreeWidget {{
            background-color: {theme['input']};
            color: {input_text};
            border: 2px solid {theme['border']};
            selection-background-color: {theme['selection']};
            selection-color: {selection_text};
        }}
        
        QTableWidget {{
            background-color: {theme['input']};
            color: {input_text};
            border: 2px solid {theme['border']};
            gridline-color: {theme['border']};
            selection-background-color: {theme['selection']};
            selection-color: {selection_text};
        }}
        
        QHeaderView::section {{
            background-color: {theme['button']};
            color: {button_text};
            border: 2px solid {theme['border']};
            padding: 3px;
            font-weight: bold;
        }}
        
        QMenuBar {{
            background-color: {theme['menu']};
            color: {theme['text']};
            border-bottom: 2px solid {theme['border']};
        }}
        
        QMenuBar::item {{
            background-color: transparent;
            padding: 4px 8px;
            color: {theme['text']};
        }}
        
        QMenuBar::item:selected {{
            background-color: {theme['menu_hover']};
            color: {menu_hover_text};
        }}
        
        QMenu {{
            background-color: {theme['menu']};
            color: {theme['text']};
            border: 2px solid {theme['border']};
        }}
        
        QMenu::item {{
            padding: 4px 20px;
            color: {theme['text']};
        }}
        
        QMenu::item:selected {{
            background-color: {theme['menu_hover']};
            color: {menu_hover_text};
        }}
        
        QGroupBox {{
            color: {theme['text']};
            border: 2px solid {theme['border']};
            border-radius: 3px;
            margin-top: 10px;
            font-weight: bold;
        }}
        
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px 0 5px;
            color: {theme['text']};
        }}
        
        QLabel {{
            color: {theme['text']};
            font-weight: bold;
        }}
        
        QComboBox {{
            background-color: {theme['input']};
            color: {input_text};
            border: 2px solid {theme['border']};
            padding: 3px;
            border-radius: 2px;
            font-weight: bold;
        }}
        
        QComboBox::drop-down {{
            border: none;
            background-color: {theme['button']};
        }}
        
        QComboBox::down-arrow {{
            border: none;
            color: {button_text};
        }}
        
        QScrollBar:vertical {{
            background-color: {theme['background']};
            width: 15px;
            border: 2px solid {theme['border']};
        }}
        
        QScrollBar::handle:vertical {{
            background-color: {theme['button']};
            border-radius: 3px;
            border: 1px solid {theme['border']};
            min-height: 20px;
        }}
        
        QScrollBar::handle:vertical:hover {{
            background-color: {theme['button_hover']};
        }}
        
        QScrollBar::add-line:vertical {{
            background-color: {theme['button']};
            height: 15px;
            border: 1px solid {theme['border']};
            subcontrol-position: bottom;
            subcontrol-origin: margin;
        }}
        
        QScrollBar::sub-line:vertical {{
            background-color: {theme['button']};
            height: 15px;
            border: 1px solid {theme['border']};
            subcontrol-position: top;
            subcontrol-origin: margin;
        }}
        
        QScrollBar::add-line:vertical:hover, QScrollBar::sub-line:vertical:hover {{
            background-color: {theme['button_hover']};
        }}
        
        QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {{
            width: 8px;
            height: 8px;
            background-color: {button_text};
        }}
        
        QScrollBar:horizontal {{
            background-color: {theme['background']};
            height: 15px;
            border: 2px solid {theme['border']};
        }}
        
        QScrollBar::handle:horizontal {{
            background-color: {theme['button']};
            border-radius: 3px;
            border: 1px solid {theme['border']};
            min-width: 20px;
        }}
        
        QScrollBar::handle:horizontal:hover {{
            background-color: {theme['button_hover']};
        }}
        
        QScrollBar::add-line:horizontal {{
            background-color: {theme['button']};
            width: 15px;
            border: 1px solid {theme['border']};
            subcontrol-position: right;
            subcontrol-origin: margin;
        }}
        
        QScrollBar::sub-line:horizontal {{
            background-color: {theme['button']};
            width: 15px;
            border: 1px solid {theme['border']};
            subcontrol-position: left;
            subcontrol-origin: margin;
        }}
        
        QScrollBar::add-line:horizontal:hover, QScrollBar::sub-line:horizontal:hover {{
            background-color: {theme['button_hover']};
        }}
        
        QScrollBar::left-arrow:horizontal, QScrollBar::right-arrow:horizontal {{
            width: 8px;
            height: 8px;
            background-color: {button_text};
        }}
        
        QDialog {{
            background-color: {theme['background']};
            color: {theme['text']};
            border: 2px solid {theme['border']};
        }}
        
        QRadioButton {{
            color: {theme['text']};
            font-weight: bold;
        }}
        
        QRadioButton::indicator {{
            width: 15px;
            height: 15px;
            border: 2px solid {theme['border']};
            border-radius: 8px;
            background-color: {theme['input']};
        }}
        
        QRadioButton::indicator:checked {{
            background-color: {theme['selection']};
            border: 2px solid {theme['border']};
        }}
        
        QTabWidget::pane {{
            border: 2px solid {theme['border']};
            background-color: {theme['background']};
        }}
        
        QTabBar::tab {{
            background-color: {theme['button']};
            color: {button_text};
            border: 2px solid {theme['border']};
            padding: 8px 12px;
            margin-right: 2px;
            border-bottom: none;
            font-weight: bold;
            min-width: 80px;
            max-width: 200px;
        }}
        
        QTabBar::tab:selected {{
            background-color: {theme['background']};
            color: {theme['text']};
            border-bottom: 2px solid {theme['background']};
        }}
        
        QTabBar::tab:hover {{
            background-color: {theme['button_hover']};
            color: {button_hover_text};
        }}
        
        QTabBar::close-button {{
            background-color: {theme['button']};
            color: {button_text};
            border: 1px solid {theme['border']};
            border-radius: 2px;
            width: 16px;
            height: 16px;
            margin: 2px;
            font-weight: bold;
            font-size: 12px;
            subcontrol-position: right;
        }}
        
        QTabBar::close-button:hover {{
            background-color: {theme['selection']};
            color: {selection_text};
            border: 1px solid {theme['text']};
        }}
        
        QTabBar::close-button:pressed {{
            background-color: {theme['border']};
            color: {button_hover_text};
        }}
        """
        
        self.setStyleSheet(stylesheet)
        
        # Update theme action states
        for action in self.theme_group.actions():
            action.setChecked(False)
            
        if theme_name == 'light':
            self.light_theme_action.setChecked(True)
        elif theme_name == 'dark':
            self.dark_theme_action.setChecked(True)
        elif theme_name == 'blue':
            self.blue_theme_action.setChecked(True)
        elif theme_name == 'green':
            self.green_theme_action.setChecked(True)
        elif theme_name == 'high_contrast':
            self.high_contrast_theme_action.setChecked(True)
        
        # Save theme preference
        self.current_theme = theme_name
        self.save_theme_preference()
    
    def save_theme_preference(self):
        """Save the current theme preference to a file"""
        try:
            import json
            import os
            
            config_dir = os.path.expanduser('~/.duckdb_sql_app')
            os.makedirs(config_dir, exist_ok=True)
            
            config_file = os.path.join(config_dir, 'config.json')
            config = {}
            
            if os.path.exists(config_file):
                with open(config_file, 'r') as f:
                    config = json.load(f)
            
            config['theme'] = self.current_theme
            
            with open(config_file, 'w') as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Error saving theme preference: {e}")
    
    def load_theme_preference(self):
        """Load the saved theme preference"""
        try:
            import json
            import os
            
            config_file = os.path.expanduser('~/.duckdb_sql_app/config.json')
            
            if os.path.exists(config_file):
                with open(config_file, 'r') as f:
                    config = json.load(f)
                    
                theme = config.get('theme', 'light')
                self.set_theme(theme)
            else:
                self.set_theme('light')
        except Exception as e:
            print(f"Error loading theme preference: {e}")
            self.set_theme('light')
    

    

    

    


    def add_new_query_tab(self):
        """Add a new query tab with SQL editor and results table"""
        self.tab_counter += 1
        tab_name = f"Query {self.tab_counter}"
        
        # Create tab widget
        tab_widget = QWidget()
        tab_layout = QVBoxLayout(tab_widget)
        
        # Create splitter for SQL editor and results
        splitter = QSplitter(Qt.Vertical)
        
        # SQL Editor section
        sql_group = QGroupBox('SQL Query Editor')
        sql_layout = QVBoxLayout(sql_group)
        
        sql_editor = SQLTextEdit(self)
        sql_editor.setFont(QFont('Consolas', 10))
        sql_editor.setPlaceholderText('Enter your SQL query here...\nExample: SELECT * FROM localdb.your_table_name LIMIT 10;')
        
        # Update autocomplete with current table names
        self.update_autocomplete_for_editor(sql_editor)
        
        # Button layout for Execute and New Query buttons
        button_layout = QHBoxLayout()
        execute_btn = QPushButton('Execute Query')
        new_query_btn = QPushButton('New Query')
        split_screen_btn = QPushButton('Split Screen')
        
        # Export buttons
        export_excel_btn = QPushButton('Export to Excel')
        export_csv_btn = QPushButton('Export to CSV')
        export_json_btn = QPushButton('Export to JSON')
        export_parquet_btn = QPushButton('Export to Parquet')

        
        execute_btn.clicked.connect(lambda: self.execute_query_for_tab(self.query_tab_widget.currentIndex()))
        new_query_btn.clicked.connect(self.add_new_query_tab)
        split_screen_btn.clicked.connect(self.toggle_split_screen)
        export_excel_btn.clicked.connect(lambda: self.export_results('excel', self.query_tab_widget.currentIndex()))
        export_csv_btn.clicked.connect(lambda: self.export_results('csv', self.query_tab_widget.currentIndex()))
        export_json_btn.clicked.connect(lambda: self.export_results('json', self.query_tab_widget.currentIndex()))
        export_parquet_btn.clicked.connect(lambda: self.export_results('parquet', self.query_tab_widget.currentIndex()))

        
        button_layout.addWidget(execute_btn)
        button_layout.addWidget(new_query_btn)
        button_layout.addWidget(split_screen_btn)
        button_layout.addWidget(export_excel_btn)
        button_layout.addWidget(export_csv_btn)
        button_layout.addWidget(export_json_btn)
        button_layout.addWidget(export_parquet_btn)

        button_layout.addStretch()
        
        sql_layout.addWidget(sql_editor)
        sql_layout.addLayout(button_layout)
        
        splitter.addWidget(sql_group)
        
        # Results section
        results_group = QGroupBox('Query Results')
        results_layout = QVBoxLayout(results_group)
        
        # Pagination controls
        pagination_layout = QHBoxLayout()
        
        # Page info and controls
        page_info_label = QLabel('No results')
        first_page_btn = QPushButton('First')
        prev_page_btn = QPushButton('Previous')
        next_page_btn = QPushButton('Next')
        last_page_btn = QPushButton('Last')
        
        # Page size selector
        page_size_label = QLabel('Rows per page:')
        page_size_combo = QComboBox()
        page_size_combo.addItems(['1000', '5000', '10000', '25000', '50000'])
        page_size_combo.setCurrentText('10000')
        
        # Cancel query button
        cancel_btn = QPushButton('Cancel Query')
        cancel_btn.setEnabled(False)
        
        # Progress bar
        progress_bar = QProgressBar()
        progress_bar.setVisible(False)
        progress_bar.setMaximum(100)
        
        pagination_layout.addWidget(page_info_label)
        pagination_layout.addStretch()
        pagination_layout.addWidget(first_page_btn)
        pagination_layout.addWidget(prev_page_btn)
        pagination_layout.addWidget(next_page_btn)
        pagination_layout.addWidget(last_page_btn)
        pagination_layout.addStretch()
        pagination_layout.addWidget(page_size_label)
        pagination_layout.addWidget(page_size_combo)
        pagination_layout.addWidget(cancel_btn)
        
        # Add progress bar below pagination controls
        results_layout.addWidget(progress_bar)
        
        results_table = QTableWidget()
        
        results_layout.addLayout(pagination_layout)
        results_layout.addWidget(results_table)
        
        splitter.addWidget(results_group)
        
        # Set splitter proportions
        splitter.setSizes([300, 500])
        
        tab_layout.addWidget(splitter)
        
        # Add tab to tab widget
        tab_index = self.query_tab_widget.addTab(tab_widget, tab_name)
        
        # Set up context menu for results table (after tab_index is defined)
        results_table.setContextMenuPolicy(Qt.CustomContextMenu)
        results_table.customContextMenuRequested.connect(lambda pos, tab_idx=tab_index: self.show_results_context_menu(pos, tab_idx))
        
        # Set up context menu for table headers
        results_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        results_table.horizontalHeader().customContextMenuRequested.connect(lambda pos, tab_idx=tab_index: self.show_header_context_menu(pos, tab_idx))
        
        # Create custom close button for this tab
        close_button = QPushButton('×')
        close_button.setFixedSize(16, 16)
        close_button.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                border: none;
                color: {self.get_current_theme_color('text')};
                font-weight: bold;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: red;
                color: white;
                border-radius: 8px;
            }}
        """)
        close_button.setToolTip('Close tab')
        close_button.clicked.connect(lambda: self.close_query_tab(tab_index))
        
        # Add close button to tab
        self.query_tab_widget.tabBar().setTabButton(tab_index, self.query_tab_widget.tabBar().RightSide, close_button)
        
        # Connect pagination controls
        first_page_btn.clicked.connect(lambda: self.go_to_page(tab_index, 0))
        prev_page_btn.clicked.connect(lambda: self.prev_page(tab_index))
        next_page_btn.clicked.connect(lambda: self.next_page(tab_index))
        last_page_btn.clicked.connect(lambda: self.go_to_last_page(tab_index))
        page_size_combo.currentTextChanged.connect(lambda: self.change_page_size(tab_index))
        cancel_btn.clicked.connect(lambda: self.cancel_query(tab_index))
        
        # Store tab components
        self.query_tabs[tab_index] = {
            'sql_editor': sql_editor,
            'results_table': results_table,
            'query_thread': None,
            'streaming_thread': None,
            'close_button': close_button,
            'page_info_label': page_info_label,
            'first_page_btn': first_page_btn,
            'prev_page_btn': prev_page_btn,
            'next_page_btn': next_page_btn,
            'last_page_btn': last_page_btn,
            'page_size_combo': page_size_combo,
            'cancel_btn': cancel_btn,
            'progress_bar': progress_bar,
            'current_page': 0,
            'total_rows': 0,
            'current_query': '',
            'columns': []
        }
        
        # Switch to new tab
        self.query_tab_widget.setCurrentIndex(tab_index)
        
        return tab_index
    
    def get_current_theme_color(self, color_type):
        """Get color from current theme"""
        if hasattr(self, 'themes') and hasattr(self, 'current_theme'):
            theme = self.themes[self.current_theme]
            return theme.get(color_type, '#000000')
        else:
            # Default colors when themes aren't loaded yet
            defaults = {
                'text': '#000000',
                'background': '#ffffff',
                'button': '#f0f0f0'
            }
            return defaults.get(color_type, '#000000')
    
    def close_query_tab(self, index):
        """Close a query tab with comprehensive memory cleanup"""
        if self.query_tab_widget.count() <= 1:
            # Don't close the last tab
            return
            
        # Stop any running queries and clean up threads
        if index in self.query_tabs:
            tab_data = self.query_tabs[index]
            
            # Stop streaming thread
            if tab_data.get('streaming_thread'):
                tab_data['streaming_thread'].cancel()
                tab_data['streaming_thread'].wait()
                tab_data['streaming_thread'] = None
                
            # Stop query thread
            if tab_data.get('query_thread'):
                tab_data['query_thread'].terminate()
                tab_data['query_thread'].wait()
                tab_data['query_thread'] = None
            
            # Clear table data to free memory
            if 'results_table' in tab_data:
                tab_data['results_table'].clearContents()
                tab_data['results_table'].setRowCount(0)
                tab_data['results_table'].setColumnCount(0)
            
            # Clear stored data
            tab_data.clear()
            del self.query_tabs[index]
            
        # Remove the tab
        self.query_tab_widget.removeTab(index)
        
        # Update tab indices in query_tabs dict
        new_query_tabs = {}
        for i in range(self.query_tab_widget.count()):
            old_index = list(self.query_tabs.keys())[i] if i < len(self.query_tabs) else i
            if old_index in self.query_tabs:
                new_query_tabs[i] = self.query_tabs[old_index]
        self.query_tabs = new_query_tabs
        
        # Force garbage collection to free memory
        gc.collect()
    
    def execute_query_for_tab(self, tab_index):
        """Execute the SQL query for a specific tab using streaming"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        sql_editor = tab_data['sql_editor']
        
        # Check if there's selected text first
        cursor = sql_editor.textCursor()
        if cursor.hasSelection():
            # Fix for comments: Get selected text properly preserving line breaks
            # cursor.selectedText() converts line breaks to \u2029 which breaks SQL comments
            start_pos = cursor.selectionStart()
            end_pos = cursor.selectionEnd()
            full_text = sql_editor.toPlainText()
            query_text = full_text[start_pos:end_pos].strip()
        else:
            query_text = sql_editor.toPlainText().strip()
        
        if not query_text:
            QMessageBox.warning(self, 'Empty Query', 'Please enter a SQL query.')
            return
        
        # Check if we have any data available (local tables or remote connections)
        if not self.loaded_tables and not self.active_connections:
            QMessageBox.warning(self, 'No Data', 'Please load a data file or connect to a database first.')
            return
        
        # Split queries by semicolon (handle multiple statements)
        queries = self.split_sql_statements(query_text)
        
        if len(queries) > 1:
            # Multiple queries - execute them sequentially and show multiple results
            self.execute_multiple_queries(tab_index, queries)
        else:
            # Single query - use existing streaming approach
            tab_data['current_query'] = queries[0]
            tab_data['current_page'] = 0
            tab_data['total_rows'] = 0
            
            # Execute streaming query
            self.execute_streaming_query(tab_index)
    
    def split_sql_statements(self, query_text):
        """Split SQL text into individual statements, handling strings and comments"""
        statements = []
        current_statement = []
        in_string = False
        in_comment = False
        string_char = None
        
        lines = query_text.split('\n')
        
        for line in lines:
            i = 0
            while i < len(line):
                char = line[i]
                
                # Handle line comments
                if not in_string and i < len(line) - 1 and line[i:i+2] == '--':
                    # Rest of line is comment
                    current_statement.append(line[i:])
                    break
                
                # Handle string literals
                if char in ('"', "'") and not in_comment:
                    if not in_string:
                        in_string = True
                        string_char = char
                    elif char == string_char:
                        # Check if it's escaped
                        if i > 0 and line[i-1] == '\\':
                            pass  # Escaped quote
                        else:
                            in_string = False
                            string_char = None
                
                # Handle semicolon (statement separator)
                if char == ';' and not in_string and not in_comment:
                    current_statement.append(char)
                    stmt = ''.join(current_statement).strip()
                    if stmt and stmt != ';':
                        statements.append(stmt.rstrip(';').strip())
                    current_statement = []
                    i += 1
                    continue
                
                current_statement.append(char)
                i += 1
            
            # Add newline if not at end
            if current_statement:
                current_statement.append('\n')
        
        # Add remaining statement
        stmt = ''.join(current_statement).strip()
        if stmt and stmt != ';':
            statements.append(stmt.rstrip(';').strip())
        
        return [s for s in statements if s]  # Filter empty statements
    
    def execute_multiple_queries(self, tab_index, queries):
        """Execute multiple SQL queries and display results in tabs"""
        if tab_index not in self.query_tabs:
            return
        
        tab_data = self.query_tabs[tab_index]
        
        # Hide the original single-result widgets
        tab_data['results_table'].setVisible(False)
        tab_data['progress_bar'].setVisible(False)
        tab_data['page_info_label'].setVisible(False)
        tab_data['first_page_btn'].setVisible(False)
        tab_data['prev_page_btn'].setVisible(False)
        tab_data['next_page_btn'].setVisible(False)
        tab_data['last_page_btn'].setVisible(False)
        tab_data['page_size_combo'].setVisible(False)
        tab_data['cancel_btn'].setVisible(False)
        
        # Find the results group widget
        results_group = tab_data['results_table'].parent()
        while results_group and not isinstance(results_group, QGroupBox):
            results_group = results_group.parent()
        
        if not results_group:
            return
        
        # Get the layout
        results_layout = results_group.layout()
        
        # Check if we already have a multi-query tab widget
        multi_query_tab_widget = None
        for i in range(results_layout.count()):
            widget = results_layout.itemAt(i).widget()
            if isinstance(widget, QTabWidget) and hasattr(widget, '_is_multi_query_widget'):
                multi_query_tab_widget = widget
                break
        
        # If no multi-query widget exists, create one
        if not multi_query_tab_widget:
            multi_query_tab_widget = QTabWidget()
            multi_query_tab_widget._is_multi_query_widget = True
            multi_query_tab_widget.setTabPosition(QTabWidget.South)
            multi_query_tab_widget.setMovable(True)
            
            # Style the tabs to be compact
            multi_query_tab_widget.setStyleSheet("""
                QTabWidget::pane {
                    border: 1px solid #cccccc;
                    background: white;
                }
                QTabBar::tab {
                    background: #e0e0e0;
                    border: 1px solid #cccccc;
                    padding: 4px 10px;
                    margin-right: 2px;
                    border-bottom: none;
                    min-width: 70px;
                    max-width: 150px;
                    font-size: 11px;
                }
                QTabBar::tab:selected {
                    background: white;
                    border-bottom: 2px solid #3498db;
                    font-weight: bold;
                    color: #2c3e50;
                }
                QTabBar::tab:hover {
                    background: #d0d0d0;
                }
            """)
            
            results_layout.addWidget(multi_query_tab_widget)
            multi_query_tab_widget.setVisible(True)
        else:
            # Clear existing tabs
            multi_query_tab_widget.clear()
            multi_query_tab_widget.setVisible(True)
        
        # Update title
        results_group.setTitle(f'Query Results ({len(queries)} statements)')
        
        # Store reference to multi-query widget
        tab_data['multi_query_widget'] = multi_query_tab_widget
        
        # Execute each query and create a tab for each result
        for idx, query in enumerate(queries, 1):
            # Create widget for this result
            result_widget = QWidget()
            result_layout = QVBoxLayout(result_widget)
            result_layout.setContentsMargins(5, 5, 5, 5)
            
            # Pagination controls (same style as single query)
            pagination_layout = QHBoxLayout()
            
            # Page info and controls
            page_info_label = QLabel('No results')
            first_page_btn = QPushButton('First')
            prev_page_btn = QPushButton('Previous')
            next_page_btn = QPushButton('Next')
            last_page_btn = QPushButton('Last')
            
            # Page size selector
            page_size_label = QLabel('Rows per page:')
            page_size_combo = QComboBox()
            page_size_combo.addItems(['1000', '5000', '10000', '25000', '50000'])
            page_size_combo.setCurrentText('10000')
            
            # Progress bar
            progress_bar = QProgressBar()
            progress_bar.setVisible(False)
            progress_bar.setMaximum(100)
            
            pagination_layout.addWidget(page_info_label)
            pagination_layout.addStretch()
            pagination_layout.addWidget(first_page_btn)
            pagination_layout.addWidget(prev_page_btn)
            pagination_layout.addWidget(next_page_btn)
            pagination_layout.addWidget(last_page_btn)
            pagination_layout.addStretch()
            pagination_layout.addWidget(page_size_label)
            pagination_layout.addWidget(page_size_combo)
            
            result_layout.addWidget(progress_bar)
            result_layout.addLayout(pagination_layout)
            
            # Create result table for this query
            result_table = QTableWidget()
            result_table.setAlternatingRowColors(False)
            result_table.setSortingEnabled(False)
            
            # Set selection behavior to select entire rows (like single query mode)
            result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
            
            # Set up context menu for the result table
            result_table.setContextMenuPolicy(Qt.CustomContextMenu)
            
            result_layout.addWidget(result_table)
            
            # Store pagination state for this tab
            tab_state = {
                'query': query,
                'current_page': 0,
                'total_rows': 0,
                'columns': [],
                'all_data': [],
                'page_info_label': page_info_label,
                'first_page_btn': first_page_btn,
                'prev_page_btn': prev_page_btn,
                'next_page_btn': next_page_btn,
                'last_page_btn': last_page_btn,
                'page_size_combo': page_size_combo,
                'result_table': result_table,
                'progress_bar': progress_bar,
                'tab_index': tab_index,
                'result_index': idx
            }
            
            # Connect context menu
            result_table.customContextMenuRequested.connect(
                lambda pos, ts=tab_state: self.show_multi_query_context_menu(pos, ts)
            )
            
            # Create compact tab label
            query_preview = query.replace('\n', ' ').strip()
            tab_label = f'Result {idx}'
            
            # Add tab with tooltip showing full query
            tab_idx = multi_query_tab_widget.addTab(result_widget, tab_label)
            multi_query_tab_widget.setTabToolTip(tab_idx, query_preview)
            
            # Execute the query
            try:
                self.status_label.setText(f'Executing statement {idx} of {len(queries)}...')
                QApplication.processEvents()
                
                cursor = self.connection.execute(query)
                
                # Check if query returns results
                if cursor.description:
                    columns = [desc[0] for desc in cursor.description]
                    all_data = cursor.fetchall()
                    
                    # Store data in tab state
                    tab_state['columns'] = columns
                    tab_state['all_data'] = all_data
                    tab_state['total_rows'] = len(all_data)
                    
                    # Display first page
                    page_size = int(page_size_combo.currentText())
                    self.display_multi_query_page(tab_state, 0, page_size)
                    
                    # Connect pagination controls
                    first_page_btn.clicked.connect(lambda checked, ts=tab_state: self.multi_query_first_page(ts))
                    prev_page_btn.clicked.connect(lambda checked, ts=tab_state: self.multi_query_prev_page(ts))
                    next_page_btn.clicked.connect(lambda checked, ts=tab_state: self.multi_query_next_page(ts))
                    last_page_btn.clicked.connect(lambda checked, ts=tab_state: self.multi_query_last_page(ts))
                    page_size_combo.currentTextChanged.connect(lambda text, ts=tab_state: self.multi_query_change_page_size(ts))
                    
                    # Update tab label with row count
                    multi_query_tab_widget.setTabText(tab_idx, f'Result {idx} ({len(all_data)} rows)')
                else:
                    # Query doesn't return results (INSERT, UPDATE, DELETE, etc.)
                    result_table.setVisible(False)
                    page_info_label.setText('✓ Query executed successfully (no results returned)')
                    page_info_label.setStyleSheet('color: green; font-weight: bold;')
                    
                    # Hide pagination controls
                    first_page_btn.setVisible(False)
                    prev_page_btn.setVisible(False)
                    next_page_btn.setVisible(False)
                    last_page_btn.setVisible(False)
                    page_size_label.setVisible(False)
                    page_size_combo.setVisible(False)
                    
                    multi_query_tab_widget.setTabText(tab_idx, f'Result {idx} (OK)')
                    
            except Exception as e:
                result_table.setVisible(False)
                error_text = str(e)
                page_info_label.setText(f'✗ Error: {error_text}')
                page_info_label.setStyleSheet('color: red; font-weight: bold;')
                page_info_label.setWordWrap(True)
                
                # Hide pagination controls
                first_page_btn.setVisible(False)
                prev_page_btn.setVisible(False)
                next_page_btn.setVisible(False)
                last_page_btn.setVisible(False)
                page_size_label.setVisible(False)
                page_size_combo.setVisible(False)
                
                multi_query_tab_widget.setTabText(tab_idx, f'Result {idx} (Error)')
                multi_query_tab_widget.setTabIcon(tab_idx, self.style().standardIcon(QStyle.SP_MessageBoxCritical))
        
        self.status_label.setText(f'Executed {len(queries)} statements')
    
    def display_multi_query_page(self, tab_state, page_num, page_size):
        """Display a specific page of data for a multi-query result tab"""
        all_data = tab_state['all_data']
        columns = tab_state['columns']
        result_table = tab_state['result_table']
        page_info_label = tab_state['page_info_label']
        
        total_rows = len(all_data)
        total_pages = (total_rows + page_size - 1) // page_size if page_size > 0 else 1
        
        # Clamp page number
        page_num = max(0, min(page_num, total_pages - 1))
        tab_state['current_page'] = page_num
        
        # Calculate slice
        start_idx = page_num * page_size
        end_idx = min(start_idx + page_size, total_rows)
        page_data = all_data[start_idx:end_idx]
        
        # Populate table
        result_table.setColumnCount(len(columns))
        result_table.setHorizontalHeaderLabels(columns)
        result_table.setRowCount(len(page_data))
        
        for row_idx, row_data in enumerate(page_data):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value) if value is not None else '')
                result_table.setItem(row_idx, col_idx, item)
        
        # Resize columns intelligently
        if len(page_data) < 1000:
            result_table.resizeColumnsToContents()
            for col in range(result_table.columnCount()):
                if result_table.columnWidth(col) > 300:
                    result_table.setColumnWidth(col, 300)
        else:
            result_table.horizontalHeader().setDefaultSectionSize(120)
        
        # Update page info
        page_info_label.setText(f'Page {page_num + 1} of {total_pages} ({total_rows} total rows)')
        
        # Update button states
        tab_state['first_page_btn'].setEnabled(page_num > 0)
        tab_state['prev_page_btn'].setEnabled(page_num > 0)
        tab_state['next_page_btn'].setEnabled(page_num < total_pages - 1)
        tab_state['last_page_btn'].setEnabled(page_num < total_pages - 1)
    
    def multi_query_first_page(self, tab_state):
        """Go to first page in multi-query result"""
        page_size = int(tab_state['page_size_combo'].currentText())
        self.display_multi_query_page(tab_state, 0, page_size)
    
    def multi_query_prev_page(self, tab_state):
        """Go to previous page in multi-query result"""
        page_size = int(tab_state['page_size_combo'].currentText())
        current_page = tab_state['current_page']
        self.display_multi_query_page(tab_state, current_page - 1, page_size)
    
    def multi_query_next_page(self, tab_state):
        """Go to next page in multi-query result"""
        page_size = int(tab_state['page_size_combo'].currentText())
        current_page = tab_state['current_page']
        self.display_multi_query_page(tab_state, current_page + 1, page_size)
    
    def multi_query_last_page(self, tab_state):
        """Go to last page in multi-query result"""
        page_size = int(tab_state['page_size_combo'].currentText())
        total_rows = tab_state['total_rows']
        last_page = (total_rows - 1) // page_size if page_size > 0 else 0
        self.display_multi_query_page(tab_state, last_page, page_size)
    
    def multi_query_change_page_size(self, tab_state):
        """Handle page size change in multi-query result"""
        page_size = int(tab_state['page_size_combo'].currentText())
        # Reset to first page when changing page size
        self.display_multi_query_page(tab_state, 0, page_size)
    
    def show_multi_query_context_menu(self, pos, tab_state):
        """Show context menu for multi-query result table"""
        result_table = tab_state['result_table']
        
        # Create context menu
        menu = QMenu(self)
        
        # Get current cell/selection
        current_item = result_table.itemAt(pos)
        has_selection = result_table.selectionModel().hasSelection()
        
        # Copy Cell Value
        copy_cell_action = menu.addAction('Copy Cell Value')
        copy_cell_action.setEnabled(current_item is not None)
        copy_cell_action.triggered.connect(lambda: self.copy_multi_query_cell(tab_state))
        
        # Copy Column with Header
        copy_column_action = menu.addAction('Copy Column with Header')
        copy_column_action.setEnabled(current_item is not None)
        copy_column_action.triggered.connect(lambda: self.copy_multi_query_column(tab_state))
        
        # Copy Row with Header
        copy_row_action = menu.addAction('Copy Row with Header')
        copy_row_action.setEnabled(has_selection)
        copy_row_action.triggered.connect(lambda: self.copy_multi_query_row(tab_state))
        
        menu.addSeparator()
        
        # Copy Entire Table
        copy_table_action = menu.addAction('Copy Entire Table')
        copy_table_action.setEnabled(result_table.rowCount() > 0)
        copy_table_action.triggered.connect(lambda: self.copy_multi_query_entire_table(tab_state))
        
        menu.addSeparator()
        
        # Graph Data
        graph_action = menu.addAction('Graph Data')
        graph_action.setEnabled(result_table.rowCount() > 0 and PLOTLY_AVAILABLE)
        graph_action.triggered.connect(lambda: self.graph_multi_query_data(tab_state))
        
        # Show menu at cursor position
        menu.exec_(result_table.viewport().mapToGlobal(pos))
    
    def copy_multi_query_cell(self, tab_state):
        """Copy selected cell value from multi-query result"""
        result_table = tab_state['result_table']
        current_item = result_table.currentItem()
        if current_item:
            QApplication.clipboard().setText(current_item.text())
    
    def copy_multi_query_column(self, tab_state):
        """Copy entire column with header from multi-query result"""
        result_table = tab_state['result_table']
        current_column = result_table.currentColumn()
        
        if current_column < 0:
            return
        
        # Get header
        header_item = result_table.horizontalHeaderItem(current_column)
        header = header_item.text() if header_item else f"Column_{current_column}"
        
        # Get all values in column
        values = [header]
        for row in range(result_table.rowCount()):
            item = result_table.item(row, current_column)
            values.append(item.text() if item else '')
        
        QApplication.clipboard().setText('\n'.join(values))
    
    def copy_multi_query_row(self, tab_state):
        """Copy selected row with headers from multi-query result"""
        result_table = tab_state['result_table']
        current_row = result_table.currentRow()
        
        if current_row < 0:
            return
        
        # Get headers
        headers = []
        for col in range(result_table.columnCount()):
            header_item = result_table.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else f"Column_{col}")
        
        # Get row values
        values = []
        for col in range(result_table.columnCount()):
            item = result_table.item(current_row, col)
            values.append(item.text() if item else '')
        
        # Format as tab-separated
        result = '\t'.join(headers) + '\n' + '\t'.join(values)
        QApplication.clipboard().setText(result)
    
    def copy_multi_query_entire_table(self, tab_state):
        """Copy entire table with headers from multi-query result"""
        result_table = tab_state['result_table']
        
        # Get headers
        headers = []
        for col in range(result_table.columnCount()):
            header_item = result_table.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else f"Column_{col}")
        
        # Build table data
        lines = ['\t'.join(headers)]
        
        for row in range(result_table.rowCount()):
            row_data = []
            for col in range(result_table.columnCount()):
                item = result_table.item(row, col)
                row_data.append(item.text() if item else '')
            lines.append('\t'.join(row_data))
        
        QApplication.clipboard().setText('\n'.join(lines))
    
    def graph_multi_query_data(self, tab_state):
        """Open Eel dashboard with data from multi-query result"""
        if not EEL_AVAILABLE:
            QMessageBox.warning(self, 'Feature Unavailable', 
                              'Interactive Dashboard is not available. Please install Eel:\n'
                              'pip install eel')
            return
        
        if not PANDAS_AVAILABLE:
            QMessageBox.warning(self, 'Feature Unavailable', 
                              'Pandas is required. Please install:\n'
                              'pip install pandas')
            return
        
        try:
            result_table = tab_state['result_table']
            
            # Convert table data to pandas DataFrame
            df = self.table_to_dataframe(result_table)
            
            if df.empty:
                QMessageBox.information(self, 'No Data', 'No data available for dashboard.')
                return
            
            # Create and launch the Eel dashboard in a separate thread
            import threading
            
            def launch_dashboard():
                try:
                    create_dashboard(df, title=f"Query Results - Statement {tab_state['result_index']}")
                    while True:
                        eel.sleep(1.0)
                except Exception as e:
                    print(f"Error launching dashboard: {e}")
            
            dashboard_thread = threading.Thread(target=launch_dashboard, daemon=True)
            dashboard_thread.start()
            
            # Show info message
            QMessageBox.information(
                self, 
                'Dashboard Launched', 
                'Interactive Dashboard is opening as a desktop application.\n\n'
                'The dashboard window is independent of this application.'
            )
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to create dashboard: {str(e)}')
    
    def execute_streaming_query(self, tab_index):
        """Execute a streaming query for the specified tab"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        
        # Hide multi-query widget if it exists and show single-result widgets
        if 'multi_query_widget' in tab_data and tab_data['multi_query_widget']:
            tab_data['multi_query_widget'].setVisible(False)
        
        # Show the original single-result widgets
        tab_data['results_table'].setVisible(True)
        tab_data['progress_bar'].setVisible(True)
        tab_data['page_info_label'].setVisible(True)
        tab_data['first_page_btn'].setVisible(True)
        tab_data['prev_page_btn'].setVisible(True)
        tab_data['next_page_btn'].setVisible(True)
        tab_data['last_page_btn'].setVisible(True)
        tab_data['page_size_combo'].setVisible(True)
        tab_data['cancel_btn'].setVisible(True)
        
        # Find results group and update title
        results_group = tab_data['results_table'].parent()
        while results_group and not isinstance(results_group, QGroupBox):
            results_group = results_group.parent()
        if results_group:
            results_group.setTitle('Query Results')
        
        query = tab_data['current_query']
        page_size = int(tab_data['page_size_combo'].currentText())
        offset = tab_data['current_page'] * page_size
        
        # Update UI state
        self.status_label.setText('Executing query...')
        tab_data['cancel_btn'].setEnabled(True)
        tab_data['progress_bar'].setVisible(True)
        tab_data['progress_bar'].setValue(0)
        self.disable_pagination_controls(tab_index)
        
        # Stop any existing threads
        if tab_data['streaming_thread']:
            tab_data['streaming_thread'].cancel()
            tab_data['streaming_thread'].wait()
        if tab_data['query_thread']:
            tab_data['query_thread'].terminate()
            tab_data['query_thread'].wait()
        
        # Start streaming query
        streaming_thread = StreamingQueryThread(self.connection, query, page_size, offset)
        streaming_thread.batch_ready.connect(lambda cols, data, total, has_more: self.handle_batch_ready(tab_index, cols, data, total, has_more))
        streaming_thread.error_occurred.connect(lambda error: self.handle_streaming_error(tab_index, error))
        streaming_thread.progress_update.connect(lambda progress: self.handle_progress_update(tab_index, progress))
        
        tab_data['streaming_thread'] = streaming_thread
        streaming_thread.start()
    
    def handle_batch_ready(self, tab_index, columns, data, total_count, has_more):
        """Handle when a batch of results is ready"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        
        # Store metadata
        tab_data['columns'] = columns
        if total_count > 0:
            tab_data['total_rows'] = total_count
        
        # Display results
        self.display_results_for_tab(tab_index, columns, data)
        
        # Update pagination controls
        self.update_pagination_info(tab_index, has_more)
        
        # Re-enable controls
        tab_data['cancel_btn'].setEnabled(False)
        tab_data['progress_bar'].setVisible(False)
        self.enable_pagination_controls(tab_index)
        
        # Update status
        page_size = int(tab_data['page_size_combo'].currentText())
        start_row = tab_data['current_page'] * page_size + 1
        end_row = start_row + len(data) - 1
        if tab_data['total_rows'] > 0:
            self.status_label.setText(f'Showing rows {start_row}-{end_row} of {tab_data["total_rows"]:,} total rows')
        else:
            self.status_label.setText(f'Showing rows {start_row}-{end_row} (total unknown)')
    
    def handle_streaming_error(self, tab_index, error_message):
        """Handle streaming query errors"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        tab_data['cancel_btn'].setEnabled(False)
        tab_data['progress_bar'].setVisible(False)
        self.enable_pagination_controls(tab_index)
        
        QMessageBox.critical(self, 'Query Error', f'Query execution failed:\n{error_message}')
        self.status_label.setText('Query execution failed')
    
    def handle_progress_update(self, tab_index, progress):
        """Handle progress updates from streaming query"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        tab_data['progress_bar'].setValue(progress)
    
    def disable_pagination_controls(self, tab_index):
        """Disable pagination controls during query execution"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        tab_data['first_page_btn'].setEnabled(False)
        tab_data['prev_page_btn'].setEnabled(False)
        tab_data['next_page_btn'].setEnabled(False)
        tab_data['last_page_btn'].setEnabled(False)
        tab_data['page_size_combo'].setEnabled(False)
    
    def enable_pagination_controls(self, tab_index):
        """Enable pagination controls after query execution"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        tab_data['page_size_combo'].setEnabled(True)
        self.update_pagination_buttons(tab_index)
    
    def update_pagination_buttons(self, tab_index):
        """Update pagination button states"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        current_page = tab_data['current_page']
        total_rows = tab_data['total_rows']
        page_size = int(tab_data['page_size_combo'].currentText())
        
        # Enable/disable based on current page
        tab_data['first_page_btn'].setEnabled(current_page > 0)
        tab_data['prev_page_btn'].setEnabled(current_page > 0)
        
        # For next/last buttons, we need to check if there are more results
        if total_rows > 0:
            max_page = (total_rows - 1) // page_size
            tab_data['next_page_btn'].setEnabled(current_page < max_page)
            tab_data['last_page_btn'].setEnabled(current_page < max_page)
        else:
            # If we don't know total, enable next button and let the query determine if there are more results
            tab_data['next_page_btn'].setEnabled(True)
            tab_data['last_page_btn'].setEnabled(False)  # Can't go to last if we don't know total
    
    def update_pagination_info(self, tab_index, has_more):
        """Update pagination information display"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        current_page = tab_data['current_page']
        total_rows = tab_data['total_rows']
        page_size = int(tab_data['page_size_combo'].currentText())
        
        if total_rows > 0:
            total_pages = (total_rows - 1) // page_size + 1
            tab_data['page_info_label'].setText(f'Page {current_page + 1} of {total_pages} ({total_rows:,} total rows)')
        else:
            more_text = ' (more available)' if has_more else ''
            tab_data['page_info_label'].setText(f'Page {current_page + 1}{more_text}')
        
        self.update_pagination_buttons(tab_index)
    
    def go_to_page(self, tab_index, page_number):
        """Navigate to a specific page"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        if page_number < 0:
            page_number = 0
            
        tab_data['current_page'] = page_number
        self.execute_streaming_query(tab_index)
    
    def prev_page(self, tab_index):
        """Navigate to previous page"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        if tab_data['current_page'] > 0:
            self.go_to_page(tab_index, tab_data['current_page'] - 1)
    
    def next_page(self, tab_index):
        """Navigate to next page"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        self.go_to_page(tab_index, tab_data['current_page'] + 1)
    
    def go_to_last_page(self, tab_index):
        """Navigate to last page"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        if tab_data['total_rows'] > 0:
            page_size = int(tab_data['page_size_combo'].currentText())
            last_page = (tab_data['total_rows'] - 1) // page_size
            self.go_to_page(tab_index, last_page)
    
    def change_page_size(self, tab_index):
        """Handle page size change"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        if tab_data['current_query']:  # Only re-execute if we have a query
            tab_data['current_page'] = 0  # Reset to first page
            self.execute_streaming_query(tab_index)
    
    def cancel_query(self, tab_index):
        """Cancel the running query"""
        if tab_index not in self.query_tabs:
            return
            
        tab_data = self.query_tabs[tab_index]
        
        if tab_data['streaming_thread']:
            tab_data['streaming_thread'].cancel()
            tab_data['streaming_thread'].wait()
            tab_data['streaming_thread'] = None
            
        if tab_data['query_thread']:
            tab_data['query_thread'].terminate()
            tab_data['query_thread'].wait()
            tab_data['query_thread'] = None
            
        tab_data['cancel_btn'].setEnabled(False)
        tab_data['progress_bar'].setVisible(False)
        self.enable_pagination_controls(tab_index)
        self.status_label.setText('Query cancelled')
    
    def display_results_for_tab(self, tab_index, columns, data):
        """Display query results in the specific tab's table widget with optimized performance and memory usage"""
        if tab_index not in self.query_tabs:
            return
            
        results_table = self.query_tabs[tab_index]['results_table']
        
        # Clear existing data efficiently and force garbage collection
        results_table.clearContents()
        gc.collect()  # Free memory from previous data
        
        # Set up the table dimensions
        results_table.setRowCount(len(data))
        results_table.setColumnCount(len(columns))
        results_table.setHorizontalHeaderLabels(columns)
        
        # Optimize table performance for large datasets
        results_table.setUpdatesEnabled(False)  # Disable updates during population
        
        try:
            # Populate the table with batch processing
            batch_size = 1000  # Process in batches to avoid UI freezing
            for batch_start in range(0, len(data), batch_size):
                batch_end = min(batch_start + batch_size, len(data))
                
                for row_idx in range(batch_start, batch_end):
                    row_data = data[row_idx]
                    for col_idx, cell_data in enumerate(row_data):
                        # Convert to string efficiently and handle None values
                        if cell_data is None:
                            display_text = 'NULL'
                        elif isinstance(cell_data, (int, float)):
                            display_text = str(cell_data)
                        elif isinstance(cell_data, str):
                            # Truncate very long strings for display performance
                            display_text = cell_data[:1000] + '...' if len(cell_data) > 1000 else cell_data
                        else:
                            display_text = str(cell_data)[:1000]
                        
                        item = QTableWidgetItem(display_text)
                        # Set item flags for better performance (read-only)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        results_table.setItem(row_idx, col_idx, item)
                
                # Process events periodically to keep UI responsive
                if batch_end < len(data):
                    QApplication.processEvents()
        
        finally:
            results_table.setUpdatesEnabled(True)  # Re-enable updates
        
        # Optimize column sizing for performance
        if len(data) > 0:
            # For large datasets, use uniform column width instead of resizing to contents
            if len(data) > 10000:
                results_table.horizontalHeader().setDefaultSectionSize(120)
                results_table.horizontalHeader().setStretchLastSection(True)
            else:
                # Only resize columns to contents for smaller datasets
                results_table.resizeColumnsToContents()
                # Limit maximum column width for readability
                for col in range(results_table.columnCount()):
                    if results_table.columnWidth(col) > 300:
                        results_table.setColumnWidth(col, 300)
        
        # Disable sorting to prevent column header sorting arrows
        results_table.setSortingEnabled(False)
        
        # Set selection behavior for better performance
        results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        results_table.setAlternatingRowColors(False)  # Disable alternating row colors for consistent appearance
        
        # Enable export menu items when results are available
        if hasattr(self, 'export_excel_action'):
            self.export_excel_action.setEnabled(True)
        if hasattr(self, 'export_csv_action'):
            self.export_csv_action.setEnabled(True)
        if hasattr(self, 'export_json_action'):
            self.export_json_action.setEnabled(True)
        if hasattr(self, 'export_parquet_action'):
            self.export_parquet_action.setEnabled(True)
        
        # Clear the thread reference
        if tab_index in self.query_tabs:
            self.query_tabs[tab_index]['query_thread'] = None
    
    def handle_query_error_for_tab(self, tab_index, error_message):
        """Handle query execution errors for specific tab"""
        QMessageBox.critical(self, 'Query Error', f'Query execution failed:\n{error_message}')
        self.status_label.setText('Query execution failed')
        
        # Clear the thread reference
        if tab_index in self.query_tabs:
            self.query_tabs[tab_index]['query_thread'] = None
    
    def export_results(self, format_type, tab_index):
        """Export complete query results to Excel or CSV format"""
        if tab_index not in self.query_tabs:
            QMessageBox.warning(self, 'Export Error', 'No query results to export.')
            return
        
        tab_data = self.query_tabs[tab_index]
        
        # Check if we have a query to re-execute
        if 'current_query' not in tab_data or not tab_data['current_query']:
            QMessageBox.warning(self, 'Export Error', 'No query available for export. Please execute a query first.')
            return
        
        # Store format type and tab index for use in export callbacks
        self.export_format_type = format_type
        self.export_tab_index = tab_index
        
        # Create and show progress dialog
        from PyQt5.QtWidgets import QProgressDialog
        self.export_progress_dialog = QProgressDialog('Executing query for export...', 'Cancel', 0, 100, self)
        self.export_progress_dialog.setWindowModality(Qt.WindowModal)
        self.export_progress_dialog.setAutoClose(False)
        self.export_progress_dialog.setAutoReset(False)
        self.export_progress_dialog.show()
        
        # Create and start the full export query thread
        query = tab_data['current_query']
        self.export_query_thread = FullExportQueryThread(self.connection, query)
        self.export_query_thread.export_ready.connect(self.handle_export_data_ready)
        self.export_query_thread.error_occurred.connect(self.handle_export_error)
        self.export_query_thread.progress_update.connect(self.update_export_progress)
        
        # Connect cancel button to thread cancellation
        self.export_progress_dialog.canceled.connect(self.export_query_thread.cancel)
        
        self.export_query_thread.start()
    
    def update_export_progress(self, progress):
        """Update export progress dialog"""
        if hasattr(self, 'export_progress_dialog'):
            self.export_progress_dialog.setValue(progress)
    
    def handle_export_data_ready(self, columns, data):
        """Handle when export data is ready"""
        try:
            # Close progress dialog
            if hasattr(self, 'export_progress_dialog'):
                self.export_progress_dialog.close()
                delattr(self, 'export_progress_dialog')
            
            # Export the data
            if self.export_format_type == 'excel':
                self.export_to_excel(columns, data)
            elif self.export_format_type == 'csv':
                self.export_to_csv(columns, data)
            elif self.export_format_type == 'json':
                self.export_to_json(columns, data)
            elif self.export_format_type == 'parquet':
                self.export_to_parquet(columns, data)
                
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to export data:\n{str(e)}')
        finally:
            # Clean up
            if hasattr(self, 'export_query_thread'):
                self.export_query_thread.deleteLater()
                delattr(self, 'export_query_thread')
    
    def handle_export_error(self, error_message):
        """Handle export query errors"""
        # Close progress dialog
        if hasattr(self, 'export_progress_dialog'):
            self.export_progress_dialog.close()
            delattr(self, 'export_progress_dialog')
        
        QMessageBox.critical(self, 'Export Error', f'Failed to execute query for export:\n{error_message}')
        
        # Clean up
        if hasattr(self, 'export_query_thread'):
            self.export_query_thread.deleteLater()
            delattr(self, 'export_query_thread')
    
    def export_to_excel(self, columns, data):
        """Export data to Excel format with frozen headers"""
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(self, 'Export Error', 
                               'Excel export requires openpyxl library.\n'
                               'Please install it using: pip install openpyxl')
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Export to Excel', 'query_results.xlsx', 'Excel Files (*.xlsx)'
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Query Results'
            
            # Add headers with bold formatting
            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.font = Font(bold=True)
            
            # Add data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Freeze the top row (headers)
            ws.freeze_panes = 'A2'
            
            # Auto-adjust column widths
            for col_idx in range(1, len(columns) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(columns[col_idx - 1])
                for row_idx in range(2, len(data) + 2):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                    max_length = max(max_length, len(cell_value))
                # Set column width with some padding, but cap at reasonable maximum
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            wb.save(file_path)
            QMessageBox.information(self, 'Export Successful', 
                                  f'Data exported successfully to:\n{file_path}')
            
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to export to Excel:\n{str(e)}')
    
    def export_to_csv(self, columns, data):
        """Export data to CSV format with delimiter selection"""
        # Show delimiter selection dialog
        delimiter_dialog = ExportDelimiterDialog(self)
        if delimiter_dialog.exec_() != QDialog.Accepted:
            return
        
        delimiter = delimiter_dialog.get_delimiter()
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Export to CSV', 'query_results.csv', 'CSV Files (*.csv)'
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=delimiter)
                
                # Write headers
                writer.writerow(columns)
                
                # Write data
                writer.writerows(data)
            
            QMessageBox.information(self, 'Export Successful', 
                                  f'Data exported successfully to:\n{file_path}\n'
                                  f'Delimiter used: {repr(delimiter)}')
            
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to export to CSV:\n{str(e)}')
    
    def export_to_json(self, columns, data):
        """Export data to JSON format"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Export to JSON', 'query_results.json', 'JSON Files (*.json)'
        )
        
        if not file_path:
            return
        
        try:
            # Convert data to list of dictionaries
            json_data = []
            for row in data:
                row_dict = {}
                for i, column in enumerate(columns):
                    # Handle different data types for JSON serialization
                    value = row[i]
                    if value is None:
                        row_dict[column] = None
                    elif isinstance(value, (int, float, str, bool)):
                        row_dict[column] = value
                    else:
                        # Convert other types to string
                        row_dict[column] = str(value)
                json_data.append(row_dict)
            
            # Write JSON file
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(json_data, jsonfile, indent=2, ensure_ascii=False)
            
            QMessageBox.information(self, 'Export Successful', 
                                  f'Data exported successfully to:\n{file_path}\n'
                                  f'Records exported: {len(json_data)}')
            
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to export to JSON:\n{str(e)}')
    
    def export_to_parquet(self, columns, data):
        """Export data to Parquet format using Polars"""
        if not PARQUET_AVAILABLE:
            QMessageBox.critical(self, 'Export Error', 
                               'Parquet export requires pyarrow library.\n'
                               'Please install it using: pip install pyarrow')
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Export to Parquet', 'query_results.parquet', 'Parquet Files (*.parquet)'
        )
        
        if not file_path:
            return
        
        try:
            # Convert data to dictionary format for Polars DataFrame
            data_dict = {}
            for i, column in enumerate(columns):
                data_dict[column] = [row[i] for row in data]
            
            # Create Polars DataFrame and write to Parquet
            df = pl.DataFrame(data_dict)
            df.write_parquet(file_path)
            
            QMessageBox.information(self, 'Export Successful', 
                                  f'Data exported successfully to:\n{file_path}\n'
                                  f'Records exported: {len(data)}\n'
                                  f'Columns: {len(columns)}')
            
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to export to Parquet:\n{str(e)}')

    def toggle_split_screen(self):
        """Toggle split screen mode on/off"""
        if not self.split_screen_active:
            self.enable_split_screen()
        else:
            self.disable_split_screen()
    
    def enable_split_screen(self):
        """Enable split screen mode"""
        if self.split_screen_active:
            return
            
        # Get current tab widget
        current_tab_widget = self.query_tab_widget
        current_parent = current_tab_widget.parent()
        
        # Create horizontal splitter for split screen
        self.split_screen_widget = QSplitter(Qt.Horizontal)
        
        # Configure splitter for smooth dynamic resizing
        self.split_screen_widget.setChildrenCollapsible(False)  # Prevent panels from collapsing
        self.split_screen_widget.setHandleWidth(8)  # Set handle width for easier dragging
        self.split_screen_widget.setOpaqueResize(True)  # Enable smooth real-time resizing
        
        # Set splitter handle style for better visibility and smooth dragging
        self.split_screen_widget.setStyleSheet("""
            QSplitter::handle:horizontal {
                background-color: #d0d0d0;
                border: 1px solid #a0a0a0;
                border-radius: 3px;
                margin: 2px 0px;
                width: 8px;
            }
            QSplitter::handle:horizontal:hover {
                background-color: #b0b0b0;
            }
            QSplitter::handle:horizontal:pressed {
                background-color: #909090;
            }
        """)
        
        # Remove current tab widget from its parent
        current_tab_widget.setParent(None)
        
        # Create left side container
        left_container = QWidget()
        left_container.setMinimumWidth(200)  # Set minimum width to prevent collapse
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins for better space usage
        left_layout.addWidget(current_tab_widget)
        
        # Create right side with duplicate tab widget
        right_container = QWidget()
        right_container.setMinimumWidth(200)  # Set minimum width to prevent collapse
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins for better space usage
        
        # Create new tab widget for right side
        self.right_query_tab_widget = QTabWidget()
        self.right_query_tab_widget.setTabsClosable(False)
        
        # Add initial query tab to right side
        self.add_new_query_tab_to_widget(self.right_query_tab_widget)
        
        right_layout.addWidget(self.right_query_tab_widget)
        
        # Add both containers to splitter
        self.split_screen_widget.addWidget(left_container)
        self.split_screen_widget.addWidget(right_container)
        
        # Set equal proportions (50/50 split) and ensure proper sizing
        total_width = self.width() if hasattr(self, 'width') else 1000
        half_width = max(200, total_width // 2)  # Ensure minimum width
        self.split_screen_widget.setSizes([half_width, half_width])
        self.split_screen_widget.setStretchFactor(0, 1)  # Left panel can stretch
        self.split_screen_widget.setStretchFactor(1, 1)  # Right panel can stretch
        
        # Ensure splitter is properly enabled and interactive
        self.split_screen_widget.setEnabled(True)
        self.split_screen_widget.show()  # Explicitly show the splitter
        
        # Replace the original tab widget in the parent layout
        parent_layout = current_parent.layout()
        parent_layout.addWidget(self.split_screen_widget)
        
        self.split_screen_active = True
        
    def disable_split_screen(self):
        """Disable split screen mode"""
        if not self.split_screen_active:
            return
            
        # Get the left tab widget (original)
        left_container = self.split_screen_widget.widget(0)
        left_layout = left_container.layout()
        original_tab_widget = left_layout.itemAt(0).widget()
        
        # Remove from split screen
        original_tab_widget.setParent(None)
        
        # Get parent and add back the original tab widget
        parent_widget = self.split_screen_widget.parent()
        parent_layout = parent_widget.layout()
        
        # Remove split screen widget
        self.split_screen_widget.setParent(None)
        
        # Add original tab widget back
        parent_layout.addWidget(original_tab_widget)
        
        # Clean up
        self.right_query_tab_widget = None
        self.split_screen_widget = None
        self.split_screen_active = False
    
    def add_new_query_tab_to_widget(self, tab_widget):
        """Add a new query tab to a specific tab widget (for split screen)"""
        self.tab_counter += 1
        tab_name = f"Query {self.tab_counter}"
        
        # Create tab widget
        tab_widget_content = QWidget()
        tab_layout = QVBoxLayout(tab_widget_content)
        
        # Create splitter for SQL editor and results
        splitter = QSplitter(Qt.Vertical)
        
        # SQL Editor section
        sql_group = QGroupBox('SQL Query Editor')
        sql_layout = QVBoxLayout(sql_group)
        
        sql_editor = SQLTextEdit(self)
        sql_editor.setFont(QFont('Consolas', 10))
        sql_editor.setPlaceholderText('Enter your SQL query here...\nExample: SELECT * FROM localdb.your_table_name LIMIT 10;')
        
        # Update autocomplete with current table names
        self.update_autocomplete_for_editor(sql_editor)
        
        # Button layout for Execute and New Query buttons
        button_layout = QHBoxLayout()
        execute_btn = QPushButton('Execute Query')
        new_query_btn = QPushButton('New Query')
        split_screen_btn = QPushButton('Split Screen')
        
        # Export buttons
        export_excel_btn = QPushButton('Export to Excel')
        export_csv_btn = QPushButton('Export to CSV')
        export_json_btn = QPushButton('Export to JSON')
        export_parquet_btn = QPushButton('Export to Parquet')

        
        # Connect buttons - need to determine which tab widget this belongs to
        is_right_side = (tab_widget == getattr(self, 'right_query_tab_widget', None))
        
        execute_btn.clicked.connect(lambda: self.execute_query_for_tab_widget(tab_widget, tab_widget.currentIndex()))
        new_query_btn.clicked.connect(lambda: self.add_new_query_tab_to_widget(tab_widget))
        split_screen_btn.clicked.connect(self.toggle_split_screen)
        export_excel_btn.clicked.connect(lambda: self.export_results_for_tab_widget(tab_widget, 'excel', tab_widget.currentIndex()))
        export_csv_btn.clicked.connect(lambda: self.export_results_for_tab_widget(tab_widget, 'csv', tab_widget.currentIndex()))
        export_json_btn.clicked.connect(lambda: self.export_results_for_tab_widget(tab_widget, 'json', tab_widget.currentIndex()))
        export_parquet_btn.clicked.connect(lambda: self.export_results_for_tab_widget(tab_widget, 'parquet', tab_widget.currentIndex()))

        
        button_layout.addWidget(execute_btn)
        button_layout.addWidget(new_query_btn)
        button_layout.addWidget(split_screen_btn)
        button_layout.addWidget(export_excel_btn)
        button_layout.addWidget(export_csv_btn)
        button_layout.addWidget(export_json_btn)
        button_layout.addWidget(export_parquet_btn)

        button_layout.addStretch()
        
        sql_layout.addWidget(sql_editor)
        sql_layout.addLayout(button_layout)
        
        splitter.addWidget(sql_group)
        
        # Results section
        results_group = QGroupBox('Query Results')
        results_layout = QVBoxLayout(results_group)
        
        # Pagination controls
        pagination_layout = QHBoxLayout()
        
        # Page info and controls
        page_info_label = QLabel('No results')
        first_page_btn = QPushButton('First')
        prev_page_btn = QPushButton('Previous')
        next_page_btn = QPushButton('Next')
        last_page_btn = QPushButton('Last')
        
        # Page size selector
        page_size_label = QLabel('Rows per page:')
        page_size_combo = QComboBox()
        page_size_combo.addItems(['1000', '5000', '10000', '25000', '50000'])
        page_size_combo.setCurrentText('10000')
        
        # Cancel query button
        cancel_btn = QPushButton('Cancel Query')
        cancel_btn.setEnabled(False)
        
        # Progress bar
        progress_bar = QProgressBar()
        progress_bar.setVisible(False)
        progress_bar.setMaximum(100)
        
        pagination_layout.addWidget(page_info_label)
        pagination_layout.addStretch()
        pagination_layout.addWidget(first_page_btn)
        pagination_layout.addWidget(prev_page_btn)
        pagination_layout.addWidget(next_page_btn)
        pagination_layout.addWidget(last_page_btn)
        pagination_layout.addStretch()
        pagination_layout.addWidget(page_size_label)
        pagination_layout.addWidget(page_size_combo)
        pagination_layout.addWidget(cancel_btn)
        
        # Add progress bar below pagination controls
        results_layout.addWidget(progress_bar)
        
        results_table = QTableWidget()
        
        results_layout.addLayout(pagination_layout)
        results_layout.addWidget(results_table)
        
        splitter.addWidget(results_group)
        
        # Set splitter proportions
        splitter.setSizes([300, 500])
        
        tab_layout.addWidget(splitter)
        
        # Add tab to tab widget
        tab_index = tab_widget.addTab(tab_widget_content, tab_name)
        
        # Set up context menu for results table (after tab_index is defined)
        results_table.setContextMenuPolicy(Qt.CustomContextMenu)
        results_table.customContextMenuRequested.connect(lambda pos, tw=tab_widget, tab_idx=tab_index: self.show_results_context_menu_for_widget(pos, tw, tab_idx))
        
        # Set up context menu for table headers
        results_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        results_table.horizontalHeader().customContextMenuRequested.connect(lambda pos, tw=tab_widget, tab_idx=tab_index: self.show_header_context_menu_for_widget(pos, tw, tab_idx))
        
        # Create custom close button for this tab
        close_button = QPushButton('×')
        close_button.setFixedSize(16, 16)
        close_button.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                border: none;
                color: {self.get_current_theme_color('text')};
                font-weight: bold;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: red;
                color: white;
                border-radius: 8px;
            }}
        """)
        close_button.setToolTip('Close tab')
        close_button.clicked.connect(lambda: self.close_query_tab_for_widget(tab_widget, tab_index))
        
        # Add close button to tab
        tab_widget.tabBar().setTabButton(tab_index, tab_widget.tabBar().RightSide, close_button)
        
        # Connect pagination controls
        first_page_btn.clicked.connect(lambda: self.go_to_page_for_widget(tab_widget, tab_index, 0))
        prev_page_btn.clicked.connect(lambda: self.prev_page_for_widget(tab_widget, tab_index))
        next_page_btn.clicked.connect(lambda: self.next_page_for_widget(tab_widget, tab_index))
        last_page_btn.clicked.connect(lambda: self.go_to_last_page_for_widget(tab_widget, tab_index))
        page_size_combo.currentTextChanged.connect(lambda: self.change_page_size_for_widget(tab_widget, tab_index))
        cancel_btn.clicked.connect(lambda: self.cancel_query_for_widget(tab_widget, tab_index))
        
        # Store tab components - use unique key for split screen tabs
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if not hasattr(self, 'split_query_tabs'):
            self.split_query_tabs = {}
            
        self.split_query_tabs[tab_key] = {
            'sql_editor': sql_editor,
            'results_table': results_table,
            'query_thread': None,
            'streaming_thread': None,
            'close_button': close_button,
            'page_info_label': page_info_label,
            'first_page_btn': first_page_btn,
            'prev_page_btn': prev_page_btn,
            'next_page_btn': next_page_btn,
            'last_page_btn': last_page_btn,
            'page_size_combo': page_size_combo,
            'cancel_btn': cancel_btn,
            'progress_bar': progress_bar,
            'current_page': 0,
            'total_rows': 0,
            'current_query': '',
            'columns': [],
            'tab_widget': tab_widget
        }
        
        # Switch to new tab
        tab_widget.setCurrentIndex(tab_index)
        
        return tab_index

    # Split screen functionality methods
    def execute_query_for_tab_widget(self, tab_widget, tab_index):
        """Execute query for a specific tab widget"""
        if tab_widget == self.query_tab_widget:
            # Use existing method for left side
            self.execute_query_for_tab(tab_index)
        else:
            # Handle right side tab widget
            self.execute_query_for_split_tab(tab_widget, tab_index)
    
    def execute_query_for_split_tab(self, tab_widget, tab_index):
        """Execute query for split screen tab"""
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        sql_editor = tab_data['sql_editor']
        results_table = tab_data['results_table']
        cancel_btn = tab_data['cancel_btn']
        progress_bar = tab_data['progress_bar']
        
        query = sql_editor.toPlainText().strip()
        if not query:
            QMessageBox.warning(self, 'Warning', 'Please enter a SQL query.')
            return
        
        # Cancel any existing query
        if tab_data['query_thread'] and tab_data['query_thread'].isRunning():
            tab_data['query_thread'].terminate()
            tab_data['query_thread'].wait()
        
        # Clear previous results
        results_table.clear()
        results_table.setRowCount(0)
        results_table.setColumnCount(0)
        
        # Show progress
        progress_bar.setVisible(True)
        progress_bar.setValue(0)
        cancel_btn.setEnabled(True)
        
        # Store current query
        tab_data['current_query'] = query
        tab_data['current_page'] = 0
        
        # Create and start query thread
        tab_data['query_thread'] = SQLQueryThread(self.connection, query)
        tab_data['query_thread'].result_ready.connect(lambda cols, data: self.handle_split_query_result(tab_key, cols, data))
        tab_data['query_thread'].error_occurred.connect(lambda error: self.handle_split_query_error(tab_key, error))
        tab_data['query_thread'].start()
    
    def handle_split_query_result(self, tab_key, columns, data):
        """Handle query result for split screen tab"""
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        results_table = tab_data['results_table']
        progress_bar = tab_data['progress_bar']
        cancel_btn = tab_data['cancel_btn']
        page_info_label = tab_data['page_info_label']
        
        try:
            # Store data
            tab_data['columns'] = columns
            tab_data['data'] = data
            tab_data['total_rows'] = len(data)
            
            # Update table
            self.update_split_results_table(tab_key, columns, data)
            
            # Update pagination info
            total_rows = len(data)
            page_info_label.setText(f'{total_rows:,} rows returned')
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to display results: {str(e)}')
        finally:
            progress_bar.setVisible(False)
            cancel_btn.setEnabled(False)
    
    def handle_split_query_error(self, tab_key, error):
        """Handle query error for split screen tab"""
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        progress_bar = tab_data['progress_bar']
        cancel_btn = tab_data['cancel_btn']
        
        progress_bar.setVisible(False)
        cancel_btn.setEnabled(False)
        QMessageBox.critical(self, 'Query Error', str(error))
    
    def update_split_results_table(self, tab_key, columns, data):
        """Update results table for split screen tab"""
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        results_table = tab_data['results_table']
        
        if not data or len(data) == 0:
            results_table.setRowCount(0)
            results_table.setColumnCount(0)
            return
        
        # Set up table
        results_table.setRowCount(len(data))
        results_table.setColumnCount(len(columns))
        results_table.setHorizontalHeaderLabels([str(col) for col in columns])
        
        # Populate table
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                if value is None:
                    item_text = 'NULL'
                else:
                    item_text = str(value)
                
                item = QTableWidgetItem(item_text)
                results_table.setItem(row_idx, col_idx, item)
        
        # Auto-resize columns
        results_table.resizeColumnsToContents()
        
        # Set selection behavior
        results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
    
    def update_split_pagination_buttons(self, tab_key):
        """Update pagination buttons for split screen tab"""
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        page_size = int(tab_data['page_size_combo'].currentText())
        total_pages = (tab_data['total_rows'] + page_size - 1) // page_size
        current_page = tab_data['current_page']
        
        tab_data['first_page_btn'].setEnabled(current_page > 0)
        tab_data['prev_page_btn'].setEnabled(current_page > 0)
        tab_data['next_page_btn'].setEnabled(current_page < total_pages - 1)
        tab_data['last_page_btn'].setEnabled(current_page < total_pages - 1)
    
    def export_results_for_tab_widget(self, tab_widget, format_type, tab_index):
        """Export results for a specific tab widget"""
        if tab_widget == self.query_tab_widget:
            self.export_results(format_type, tab_index)
        else:
            # Handle right side tab widget export
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                self.export_split_results(tab_key, format_type)
    
    def export_split_results(self, tab_key, format_type):
        """Export results for split screen tab"""
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        query = tab_data['current_query']
        
        if not query:
            QMessageBox.warning(self, 'Warning', 'No query results to export.')
            return
        
        # Use existing export logic but with split screen query
        try:
            # Execute full query for export
            df = self.conn.execute(query).pl()
            
            if format_type == 'excel':
                self.export_to_excel_split(df)
            elif format_type == 'csv':
                self.export_to_csv_split(df)
            elif format_type == 'json':
                self.export_to_json_split(df)
            elif format_type == 'parquet':
                self.export_to_parquet_split(df)
                
        except Exception as e:
            QMessageBox.critical(self, 'Export Error', f'Failed to export: {str(e)}')
    
    def export_to_excel_split(self, df):
        """Export split screen results to Excel"""
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(self, 'Warning', 'Excel export requires openpyxl package.')
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, 'Export to Excel', '', 'Excel Files (*.xlsx)')
        if file_path:
            try:
                df.write_excel(file_path)
                QMessageBox.information(self, 'Success', f'Data exported to {file_path}')
            except Exception as e:
                QMessageBox.critical(self, 'Export Error', f'Failed to export to Excel:\n{str(e)}')
    
    def export_to_csv_split(self, df):
        """Export split screen results to CSV"""
        file_path, _ = QFileDialog.getSaveFileName(self, 'Export to CSV', '', 'CSV Files (*.csv)')
        if file_path:
            try:
                df.write_csv(file_path)
                QMessageBox.information(self, 'Success', f'Data exported to {file_path}')
            except Exception as e:
                QMessageBox.critical(self, 'Export Error', f'Failed to export to CSV:\n{str(e)}')
    
    def export_to_json_split(self, df):
        """Export split screen results to JSON"""
        file_path, _ = QFileDialog.getSaveFileName(self, 'Export to JSON', '', 'JSON Files (*.json)')
        if file_path:
            try:
                df.write_json(file_path)
                QMessageBox.information(self, 'Success', f'Data exported to {file_path}')
            except Exception as e:
                QMessageBox.critical(self, 'Export Error', f'Failed to export to JSON:\n{str(e)}')
    
    def export_to_parquet_split(self, df):
        """Export split screen results to Parquet"""
        if not PARQUET_AVAILABLE:
            QMessageBox.warning(self, 'Warning', 'Parquet export requires pyarrow package.')
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, 'Export to Parquet', '', 'Parquet Files (*.parquet)')
        if file_path:
            try:
                df.write_parquet(file_path)
                QMessageBox.information(self, 'Success', f'Data exported to {file_path}')
            except Exception as e:
                QMessageBox.critical(self, 'Export Error', f'Failed to export to Parquet:\n{str(e)}')
    
    def show_results_context_menu_for_widget(self, pos, tab_widget, tab_index):
        """Show context menu for results table in specific tab widget"""
        if tab_widget == self.query_tab_widget:
            self.show_results_context_menu(pos, tab_index)
        else:
            # Handle right side context menu for split screen
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key not in self.split_query_tabs:
                return
                
            tab_data = self.split_query_tabs[tab_key]
            results_table = tab_data['results_table']
            
            # Get the item at the clicked position
            item = results_table.itemAt(pos)
            if not item:
                return
                
            row = item.row()
            column = item.column()
            
            # Create context menu
            context_menu = QMenu(self)
            
            # Copy Cell Value action
            copy_cell_action = QAction('Copy Cell Value', self)
            copy_cell_action.triggered.connect(lambda: self.copy_cell_value_for_split(tab_widget, tab_index, row, column))
            context_menu.addAction(copy_cell_action)
            
            # Copy Column with Header action
            copy_column_action = QAction('Copy Column with Header', self)
            copy_column_action.triggered.connect(lambda: self.copy_column_with_header_for_split(tab_widget, tab_index, column))
            context_menu.addAction(copy_column_action)
            
            # Copy Row with Header action
            copy_row_action = QAction('Copy Row with Header', self)
            copy_row_action.triggered.connect(lambda: self.copy_row_with_header_for_split(tab_widget, tab_index, row))
            context_menu.addAction(copy_row_action)
            
            # Copy Entire Table action
            copy_table_action = QAction('Copy Entire Table', self)
            copy_table_action.triggered.connect(lambda: self.copy_entire_table_for_split(tab_widget, tab_index))
            context_menu.addAction(copy_table_action)
            
            # Add separator and Graph Data option if Plotly is available
            if PLOTLY_AVAILABLE:
                context_menu.addSeparator()
                graph_data_action = QAction('Graph Data', self)
                graph_data_action.triggered.connect(lambda: self.graph_data_for_split(tab_widget, tab_index))
                context_menu.addAction(graph_data_action)
            
            # Show the context menu
            context_menu.exec_(results_table.mapToGlobal(pos))
    
    def show_header_context_menu_for_widget(self, pos, tab_widget, tab_index):
        """Show context menu for table headers in specific tab widget"""
        if tab_widget == self.query_tab_widget:
            self.show_header_context_menu(pos, tab_index)
        else:
            # Handle right side header context menu for split screen
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key not in self.split_query_tabs:
                return
                
            tab_data = self.split_query_tabs[tab_key]
            results_table = tab_data['results_table']
            header = results_table.horizontalHeader()
            
            # Get the column index from the position
            column = header.logicalIndexAt(pos)
            if column < 0:
                return
                
            menu = QMenu(self)
            
            # Copy header value
            copy_header_action = QAction('Copy Header', self)
            copy_header_action.triggered.connect(lambda: self.copy_header_value_for_split(tab_widget, tab_index, column))
            menu.addAction(copy_header_action)
            
            menu.exec_(header.mapToGlobal(pos))
    
    def close_query_tab_for_widget(self, tab_widget, tab_index):
        """Close a query tab in specific tab widget"""
        if tab_widget == self.query_tab_widget:
            self.close_query_tab(tab_index)
        else:
            # Handle right side tab closing
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                # Cancel any running query
                tab_data = self.split_query_tabs[tab_key]
                if tab_data['query_thread'] and tab_data['query_thread'].isRunning():
                    tab_data['query_thread'].terminate()
                    tab_data['query_thread'].wait()
                
                # Remove from tracking
                del self.split_query_tabs[tab_key]
            
            # Remove tab
            tab_widget.removeTab(tab_index)
    
    def go_to_page_for_widget(self, tab_widget, tab_index, page):
        """Go to specific page in tab widget"""
        if tab_widget == self.query_tab_widget:
            self.go_to_page(tab_index, page)
        else:
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                self.go_to_split_page(tab_key, page)
    
    def prev_page_for_widget(self, tab_widget, tab_index):
        """Go to previous page in tab widget"""
        if tab_widget == self.query_tab_widget:
            self.prev_page(tab_index)
        else:
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                current_page = self.split_query_tabs[tab_key]['current_page']
                if current_page > 0:
                    self.go_to_split_page(tab_key, current_page - 1)
    
    def next_page_for_widget(self, tab_widget, tab_index):
        """Go to next page in tab widget"""
        if tab_widget == self.query_tab_widget:
            self.next_page(tab_index)
        else:
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                tab_data = self.split_query_tabs[tab_key]
                page_size = int(tab_data['page_size_combo'].currentText())
                total_pages = (tab_data['total_rows'] + page_size - 1) // page_size
                current_page = tab_data['current_page']
                if current_page < total_pages - 1:
                    self.go_to_split_page(tab_key, current_page + 1)
    
    def go_to_last_page_for_widget(self, tab_widget, tab_index):
        """Go to last page in tab widget"""
        if tab_widget == self.query_tab_widget:
            self.go_to_last_page(tab_index)
        else:
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                tab_data = self.split_query_tabs[tab_key]
                page_size = int(tab_data['page_size_combo'].currentText())
                total_pages = (tab_data['total_rows'] + page_size - 1) // page_size
                self.go_to_split_page(tab_key, total_pages - 1)
    
    def go_to_split_page(self, tab_key, page):
        """Go to specific page for split screen tab"""
        if tab_key not in self.split_query_tabs:
            return
            
        tab_data = self.split_query_tabs[tab_key]
        page_size = int(tab_data['page_size_combo'].currentText())
        offset = page * page_size
        
        query = tab_data['current_query']
        paginated_query = f"{query} LIMIT {page_size} OFFSET {offset}"
        
        try:
            df = self.conn.execute(paginated_query).pl()
            tab_data['current_page'] = page
            self.update_split_results_table(tab_key, df)
            
            # Update page info
            total_pages = (tab_data['total_rows'] + page_size - 1) // page_size
            current_page = page + 1
            tab_data['page_info_label'].setText(f'Page {current_page} of {total_pages} ({tab_data["total_rows"]:,} total rows)')
            
            # Update buttons
            self.update_split_pagination_buttons(tab_key)
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load page: {str(e)}')
    
    def change_page_size_for_widget(self, tab_widget, tab_index):
        """Change page size in tab widget"""
        if tab_widget == self.query_tab_widget:
            self.change_page_size(tab_index)
        else:
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                # Reset to first page with new page size
                self.go_to_split_page(tab_key, 0)
    
    def cancel_query_for_widget(self, tab_widget, tab_index):
        """Cancel query in tab widget"""
        if tab_widget == self.query_tab_widget:
            self.cancel_query(tab_index)
        else:
            tab_key = f"{id(tab_widget)}_{tab_index}"
            if tab_key in self.split_query_tabs:
                self.cancel_split_query(tab_key)
    
    def open_eel_dashboard(self, tab_index):
        """Open Eel interactive dashboard with data from the specified tab"""
        if not EEL_AVAILABLE:
            QMessageBox.warning(self, 'Feature Unavailable', 
                              'Interactive Dashboard is not available. Please install Eel:\n'
                              'pip install eel')
            return
        
        if not PANDAS_AVAILABLE:
            QMessageBox.warning(self, 'Feature Unavailable', 
                              'Pandas is required for dashboard. Please install:\n'
                              'pip install pandas')
            return
            
        if tab_index not in self.query_tabs:
            return
            
        try:
            # Get the results table
            results_table = self.query_tabs[tab_index]['results_table']
            
            # Convert table data to pandas DataFrame
            df = self.table_to_dataframe(results_table)
            
            if df.empty:
                QMessageBox.information(self, 'No Data', 'No data available for dashboard.')
                return
            
            # Create and launch the Eel dashboard in a separate thread
            import threading
            
            def launch_dashboard():
                try:
                    create_dashboard(df, title=f"Query Results - Tab {tab_index + 1}")
                    # Keep the dashboard running
                    while True:
                        eel.sleep(1.0)
                except Exception as e:
                    print(f"Error launching dashboard: {e}")
            
            dashboard_thread = threading.Thread(target=launch_dashboard, daemon=True)
            dashboard_thread.start()
            
            # Small delay to let dashboard start
            import time
            time.sleep(0.5)
            
            # Show info message
            QMessageBox.information(
                self, 
                'Dashboard Launched', 
                'Interactive Dashboard is opening as a desktop application.\n\n'
                'Features:\n'
                '• Drag and drop to create charts\n'
                '• Multiple chart types\n'
                '• Filters and interactivity\n'
                '• Export as standalone HTML\n\n'
                'The dashboard window is independent of this application.'
            )
            
        except Exception as e:
            # Only show error if dashboard didn't start
            if 'create_dashboard' not in str(e):
                QMessageBox.critical(self, 'Error', f'Failed to create dashboard: {str(e)}')
    
    def graph_data_for_split(self, tab_widget, tab_index):
        """Open Eel dashboard with data from the specified split tab"""
        if not EEL_AVAILABLE:
            QMessageBox.warning(self, 'Feature Unavailable', 
                              'Interactive Dashboard is not available. Please install Eel:\n'
                              'pip install eel')
            return
        
        if not PANDAS_AVAILABLE:
            QMessageBox.warning(self, 'Feature Unavailable', 
                              'Pandas is required. Please install:\n'
                              'pip install pandas')
            return
            
        tab_key = f"{id(tab_widget)}_{tab_index}"
        if tab_key not in self.split_query_tabs:
            return
            
        try:
            # Get the results table
            results_table = self.split_query_tabs[tab_key]['results_table']
            
            # Convert table data to pandas DataFrame
            df = self.table_to_dataframe(results_table)
            
            if df.empty:
                QMessageBox.information(self, 'No Data', 'No data available for dashboard.')
                return
                
            # Create and launch the Eel dashboard in a separate thread
            import threading
            
            def launch_dashboard():
                try:
                    create_dashboard(df, title=f"Split Query Results - Tab {tab_index + 1}")
                    while True:
                        eel.sleep(1.0)
                except Exception as e:
                    print(f"Error launching dashboard: {e}")
            
            dashboard_thread = threading.Thread(target=launch_dashboard, daemon=True)
            dashboard_thread.start()
            
            # Show info message
            QMessageBox.information(
                self, 
                'Dashboard Launched', 
                'Interactive Dashboard is opening as a desktop application.\n\n'
                'The dashboard window is independent of this application.'
            )
            
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to create dashboard: {str(e)}')
    
    def table_to_dataframe(self, table_widget):
        """Convert QTableWidget data to pandas DataFrame"""
        if not table_widget or table_widget.rowCount() == 0:
            return pd.DataFrame()
            
        # Get column headers
        columns = []
        for col in range(table_widget.columnCount()):
            header_item = table_widget.horizontalHeaderItem(col)
            columns.append(header_item.text() if header_item else f"Column_{col}")
        
        # Get data
        data = []
        for row in range(table_widget.rowCount()):
            row_data = []
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                value = item.text() if item else ""
                
                # Try to convert to numeric if possible
                try:
                    if '.' in value:
                        value = float(value)
                    elif value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                        value = int(value)
                except (ValueError, AttributeError):
                    pass  # Keep as string
                    
                row_data.append(value)
            data.append(row_data)
        
        return pd.DataFrame(data, columns=columns)

def main():
    # Set Qt attribute for WebEngine before creating QApplication
    QApplication.setAttribute(Qt.AA_ShareOpenGLContexts)
    app = QApplication(sys.argv)
    window = DuckDBSQLApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()